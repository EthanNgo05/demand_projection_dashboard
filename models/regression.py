"""
Updates projections for all SKUs based on combined demand projections for all companies.

Uses the last 8 *completed* weeks of historical POS data to calculate a 15-week
projection. The in-progress week is excluded so a partially-elapsed week's POS
never drags the average (see week_anchors).

    projected_pos(week_k) = avg_pos_over_8_weeks + adjustment(k)
        adjustment(k) = slope * TREND_WEIGHT * k     (k = 1 .. 15 weeks out)

    - slope        : slope from a linear regression of the last 8 weeks of POS
    - TREND_WEIGHT : 0.25 (dampens the trend so the forecast doesn't over-react)
                        0 -> pure 8-week average, 1 -> pure trend
    - updated_projection_avg (per SKU) = mean of the 15 weekly projected_pos values
                       (equivalently: avg + slope * TREND_WEIGHT * 6.5,
                        since mean(1..15) = 8.0)

Two kinds of output are produced:
    1) Per-group     : one file per Customer Grouping (see COMBINED_GROUPING),
                       summing POS across the customers that make up the group
                       (e.g. the three Amazon-DC channels become one AMAZON-DC
                       forecast). Customers not in the mapping become their own
                       single-member group.
    2) All-customers : a single file forecasting each SKU "as a whole", built by
                       summing POS (and the existing Projection) across every
                       customer for each SKU-week before fitting.
"""

import os
import re
import glob
import traceback

import numpy as np
import pandas as pd

# How much weight to give the trend direction (0 = pure average, 1 = pure trend)
TREND_WEIGHT = 0.25

# Header caption shown by the dashboard when this model is selected.
DASHBOARD_CAPTION = (
    "15-week simple-regression forecast: anchored to the 8-week average of the "
    "historical demand window (POS where available, else Orders), nudged by a "
    f"dampened linear-regression slope (trend weight = {TREND_WEIGHT})."
)

# RAW_INPUTS_FOLDER is the constant the dashboard reads to discover raw files,
# matching the exponential-smoothing / XGBoost pipelines so switching models
# never changes which folder is scanned.
RAW_INPUTS_FOLDER = "raw_inputs/demand_projections"
INPUT_GLOB = os.path.join(RAW_INPUTS_FOLDER, "all_demand_projections_*.xlsx")

# List-price workbook (SKU -> List Price USD), used to value the change in
# forecast as a revenue risk. See load_list_prices(). Matches the other
# pipelines so switching models never changes which folder is scanned.
LIST_PRICE_GLOB = os.path.join("raw_inputs/list_prices", "list_prices_*.xlsx")

CUSTOMERS_TO_IGNORE = [
    "Others - HK", "Others - KR", "Others - MALDIV", "Others - MX",
    "Others - UAE", "Others - UK", "Others - ZZ",
]

# Maps each raw CUSTNMBR onto a consolidated customer group. Several customers
# fold into one group (e.g. the three Amazon-DC channels), so forecasts are built
# per group rather than per raw customer. Customers not listed here fall back to
# their own name as a single-member group.
US_GROUPING = {
    'AMAZON-DC': 'AMAZON-DC',
    'AMAZON-DS': 'AMAZON-DC',
    'MARVAL-FBM': 'AMAZON-DC',
    'TARGET-HQ': 'TARGET-HQ',
    'TARGET-DS': 'TARGET-DS',
    'TARGET-PLUS': 'TARGET-DS',
    'HOMDEP': 'HOMDEP',
    'LOWES-HQ': 'LOWES-HQ',
    'LOWES-SOS': 'LOWES-SOS',
    'Web Sales - US': 'Web Sales + Warranty US',
    'Warranty - US': 'Web Sales + Warranty US',
    'CONSTO': 'CONSTO',
    'COSTCO.COM': 'COSTCO.COM',
    'ULINE': 'ULINE',
    'AAFES': 'AAFES',
    'STAPADV.COM': 'STAPADV.COM',
    'STAPLES.COM': 'STAPADV.COM',
    'DILLARDS': 'DILLARDS',
    'NORSTR': 'NORSTR',
    'NORCOM-DS': 'NORSTR',
    'NORCOM': 'NORSTR',
    'Others - US': 'Others - US'
}

EU_GROUPING = {
    'AMAZON-EU': 'AMAZON-EU',
    'Web Sales - EU': 'Web Sales + Warranty EU',
    'CASTOR': 'CASTOR',
    'EBREUNINGER': 'EBREUNINGER',
    'FSKUSTERMANN': 'FSKUSTERMANN',
    'GALLAF': 'GALLAF',
    'LERMER-FR': 'LERMER-FR',
    'SANIKAL-KG': 'SANIKAL-KG',
    'Others - EU': 'Others - EU',
    'Others - IT': 'Others - EU'
}

AU_GROUPING = {
    'Web Sales - AU': 'Web Sales - AU',
    'Others - AU': 'Others - AU'
}

CA_GROUPING = {
    'AMAZON-DCCA': 'AMAZON-DCCA',
    'CANTIR': 'CANTIR',
    'COSTCO-CAN': 'COSTCO-CAN',
    'HOMEDEP-CA.COM': 'HOMEDEP-CA.COM',
    'HOMEDEPOT-CA': 'HOMEDEPOT-CA',
    'Others - CA': 'Others - CA',
    'RONA-HQ': 'RONA-HQ',
    'Web Sales - CA': 'Web Sales - CA',
    'WINMER': 'WINMER'
}

JP_GROUPING = {
    'Web Sales - JP': 'Web Sales - JP',
    'Others - JP': 'Others - JP'
}

COMBINED_GROUPING = {**US_GROUPING, **EU_GROUPING, **AU_GROUPING, **CA_GROUPING, **JP_GROUPING}

# Group labels belonging to each region (the *values* of the maps above), used to
# file each group's output under US / EU / NON-US-EU. A group that came from
# neither map (an unmapped customer kept as its own group) lands in NON-US-EU.
US_GROUPS = set(US_GROUPING.values())
EU_GROUPS = set(EU_GROUPING.values())
AU_GROUPS = set(AU_GROUPING.values())
CA_GROUPS = set(CA_GROUPING.values())
JP_GROUPS = set(JP_GROUPING.values())

def region_for_group(group):
    if group in US_GROUPS:
        return "US (LBC+NJ)"
    elif group in EU_GROUPS:
        return "EU (SH-CTS)"
    elif group in AU_GROUPS:
        return "AU (ACR)"
    elif group in CA_GROUPS:
        return "CA (YYZ5)"
    elif group in JP_GROUPS:
        return "JP (NETDEPOT)"
    return "Other"

# Label used for the all-customers (combined) output.
ALL_CUSTOMERS_LABEL = "ALL CUSTOMERS"

# Internal (snake_case) column names -> display names used in the output sheets.
DISPLAY_NAMES = {
    "weeks_with_data": "Weeks with data",
    "8_week_pos_avg": "8 Week POS/Orders Average",
    "initial_projection_avg": "Initial Projection Average",
    "updated_projection_avg": "Updated Projection Average",
    "projection_difference": "Projection Difference",
    "list_price_usd": "List Price (USD)",
    "revenue_risk_usd": "Revenue Risk (USD)",
}

# Final column order for every summary sheet
SUMMARY_COLUMNS = [
    "SKU",
    "Description",
    "Customer Grouping",
    "Data Source",
    "Weeks with data",
    "8 Week POS/Orders Average",
    "Initial Projection Average",
    "Updated Projection Average",
    "Projection Difference",
    "List Price (USD)",
    "Revenue Risk (USD)",
]


def resolve_input_file():
    """Pick the most recent raw data file and anchor the forecast to its snapshot date.

    Anchoring TODAY to the date in the filename (rather than the wall clock) keeps
    the 8-week lookback and the 15-week forecast aligned with the data snapshot,
    even if the script is run a day or two after the file was generated.
    """
    candidates = []
    for path in glob.glob(INPUT_GLOB):
        m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
        if m:
            candidates.append((m.group(1), path))

    if not candidates:
        raise FileNotFoundError(
            f"No input files matching {INPUT_GLOB}. "
            f"Expected e.g. {RAW_INPUTS_FOLDER}/all_demand_projections_YYYY-MM-DD.xlsx"
        )

    today_str, input_file = max(candidates)  # latest date wins
    return input_file, today_str, pd.Timestamp(today_str)


def load_list_prices(path=None):
    """Load a SKU -> List Price (USD) lookup from a price workbook.

    If ``path`` is given, that workbook is read directly. Otherwise the most
    recently modified file matching LIST_PRICE_GLOB is used. The sheet carries
    two columns, "SKU" and "List Price USD". SKUs with a blank price are dropped
    from the lookup so they map to NaN downstream -- their revenue risk is left
    blank rather than treated as $0 (an unknown price is not zero risk).

    Returns a pandas Series of price indexed by SKU (str), or None if no price
    file is found (in which case revenue risk is left blank for every SKU).
    """
    if path is None:
        candidates = glob.glob(LIST_PRICE_GLOB)
        if not candidates:
            print(
                f"No list-price file matching {LIST_PRICE_GLOB}; "
                f"revenue risk will be left blank.\n"
            )
            return None
        path = max(candidates, key=os.path.getmtime)  # most recently modified

    prices = pd.read_excel(path)
    prices.columns = [str(c).strip() for c in prices.columns]
    prices = prices[["SKU", "List Price USD"]].dropna(subset=["SKU"])
    prices["SKU"] = prices["SKU"].astype(str).str.strip()
    prices["List Price USD"] = pd.to_numeric(prices["List Price USD"], errors="coerce")
    prices = prices.dropna(subset=["List Price USD"]).drop_duplicates("SKU", keep="last")
    lookup = prices.set_index("SKU")["List Price USD"]

    print(f"Loaded {len(lookup)} list prices from {os.path.basename(path)}\n")
    return lookup


def week_anchors(today):
    """Resolve the week boundaries for the forecast.

    WeekDate is the Sunday that *starts* each 7-day week, so the week labelled
    W covers [W, W+6]. A week is only "completely over" once today is past its
    Saturday. To avoid feeding a partially-elapsed week's POS into the average,
    the historical window ends at the last fully-completed week -- but the
    in-progress week is still projected, so it gets a forecast value rather than
    being left as a gap between the actuals and the forecast.

    Example (run on Thu 2026-07-02): the week of 2026-06-28 is still in progress
    (runs through Sat 2026-07-04), so the historical window ends at the last
    completed week, 2026-06-21, and spans the 8 weeks 2026-05-03 .. 2026-06-21.
    The forecast starts at that same in-progress week (2026-06-28): its partial
    POS is excluded from the historical average, but it still receives a
    projection so there is no gap.

    Returns (lookback_start, last_complete_week, first_forecast_week).
    """
    days_since_sunday = (today.weekday() + 1) % 7          # Sun=0 ... Sat=6
    current_week_start = today - pd.Timedelta(days=days_since_sunday)
    last_complete_week = current_week_start - pd.Timedelta(weeks=1)
    lookback_start = last_complete_week - pd.Timedelta(weeks=7)   # 8 weeks inclusive
    # Forecast begins at the in-progress week itself: its partial actuals are
    # kept out of the historical average (window ends at last_complete_week),
    # but it is projected so there's no gap between actuals and the forecast.
    first_forecast_week = current_week_start
    return lookback_start, last_complete_week, first_forecast_week


def aggregate_to_sku_week(df):
    """Collapse rows to one row per SKU-week, summing POS and Projection.

    For a single customer this is an identity (the data already has one row per
    SKU-week), so the per-customer logic is unchanged. For the all-customers view
    it sums each SKU's POS / Projection across every CUSTNMBR, giving the SKU's
    total demand. ``min_count=1`` keeps an all-NaN group as NaN (no data) rather
    than collapsing it to a real 0.
    """
    grp = df.groupby(["SKU", "WeekDate"])
    pos = grp["POS"].sum(min_count=1)
    orders = grp["Orders"].sum(min_count=1)
    proj = grp["Projection"].sum(min_count=1)
    agg = pd.concat([pos, orders, proj], axis=1).reset_index()

    # Description is consistent per SKU; attach the first non-null one.
    desc = df.dropna(subset=["Description"]).groupby("SKU")["Description"].first()
    agg["Description"] = agg["SKU"].map(desc)
    return agg


def top_volume_customers(df, today, top_n=3):
    """For each SKU, rank the customer groups driving the most volume.

    Uses the same 8-week completed window the forecast is built from. Volume is
    measured with POS where the SKU has any, otherwise with Orders (mirroring the
    forecast's POS-then-Orders fallback), summed per Customer Grouping and turned
    into a share of the SKU total. The top ``top_n`` are returned as one string,
    e.g. "AMAZON-DC (61%); Web Sales + Warranty US (31%); AMAZON-EU (3%)".

    Returns a DataFrame [SKU, Top Volume Customer Groups].
    """
    lookback_start, last_complete_week, _ = week_anchors(today)
    win = df[
        (df["WeekDate"] >= lookback_start)
        & (df["WeekDate"] <= last_complete_week)
        & (df["POS"].notna() | df["Orders"].notna())
        & ~df["SKU"].astype(str).str.endswith("*")
    ].copy()

    # Per SKU: rank by POS volume if the SKU has any POS, else by Orders volume.
    has_pos = win.groupby("SKU")["POS"].transform(lambda s: s.notna().any())
    win["vol"] = np.where(has_pos, win["POS"], win["Orders"])

    by_grp = win.groupby(["SKU", "Customer Grouping"], as_index=False)["vol"].sum()
    by_grp = by_grp[by_grp["vol"] > 0]
    by_grp["sku_total"] = by_grp.groupby("SKU")["vol"].transform("sum")
    by_grp = by_grp[by_grp["sku_total"] > 0]
    by_grp["share"] = by_grp["vol"] / by_grp["sku_total"]
    by_grp = by_grp.sort_values(["SKU", "vol"], ascending=[True, False])

    rows = []
    for sku, grp in by_grp.groupby("SKU"):
        head = grp.head(top_n)
        label = "; ".join(
            f"{name} ({share * 100:.0f}%)"
            for name, share in zip(head["Customer Grouping"], head["share"])
        )
        rows.append({"SKU": sku, "Top Volume Customer Groups": label})

    return pd.DataFrame(rows, columns=["SKU", "Top Volume Customer Groups"])


def fit_regression(df, today, grouping_label, breakdown_df=None, list_prices=None):
    """Build a 15-week forecast from the last 8 completed weeks of demand.

    ``df`` must be at SKU-week granularity (see ``aggregate_to_sku_week``).
    For each SKU the forecast is built from POS where available; if a SKU has no
    POS in the window, it falls back to the Orders signal using the identical
    8-week-average + dampened-trend logic. The "Data Source" column records which
    one was used. SKUs with neither POS nor Orders are skipped.
    ``grouping_label`` is written into the "Customer Grouping" column (the group
    name for a per-group file, or ALL_CUSTOMERS_LABEL for the combined file).
    If ``breakdown_df`` (rows that still carry "Customer Grouping") is provided, a
    "Top Volume Customer Groups" column is appended.
    If ``list_prices`` (a SKU -> List Price USD Series, see ``load_list_prices``)
    is provided, two columns are added: "List Price (USD)" and "Revenue Risk
    (USD)" = projection_difference * list price. A negative value means the
    updated forecast fell below the original (revenue at risk on the downside);
    a positive value is upside. SKUs without a known price are left blank.
    Returns (summary_df, weekly_df), or (None, None) if no SKU has POS or Orders
    in the historical window (nothing to forecast from).
    """
    lookback_start, last_complete_week, first_forecast_week = week_anchors(today)

    # Last 8 *completed* weeks (the in-progress week, whose data is only partial,
    # is excluded). Discontinued items (SKU ends in '*') are dropped entirely.
    # Rows are kept if they carry POS OR Orders, so an orders-only SKU survives.
    window = df[
        (df["WeekDate"] >= lookback_start)
        & (df["WeekDate"] <= last_complete_week)
        & (df["POS"].notna() | df["Orders"].notna())
        & ~df["SKU"].astype(str).str.endswith("*")
    ].sort_values(["SKU", "WeekDate"])

    if window.empty:
        return None, None

    # Project 15 weeks forward starting from the first full week after today
    forecast_weeks = pd.date_range(start=first_forecast_week, periods=15, freq="W-SUN")

    summary_rows = []
    weekly_rows = []

    for (sku, desc), grp in window.groupby(["SKU", "Description"]):
        # Prefer POS; fall back to Orders only when the SKU has no POS at all.
        pos_grp = grp[grp["POS"].notna()]
        if not pos_grp.empty:
            source, src_grp = "POS", pos_grp
        else:
            orders_grp = grp[grp["Orders"].notna()]
            if orders_grp.empty:
                continue  # no POS and no Orders -> nothing to forecast
            source, src_grp = "Orders", orders_grp

        src_grp = src_grp.sort_values("WeekDate").reset_index(drop=True)
        y = src_grp[source].values
        n = len(src_grp)
        mean_val = y.mean()

        # Use real week offsets (not positional index) so gaps in the history
        # don't distort the slope.
        x = ((src_grp["WeekDate"] - src_grp["WeekDate"].min()).dt.days / 7).round().values
        slope = np.polyfit(x, y, 1)[0] if n >= 2 and len(set(x)) >= 2 else 0.0

        # Anchor to the mean; nudge by the dampened slope each week out.
        projected_15 = [
            max(round(mean_val + slope * TREND_WEIGHT * (k + 1), 1), 0)
            for k in range(15)
        ]

        summary_rows.append(
            {
                "SKU": sku,
                "Description": desc,
                "Data Source": source,
                "weeks_with_data": n,
                "8_week_pos_avg": round(mean_val, 1),
                "updated_projection_avg": int(round(np.mean(projected_15))),
            }
        )

        for week, projected in zip(forecast_weeks, projected_15):
            weekly_rows.append(
                {
                    "SKU": sku,
                    "Description": desc,
                    "WeekDate": week.date(),
                    "projected_pos": projected,
                }
            )

    if not summary_rows:
        return None, None

    print(f"  Historical window: {lookback_start.date()} -> {last_complete_week.date()} (8 completed weeks)")
    print(f"  Forecast window:   {forecast_weeks[0].date()} -> {forecast_weeks[-1].date()}")
    print(f"  SKUs projected:    {len(summary_rows)}")

    # initial_projection_avg: average of the existing system Projection from the
    # first forecast week (first_forecast_week -- the in-progress week, e.g.
    # 2026-06-28) through the last week that actually has a projection. Anchoring
    # this to the same first_forecast_week keeps it aligned with the updated
    # forecast (both start at the in-progress week). Weeks with a missing
    # projection are excluded from the average (mean() skips NaN), so a SKU whose
    # projection runs out at, say, 2026-11-22 is not penalised for a blank
    # 2026-11-29.
    avg_initial = (
        df[df["WeekDate"] >= first_forecast_week]
        .dropna(subset=["Projection"])
        .groupby("SKU")["Projection"]
        .mean()
        .reset_index()
        .rename(columns={"Projection": "initial_projection_avg"})
    )

    summary_df = (
        pd.DataFrame(summary_rows)
        .merge(avg_initial, on="SKU", how="left")
        .assign(
            initial_projection_avg=lambda d: d["initial_projection_avg"]
            .round()
            .astype("Int64")
        )
        .sort_values("SKU")
        .reset_index(drop=True)
    )


    summary_df['projection_difference'] = summary_df['updated_projection_avg'] - summary_df['initial_projection_avg']

    # Revenue risk: value the change in forecast at each SKU's list price.
    #   revenue_risk = projection_difference * list_price
    # SKUs absent from the price list (or with a blank price) map to NaN, so
    # their revenue risk is left blank rather than treated as $0.
    if list_prices is not None:
        summary_df["list_price_usd"] = summary_df["SKU"].astype(str).map(list_prices)
    else:
        summary_df["list_price_usd"] = np.nan
    summary_df["revenue_risk_usd"] = (
        summary_df["projection_difference"].astype("float64")
        * summary_df["list_price_usd"]
    ).round(2)

    # Stamp the group this file represents, then switch to display column names.
    summary_df["Customer Grouping"] = grouping_label
    summary_df = summary_df.rename(columns=DISPLAY_NAMES)
    summary_df = summary_df[SUMMARY_COLUMNS]

    # All-customers view only: identify which groups drive each SKU's volume.
    if breakdown_df is not None:
        contributors = top_volume_customers(breakdown_df, today)
        summary_df = summary_df.merge(contributors, on="SKU", how="left")
        # SKUs whose POS is present but all zero have no volume to attribute.
        summary_df["Top Volume Customer Groups"] = summary_df[
            "Top Volume Customer Groups"
        ].fillna("(no volume)")

    weekly_df = (
        pd.DataFrame(weekly_rows).sort_values(["SKU", "WeekDate"]).reset_index(drop=True)
    )

    return summary_df, weekly_df


def _autosize_to_headers(worksheet, df, padding=2):
    """Widen each column so its title is fully visible (title length + padding)."""
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = 'A2'

    for i, col in enumerate(df.columns, start=1):
        if i == 1:
            worksheet.column_dimensions[get_column_letter(i)].width = len(str(col)) + 10
        else:
            worksheet.column_dimensions[get_column_letter(i)].width = len(str(col)) + padding


def write_forecast(summary_df, weekly_df, out_path):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        weekly_df.to_excel(writer, sheet_name="weekly_forecast", index=False)
        _autosize_to_headers(writer.sheets["summary"], summary_df)
        _autosize_to_headers(writer.sheets["weekly_forecast"], weekly_df)


if __name__ == "__main__":
    INPUT_FILE, today_str, TODAY = resolve_input_file()
    OUTPUT_FOLDER = f"outputs/demand_projections/regression/{today_str}"
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    print(f"Snapshot date (anchor): {today_str}\n")

    LIST_PRICES = load_list_prices()

    df = pd.read_excel(INPUT_FILE, header=2)
    df = df.rename(
        columns={
            "'Demand'[DisplaySKU]": "SKU",
            "Custnmbr": "CUSTNMBR",
            "Sum of Quantity": "Orders",
        }
    )
    df = df[["SKU", "Description", "CUSTNMBR", "WeekDate", "POS", "Orders", "Projection"]]
    df = df[~df['CUSTNMBR'].isin(CUSTOMERS_TO_IGNORE)]
    df["WeekDate"] = pd.to_datetime(df["WeekDate"])

    # Consolidated customer group. Customers absent from COMBINED_GROUPING fall
    # back to their own name (single-member group), so nothing is dropped.
    df["Customer Grouping"] = df["CUSTNMBR"].map(COMBINED_GROUPING).fillna(df["CUSTNMBR"])
    ungrouped = sorted(
        df.loc[~df["CUSTNMBR"].isin(COMBINED_GROUPING), "CUSTNMBR"].dropna().unique()
    )
    if ungrouped:
        print(
            f"{len(ungrouped)} customers not in COMBINED_GROUPING "
            f"(kept as their own group): {ungrouped}\n"
        )

    # ------------------------------------------------------------------ #
    # 1) Per-group forecasts: one file per Customer Grouping             #
    #    (member customers' POS are summed together)                     #
    # ------------------------------------------------------------------ #
    print("=== Per-group forecasts ===")
    groups = df["Customer Grouping"].dropna().unique().tolist()
    lb, lc, _ = week_anchors(TODAY)
    succeeded, no_pos, no_data, errors = [], [], [], []

    for group in groups:
        try:
            group_df = aggregate_to_sku_week(df[df["Customer Grouping"] == group])

            # Group-level data presence in the 8-week window (discontinued excluded).
            gw = group_df[
                (group_df["WeekDate"] >= lb)
                & (group_df["WeekDate"] <= lc)
                & ~group_df["SKU"].astype(str).str.endswith("*")
            ]
            has_pos = gw["POS"].notna().any()
            has_orders = gw["Orders"].notna().any()
            if not has_pos:
                no_pos.append(group)            # forecast (if any) came from Orders
            if not has_pos and not has_orders:
                no_data.append(group)           # nothing to forecast at all

            summary_df, weekly_df = fit_regression(
                group_df, TODAY, grouping_label=group, list_prices=LIST_PRICES
            )
            if summary_df is None:
                continue

            print(f"[{group}] ok")
            region = region_for_group(group)
            safe_group = group.replace("/", "-")
            out_path = (
                f"{OUTPUT_FOLDER}/{region}/{safe_group}/"
                f"{safe_group}_demand_projections_{today_str}.xlsx"
            )
            write_forecast(summary_df, weekly_df, out_path)
            succeeded.append(group)
        except Exception:
            print(traceback.format_exc())
            errors.append(group)

    # ------------------------------------------------------------------ #
    # 2) All-customers combined: one SKU-level file (POS summed across    #
    #    every customer group)                                            #
    # ------------------------------------------------------------------ #
    print("\n=== All-customers (combined SKU) forecast ===")
    combined_path = f"{OUTPUT_FOLDER}/ALL_CUSTOMERS_demand_projections_{today_str}.xlsx"
    try:
        combined_df = aggregate_to_sku_week(df)
        combined_summary, combined_weekly = fit_regression(
            combined_df, TODAY, grouping_label=ALL_CUSTOMERS_LABEL,
            breakdown_df=df, list_prices=LIST_PRICES,
        )
        if combined_summary is None:
            print("No POS data in the 8-week window for any SKU; combined file skipped.")
        else:
            write_forecast(combined_summary, combined_weekly, combined_path)
            print(f"[ALL_CUSTOMERS] ok -> {combined_path}")
    except Exception:
        print(traceback.format_exc())

    # ------------------------------------------------------------------ #
    print("\n=== Summary ===")
    print(f"Per-group forecasts written: {len(succeeded)}/{len(groups)}")
    print(f"No POS, forecast from Orders (or skipped): {len(no_pos)} -> {no_pos}")
    print(f"Skipped, no POS and no Orders: {len(no_data)} -> {no_data}")
    print(f"Errors: {len(errors)}/{len(groups)} -> {errors}")
    print(f"Input file: {INPUT_FILE}")
    print(f"Output folder: {OUTPUT_FOLDER}")
