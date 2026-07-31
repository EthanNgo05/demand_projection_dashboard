"""The dashboard Excel exporters (compute.view_to_excel / summary_to_excel)
format every downloaded workbook: bold+centered headers, AutoFilter, a frozen
header row, and columns wide enough that nothing truncates."""
from io import BytesIO

import openpyxl
import pandas as pd

from dashboard_app.compute import summary_to_excel, view_to_excel


def _load(data):
    return openpyxl.load_workbook(BytesIO(data))


def _sample():
    return pd.DataFrame(
        {
            "SKU": ["ST1001", "ST1002"],
            "Description": [
                "A very long product description that exceeds the header width",
                "Short",
            ],
            "Projection": [12345.6, 0.0],
        }
    )


def test_summary_to_excel_formats_header_and_filter():
    df = _sample()
    ws = _load(summary_to_excel(df, sheet_name="summary")).active

    assert ws["A1"].font.bold is True
    assert ws["A1"].alignment.horizontal == "center"
    # AutoFilter spans header + data rows.
    assert ws.auto_filter.ref == ws.dimensions
    assert ws.freeze_panes == "A2"


def test_summary_to_excel_column_widths_fit_content():
    df = _sample()
    ws = _load(summary_to_excel(df)).active

    for i, col in enumerate(df.columns, start=1):
        letter = openpyxl.utils.get_column_letter(i)
        assert ws.column_dimensions[letter].width >= len(str(col))

    # The long-text column is wider than its short header.
    desc_letter = openpyxl.utils.get_column_letter(list(df.columns).index("Description") + 1)
    assert ws.column_dimensions[desc_letter].width > len("Description")


def test_summary_to_excel_caps_width():
    df = pd.DataFrame({"Notes": ["x" * 500]})
    ws = _load(summary_to_excel(df)).active
    letter = openpyxl.utils.get_column_letter(1)
    assert ws.column_dimensions[letter].width <= 60


def test_summary_to_excel_empty_frame():
    df = pd.DataFrame({"SKU": [], "Projection": []})
    ws = _load(summary_to_excel(df)).active
    assert ws["A1"].font.bold is True


def test_view_to_excel_formats_both_sheets():
    summary = _sample()
    weekly = pd.DataFrame(
        {"SKU": ["ST1001"], "WeekDate": ["2026-01-05"], "Forecast": [10.0]}
    )
    wb = _load(view_to_excel(summary, weekly))

    assert set(wb.sheetnames) == {"summary", "weekly_forecast"}
    for name in ("summary", "weekly_forecast"):
        ws = wb[name]
        assert ws["A1"].font.bold is True
        assert ws["A1"].alignment.horizontal == "center"
        assert ws.auto_filter.ref == ws.dimensions
        assert ws.freeze_panes == "A2"
