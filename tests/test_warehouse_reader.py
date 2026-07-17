"""Warehouse-export reader: wide/long format sniffing + grid reconstruction.

The missing-projections feature reads the regional warehouse exports through
``data_io.warehouse_wide_to_long`` / ``combine_warehouse_projections``. Two
on-disk layouts exist:

  * the legacy *wide* matrix export (merged SKU cells, week dates as columns,
    blank cell = missing projection), and
  * the *long* table export (banner row, blank row, ``header=2`` with a literal
    ``WeekDate`` column) — what PowerBI "export data" produces today and what
    ``extract_warehouse_projections.py`` writes from the data warehouse. Here a
    missing projection is an *absent row*, so the reader reconstructs the
    pairs × weeks grid to reintroduce the NaN cells downstream code flags.

These tests cover both parsers, the format sniffing, key normalization (GP
CHAR padding, trailing ``'*'`` display markers), and the cross-file week-union
reconstruction. No SQL Server involved — everything runs in the fast suite.
"""

import io

import numpy as np
import openpyxl
import pandas as pd
import pytest

from agent import data_io

TODAY = pd.Timestamp.today().normalize()
# Week anchors safely inside compute_missing_projections' (today, today+15w]
# window, so end-to-end tests are date-stable whenever they run.
W1 = TODAY + pd.Timedelta(weeks=2)
W2 = TODAY + pd.Timedelta(weeks=3)
W3 = TODAY + pd.Timedelta(weeks=4)

LONG_HEADERS = [
    "'Projection_by_Warehouse'[DisplaySKU]", "CUSTNMBR", "WeekDate", "Sum of Proj",
]


def write_wide_export(path, week_dates, blocks, footer="Applied filters:\netc."):
    """Build a legacy wide matrix export.

    ``blocks`` is a list of (sku, [(custnmbr, [proj-per-week]), ...]); the SKU
    cell is merged vertically across its customers, as PowerBI rendered it.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Projection by Warehouse")
    for j, wd in enumerate(week_dates):
        ws.cell(row=1, column=3 + j, value=wd)
    ws.cell(row=2, column=1, value="SKU")
    ws.cell(row=2, column=2, value="CUSTNMBR")
    for j in range(len(week_dates)):
        ws.cell(row=2, column=3 + j, value="Proj")

    r = 3
    for sku, customers in blocks:
        first = r
        for cust, values in customers:
            ws.cell(row=r, column=2, value=cust)
            for j, v in enumerate(values):
                if v is not None:
                    ws.cell(row=r, column=3 + j, value=v)
            r += 1
        ws.cell(row=first, column=1, value=sku)
        if r - first > 1:
            ws.merge_cells(start_row=first, start_column=1,
                           end_row=r - 1, end_column=1)
    r += 1  # blank separator row
    ws.cell(row=r, column=1, value=footer)
    wb.save(path)
    return str(path)


def write_long_export(path, rows, banner="Applied filters:\netc."):
    """Build a long table export: banner, blank row, header=2, data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=banner)
    for j, h in enumerate(LONG_HEADERS):
        ws.cell(row=3, column=1 + j, value=h)
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            ws.cell(row=4 + i, column=1 + j, value=v)
    wb.save(path)
    return str(path)


def plytix_frame(skus, active_in="US,CA,EU"):
    return pd.DataFrame({
        "SKU": skus,
        "SKU Status": "Active",
        "SKU Type": "Product",
        "Active in": active_in,
    })


# --------------------------------------------------------------------------- #
# Wide format (regression: the original path must keep working)               #
# --------------------------------------------------------------------------- #
def test_wide_export_parses_merged_cells_and_footer(tmp_path):
    path = write_wide_export(
        tmp_path / "US_warehouse_projections_2026-07-14.xlsx",
        [W1, W2],
        [
            ("SKU1", [("CUSTA", [10, None]), ("CUSTB", [None, 5])]),
            ("SKU2", [("CUSTA", [7, 8])]),
        ],
    )
    long_df, location = data_io.warehouse_wide_to_long(path)

    assert location == "US"
    assert list(long_df.columns) == data_io.WAREHOUSE_LONG_COLS
    assert len(long_df) == 6  # 3 (SKU, cust) rows x 2 weeks
    # Merged SKU cell forward-filled onto CUSTB's row.
    assert set(long_df[long_df["CUSTNMBR"] == "CUSTB"]["SKU"]) == {"SKU1"}
    # Blank cells preserved as NaN — the "missing" signal.
    row = long_df[(long_df["SKU"] == "SKU1") & (long_df["CUSTNMBR"] == "CUSTA")
                  & (long_df["WeekDate"] == W2)]
    assert row["Projection"].isna().all()
    # Footer text never leaks in as data.
    assert not long_df["SKU"].astype(str).str.contains("Applied").any()


# --------------------------------------------------------------------------- #
# Long format                                                                 #
# --------------------------------------------------------------------------- #
def test_long_export_absent_rows_become_nan(tmp_path):
    # SKU1/CUSTA has W1 only; SKU2/CUSTB has W2 only -> the reconstructed grid
    # holds NaN for the two absent (pair, week) cells.
    path = write_long_export(
        tmp_path / "US_warehouse_projections_2026-07-14.xlsx",
        [["SKU1", "CUSTA", W1, 10], ["SKU2", "CUSTB", W2, 4]],
    )
    combined = data_io.combine_warehouse_projections([(path, path)])

    assert len(combined) == 4  # 2 pairs x 2 weeks
    assert set(combined["Location"]) == {"US"}
    missing = combined[combined["Projection"].isna()]
    assert {(r.SKU, r.CUSTNMBR, r.WeekDate) for r in missing.itertuples()} == {
        ("SKU1", "CUSTA", W2), ("SKU2", "CUSTB", W1),
    }


def test_long_export_zero_rows_are_values_not_missing(tmp_path):
    # An explicit 0 in the export rendered as 0 in the old wide grid — a real
    # (zero) projection, not a missing one.
    path = write_long_export(
        tmp_path / "US_warehouse_projections_2026-07-14.xlsx",
        [["SKU1", "CUSTA", W1, 0], ["SKU1", "CUSTA", W2, 3]],
    )
    combined = data_io.combine_warehouse_projections([(path, path)])

    zero_cell = combined[combined["WeekDate"] == W1]
    assert (zero_cell["Projection"] == 0).all()
    assert not combined["Projection"].isna().any()


def test_long_export_strips_padding_and_star(tmp_path):
    # GP CHAR padding and the '*' display marker must not break the merge
    # against Plytix's stripped SKUs.
    path = write_long_export(
        tmp_path / "US_warehouse_projections_2026-07-14.xlsx",
        [["  SKU1*  ", "  CUSTA  ", W1, 10]],
    )
    long_df, _ = data_io.warehouse_wide_to_long(path)

    assert list(long_df["SKU"]) == ["SKU1"]
    assert list(long_df["CUSTNMBR"]) == ["CUSTA"]


def test_week_union_spans_all_long_files(tmp_path):
    # W2 appears only in the US file. JP's pair must still get a NaN cell for
    # W2 — a week with no JP data at all is exactly a missing JP projection.
    us = write_long_export(
        tmp_path / "US_warehouse_projections_2026-07-14.xlsx",
        [["SKU1", "CUSTA", W1, 10], ["SKU1", "CUSTA", W2, 12]],
    )
    jp = write_long_export(
        tmp_path / "JP_warehouse_projections_2026-07-14.xlsx",
        [["SKU9", "CUSTJ", W1, 5]],
    )
    combined = data_io.combine_warehouse_projections([(us, us), (jp, jp)])

    jp_w2 = combined[(combined["Location"] == "JP") & (combined["WeekDate"] == W2)]
    assert len(jp_w2) == 1
    assert jp_w2["Projection"].isna().all()


def test_headers_only_long_file_contributes_nothing(tmp_path):
    # A region with zero projections still gets a (headers-only) file so the
    # snapshot stays 5-per-date; it must parse cleanly and add no rows.
    empty = write_long_export(
        tmp_path / "JP_warehouse_projections_2026-07-14.xlsx", [])
    us = write_long_export(
        tmp_path / "US_warehouse_projections_2026-07-14.xlsx",
        [["SKU1", "CUSTA", W1, 10]],
    )
    combined = data_io.combine_warehouse_projections([(empty, empty), (us, us)])

    assert set(combined["Location"]) == {"US"}
    assert len(combined) == 1


def test_long_export_from_bytesio_upload(tmp_path):
    # The dashboard upload path wraps the workbook in BytesIO; the sniff read
    # must rewind before the real parse.
    path = write_long_export(
        tmp_path / "EU_warehouse_projections_2026-07-14.xlsx",
        [["SKU1", "CUSTA", W1, 10]],
    )
    with open(path, "rb") as f:
        buf = io.BytesIO(f.read())
    long_df, location = data_io.warehouse_wide_to_long(
        buf, "EU_warehouse_projections_2026-07-14.xlsx")

    assert location == "EU"
    assert len(long_df) == 1


def test_non_region_prefixed_file_skipped(tmp_path):
    # The raw combined SQL dump (no region prefix) sits in the same folder and
    # must be ignored, not parsed.
    path = write_long_export(
        tmp_path / "warehouse_projections_2026-07-14.xlsx",
        [["SKU1", "CUSTA", W1, 10]],
    )
    assert data_io.warehouse_wide_to_long(path) == (None, None)
    combined = data_io.combine_warehouse_projections([(path, path)])
    assert combined.empty
    assert list(combined.columns) == data_io.WAREHOUSE_LONG_COLS


# --------------------------------------------------------------------------- #
# End-to-end: long export -> compute_missing_projections                      #
# --------------------------------------------------------------------------- #
def test_missing_projections_end_to_end_from_long_export(tmp_path):
    # SKU1 (padded + starred in the export) has W1 but not W2/W3 -> flagged
    # with the missing span. SKU2 is fully projected -> not flagged.
    path = write_long_export(
        tmp_path / "US_warehouse_projections_2026-07-14.xlsx",
        [
            ["  SKU1*", "CUSTA", W1, 10],
            ["SKU2", "CUSTA", W1, 5], ["SKU2", "CUSTA", W2, 5],
            ["SKU2", "CUSTA", W3, 5],
        ],
    )
    projections = data_io.combine_warehouse_projections([(path, path)])
    missing = data_io.compute_missing_projections(
        projections, plytix_frame(["SKU1", "SKU2"]), df=None, P=None)

    assert list(missing["SKU"]) == ["SKU1"]
    assert list(missing["Location"]) == ["US"]
    assert missing["First_WeekDate"].iloc[0] == W2
    assert missing["Last_WeekDate"].iloc[0] == W3


def test_history_only_pair_flagged_across_forward_window(tmp_path):
    # A pair whose only rows are historical (the 636-pairs case): it stays in
    # the grid universe, so every forward week is NaN -> flagged fully missing.
    past = TODAY - pd.Timedelta(weeks=4)
    path = write_long_export(
        tmp_path / "US_warehouse_projections_2026-07-14.xlsx",
        [["SKU1", "CUSTA", past, 20], ["SKU2", "CUSTA", W1, 5]],
    )
    projections = data_io.combine_warehouse_projections([(path, path)])
    missing = data_io.compute_missing_projections(
        projections, plytix_frame(["SKU1", "SKU2"]), df=None, P=None)

    flagged = missing[missing["SKU"] == "SKU1"]
    assert len(flagged) == 1
    assert flagged["First_WeekDate"].iloc[0] == W1  # first *future* week
