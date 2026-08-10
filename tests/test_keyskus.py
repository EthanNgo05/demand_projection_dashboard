"""Key-SKU membership, the display-only "Key" chip, and the ``Key SKU`` export
column (dashboard_app.keyskus).

Key SKUs used to be a whole Exceptions tab; they are now an attribute carried by
every table. The invariant these tests protect is the display/export split: the
chip is a decoration on a throwaway copy that must never reach filtering, the
detail-card lookup, or a workbook, while the export column is a real boolean.
"""
from io import BytesIO

import openpyxl
import pandas as pd
import pytest

from dashboard_app.compute import summary_to_excel, with_export_flags
from dashboard_app.keyskus import (
    CHIP_LABEL, KEY_SKU_COL, SKU_TAG_NEUTRAL,
    key_sku_mask, mark_key_sku, sku_chip_column_config, with_key_sku_column,
)
from dashboard_app.watchlist import STAR_PREFIX, WATCHLIST_COL

KEYS = frozenset({"ST1001", "ST1003"})


def _frame(skus):
    return pd.DataFrame({"SKU": skus, "Projection": range(len(skus))})


@pytest.fixture
def isolated_watchlist(tmp_path, monkeypatch):
    """Point the watchlist store at a temp file with a plain-dict session cache.

    ``with_export_flags`` adds the Watchlist column, which reads the shared store —
    neither the JSON file nor ``st.session_state`` exists under bare pytest. Mirrors
    the ``store`` fixture in test_watchlist.py.
    """
    from dashboard_app import watchlist as wl
    monkeypatch.setattr(wl, "WATCHLIST_PATH", str(tmp_path / "watchlist.json"))
    monkeypatch.setattr(wl.st, "session_state", {})


# --------------------------------------------------------------------------- #
# key_sku_mask                                                                 #
# --------------------------------------------------------------------------- #
def test_mask_matches_listed_skus():
    mask = key_sku_mask(_frame(["ST1001", "ST1002", "ST1003"]), KEYS)
    assert list(mask) == [True, False, True]


@pytest.mark.parametrize("cell", [
    f"{STAR_PREFIX}ST1001",   # watchlist star prefix (tables.mark_starred_sku)
    "ST1001*",                # data-quality asterisk suffix
    "  ST1001  ",             # stray whitespace
    f"{STAR_PREFIX}ST1001*",  # both decorations at once
])
def test_mask_sees_through_display_decoration(cell):
    """A decorated SKU cell is still the same SKU. The star prefix in particular is
    applied by the same render path that then asks for the chip, so a mask that
    matched raw strings would silently drop the chip from every starred row."""
    assert bool(key_sku_mask(_frame([cell]), KEYS).iloc[0])


def test_mask_is_all_false_without_a_key_list():
    """No snapshot extracted yet -> nothing is a key SKU, rather than an error."""
    assert not key_sku_mask(_frame(["ST1001"]), frozenset()).any()


def test_mask_is_all_false_without_a_sku_column():
    df = pd.DataFrame({"Customer Grouping": ["AMAZON-DC"], "Units": [3]})
    mask = key_sku_mask(df, KEYS)
    assert list(mask) == [False]


def test_mask_ignores_rolled_up_count_labels():
    """At the Customer/Region grain the Exceptions frame puts "12 SKUs" in the SKU
    cell (exceptions._dim_labels). Those rows are not key SKUs and must not be
    chipped, which is also what keeps the filter field from being offered there."""
    assert not key_sku_mask(_frame(["12 SKUs", "1 SKU"]), KEYS).any()


# --------------------------------------------------------------------------- #
# mark_key_sku (display) — the chip lives INSIDE the SKU cell, to its right    #
# --------------------------------------------------------------------------- #
def test_mark_puts_the_chip_inside_the_sku_cell():
    df = _frame(["ST1001", "ST1002"])
    out, sku_values = mark_key_sku(df, KEYS)

    # No new column: the SKU cell itself becomes a tag list, chip last so it renders
    # to the RIGHT of the SKU name.
    assert list(out.columns) == ["SKU", "Projection"]
    assert list(out["SKU"]) == [["ST1001", CHIP_LABEL], ["ST1002"]]
    assert sku_values == ["ST1001", "ST1002"]


def test_mark_keeps_the_star_prefix_left_of_the_sku():
    """A starred key SKU reads star, then SKU, then chip — the star is part of the
    SKU string (mark_starred_sku ran first), so it must survive into the option."""
    out, sku_values = mark_key_sku(_frame([f"{STAR_PREFIX}ST1001"]), KEYS)
    assert out["SKU"].iloc[0] == [f"{STAR_PREFIX}ST1001", CHIP_LABEL]
    assert sku_values == [f"{STAR_PREFIX}ST1001"]


def test_mark_leaves_the_source_frame_untouched():
    """The chip goes onto a copy — the caller's frame still drives filtering, the
    detail-card row lookup and the export, none of which tolerate a tag list."""
    df = _frame(["ST1001"])
    mark_key_sku(df, KEYS)
    assert list(df["SKU"]) == ["ST1001"]


def test_mark_is_a_noop_when_nothing_matches():
    """No tag styling on tables with no key SKUs — the SKU stays ordinary text."""
    df = _frame(["ST9998", "ST9999"])
    out, sku_values = mark_key_sku(df, KEYS)
    assert sku_values is None
    assert list(out["SKU"]) == ["ST9998", "ST9999"]


def test_mark_is_a_noop_without_a_sku_column():
    df = pd.DataFrame({"Month": ["2026-01"], "Revenue": [10.0]})
    out, sku_values = mark_key_sku(df, KEYS)
    assert sku_values is None
    assert list(out.columns) == ["Month", "Revenue"]


def test_chip_config_colours_only_the_key_option():
    """Colours attach to options positionally, so every SKU must be listed (else it
    renders unstyled) and "Key" must be last to pick up the blue."""
    _, sku_values = mark_key_sku(_frame(["ST1001", "ST1002"]), KEYS)
    opts = sku_chip_column_config(sku_values)["SKU"]["type_config"]["options"]

    assert [o["value"] for o in opts] == ["ST1001", "ST1002", CHIP_LABEL]
    assert [o["color"] for o in opts] == [SKU_TAG_NEUTRAL, SKU_TAG_NEUTRAL, "blue"]


# --------------------------------------------------------------------------- #
# with_key_sku_column (export)                                                 #
# --------------------------------------------------------------------------- #
def test_export_column_is_boolean_and_sits_after_sku():
    out = with_key_sku_column(_frame(["ST1001", "ST1002"]), KEYS)
    assert list(out.columns) == ["SKU", KEY_SKU_COL, "Projection"]
    assert out[KEY_SKU_COL].dtype == bool
    assert list(out[KEY_SKU_COL]) == [True, False]


def test_export_column_is_added_even_when_nothing_matches():
    """Unlike the chip: a column of FALSE is an answer, an absent column is not."""
    out = with_key_sku_column(_frame(["ST9999"]), KEYS)
    assert list(out[KEY_SKU_COL]) == [False]


def test_export_column_skips_frames_without_a_sku_column():
    """Export call sites wrap unconditionally, so the no-SKU case must no-op."""
    df = pd.DataFrame({"Week": ["2026-01-04"], "Units": [7]})
    assert with_key_sku_column(df, KEYS) is df


def test_export_column_is_not_added_twice():
    once = with_key_sku_column(_frame(["ST1001"]), KEYS)
    assert list(with_key_sku_column(once, KEYS).columns) == list(once.columns)


def test_chip_never_reaches_a_workbook(isolated_watchlist):
    """Display and export decoration are mutually exclusive by construction: exports
    wrap the undecorated frame, so a workbook carries plain SKU strings plus the flag
    columns and no tag lists (openpyxl cannot write a list, so a leak is a hard fail).

    ``with_export_flags`` is the single door every export goes through, hence the
    round-trip here rather than on ``with_key_sku_column`` alone."""
    df = pd.DataFrame({
        "SKU": ["ST1001", "ST1002"],
        "Customer Grouping": ["AMAZON-DC", "WEB-US"],
        "Projection": [1, 2],
    })
    exported = with_export_flags(df)

    ws = openpyxl.load_workbook(BytesIO(summary_to_excel(exported))).active
    assert [c.value for c in ws[1]] == [
        "SKU", KEY_SKU_COL, WATCHLIST_COL, "Customer Grouping", "Projection",
    ]
    assert [ws.cell(row=r, column=1).value for r in (2, 3)] == ["ST1001", "ST1002"]
