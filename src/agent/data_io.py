"""Shared, Streamlit-free I/O extracted from dashboard.py (Phase 2).

Single source of truth for raw/price-file discovery and raw-frame cleaning,
used by both ``dashboard.py`` (which keeps its ``@st.cache_data`` wrappers
around thin calls into this module) and the agent's ingest node.

Each function takes the pipeline module ``P`` explicitly instead of relying
on the dashboard's Streamlit-session ``pipeline_path()``/``load_pipeline()``
globals. Passing ``P=None`` falls back to the first configured model, which
is safe for discovery/cleaning because ``RAW_INPUTS_FOLDER`` /
``LIST_PRICE_GLOB`` / ``CUSTOMERS_TO_IGNORE`` / ``COMBINED_GROUPING`` are
identical across the three model files (see README, "The pipeline contract").

Must never import streamlit (directly or transitively).
"""

import glob
import os
import re
import urllib.request
from io import BytesIO
from typing import NamedTuple

import numpy as np
import openpyxl
import pandas as pd

from agent.config import ALL_CUSTOMERS_VIEW, MODEL_OPTIONS, region_from_view
from agent.model_loader import load_pipeline

# Repo root (parent of src/, the folder holding raw_inputs/ + outputs/), so
# relative RAW_INPUTS_FOLDER / LIST_PRICE_GLOB paths resolve there. This file is
# src/agent/data_io.py, so climb three levels: agent -> src -> repo root.
HERE = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)

# Public Plytix channel feed that publishes the same product data as the
# list_prices_*.xlsx export (SKU, List Price USD, SKU Status, SKU Type,
# "Active in"). Used as the default price/Plytix source so no one has to drag a
# file. Overridable via the PLYTIX_FEED_URL env var (blank disables the feed and
# falls back to the local xlsx). See read_plytix / prices_from_plytix below.
PLYTIX_FEED_URL = os.getenv(
    "PLYTIX_FEED_URL",
    "https://pim.plytix.com/channels/6a56637a49b9e8566d5f2d4f/feed",
)


def default_pipeline():
    """Load the first configured model (any model drives discovery/cleaning)."""
    if not MODEL_OPTIONS:
        raise FileNotFoundError(
            "No forecasting pipeline found — expected "
            "models/exponential_smoothing.py, models/holt_winters.py, "
            "models/xgboost.py or models/regression.py next to dashboard.py."
        )
    return load_pipeline(next(iter(MODEL_OPTIONS.values())))


def _resolve_pipeline(P):
    return default_pipeline() if P is None else P


def view_frame(df, view, P=None):
    """Rows of ``df`` belonging to ``view``: the full frame (ALL CUSTOMERS),
    one region's groups (an "All Customers - <region>" rollup), or one
    customer group. The shared view->frame step for every agent node, kept in
    exact lockstep with dashboard.compute_view's filtering.

    ``P=None`` falls back to the first configured model — safe because
    region_for_group is identical across the three model files (see the
    pipeline contract). str() on its result matches how the view string was
    built (a custom pipeline may return non-string region labels).
    """
    if view == ALL_CUSTOMERS_VIEW:
        return df
    region = region_from_view(view)
    if region is not None:
        P = _resolve_pipeline(P)
        groups = df["Customer Grouping"].map(
            lambda g: str(P.region_for_group(g))
        )
        return df[groups == region]
    return df[df["Customer Grouping"] == view]


def _raw_dir(P=None):
    """Resolve the folder holding the raw + price files.

    Honours DEMAND_RAW_DIR if set; otherwise uses the pipeline's own
    RAW_INPUTS_FOLDER constant (e.g. ``raw_inputs/demand_projections``),
    resolved relative to the repo root when it is a relative path. This means
    moving the raw folder in the pipeline is picked up here automatically.
    """
    P = _resolve_pipeline(P)
    folder = os.environ.get("DEMAND_RAW_DIR")
    if folder is None:
        folder = getattr(P, "RAW_INPUTS_FOLDER", None)
        if folder is None:
            # Older pipeline without the constant: derive it from INPUT_GLOB
            # if present, otherwise use the standard default location.
            input_glob = getattr(P, "INPUT_GLOB", None)
            folder = (
                os.path.dirname(input_glob)
                if input_glob
                else "raw_inputs/demand_projections"
            )
        if not os.path.isabs(folder):
            folder = os.path.join(HERE, folder)
    return folder


def raw_glob(P=None):
    """Build the raw-file glob, tracking the pipeline's RAW_INPUTS_FOLDER."""
    return os.path.join(_raw_dir(P), "all_demand_projections_*.xlsx")


def price_glob(P=None):
    """Build the list-price glob, mirroring the pipeline's LIST_PRICE_GLOB.

    The pipeline's glob (folder included) is used as-is, resolved relative to
    the repo root when it is a relative path — so every caller scans the same
    folder the batch pipeline does, regardless of the working directory.
    """
    P = _resolve_pipeline(P)
    pattern = getattr(
        P, "LIST_PRICE_GLOB",
        os.path.join("raw_inputs/list_prices", "list_prices_*.xlsx"),
    )
    if not os.path.isabs(pattern):
        pattern = os.path.join(HERE, pattern)
    return pattern


def discover_price_file(P=None):
    """Newest list-price file in the raw folder, or None if there isn't one."""
    matches = glob.glob(price_glob(P))
    return max(matches, key=os.path.getmtime) if matches else None


def _date_from_name(name):
    m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(name))
    return m.group(1) if m else None


def discover_raw_files(P=None):
    """Return [(date_str, path)] newest first, mirroring resolve_input_file()."""
    out = []
    for path in glob.glob(raw_glob(P)):
        d = _date_from_name(path)
        if d:
            out.append((d, path))
    return sorted(out, reverse=True)


def _clean(raw_df, P):
    """Apply the exact preprocessing from the pipeline's __main__ block.

    Mirrors the updated pipeline: 'Sum of Quantity' -> Orders, and POS /
    Orders / Projection are all carried through. Falls back gracefully if an
    older file lacks the Orders column (an all-NaN Orders column is added so
    the POS-then-Orders logic still runs without a KeyError).
    """
    rename = {"'Demand'[DisplaySKU]": "SKU", "Custnmbr": "Customer"}
    if "Sum of Quantity" in raw_df.columns:
        rename["Sum of Quantity"] = "Orders"
    df = raw_df.rename(columns=rename)

    if "Orders" not in df.columns:
        df["Orders"] = np.nan  # legacy file without an Orders/Sum of Quantity col

    df = df[["SKU", "Description", "Customer", "WeekDate", "POS", "Orders", "Projection"]]
    # The fixed-width warehouse export space-pads its key columns (e.g.
    # 'BT1028      ', 'AMAZON-DS      '). Strip that surrounding whitespace here —
    # the single ingestion boundary both the dashboard and agent share — before
    # any key-based join or lookup runs. SKU padding made every SKU miss the
    # (stripped) list-price index (blank revenue risk) and the Plytix SKU sets
    # (active-in / discontinued checks silently ran on nothing). Customer padding
    # made padded customers miss CUSTOMERS_TO_IGNORE and COMBINED_GROUPING, so a
    # group like AMAZON-DC fragmented across its padded/clean spellings instead of
    # folding together. Strip Customer *before* the ignore filter and grouping map.
    df["SKU"] = df["SKU"].astype(str).str.strip()
    df["Customer"] = df["Customer"].astype(str).str.strip()
    df = df[~df["Customer"].isin(P.CUSTOMERS_TO_IGNORE)]
    df["WeekDate"] = pd.to_datetime(df["WeekDate"])
    df["Customer Grouping"] = (
        df["Customer"].map(P.COMBINED_GROUPING).fillna(df["Customer"])
    )
    return df


def _parquet_sidecar_path(xlsx_path):
    """The ``.parquet`` sidecar path for a snapshot ``.xlsx`` (same basename)."""
    root, _ = os.path.splitext(xlsx_path)
    return root + ".parquet"


def read_raw_frame(path):
    """Read a raw demand snapshot into the pre-``_clean`` frame, fast path first.

    The nightly extract writes a ``.parquet`` sidecar next to each
    ``all_demand_projections_<date>.xlsx`` (see
    ``extract_demand_details.write_parquet_sidecar``). Parquet preserves dtypes
    and carries no banner rows, so it loads far faster than re-parsing the
    workbook with openpyxl. This helper prefers that sidecar and falls back to
    the ``.xlsx`` (``header=2`` for the two-row PowerBI banner):

      1. sidecar exists and is at least as new as the ``.xlsx`` -> read Parquet;
      2. otherwise read the ``.xlsx`` and **lazily backfill** the sidecar for
         next time (best-effort — swallowed on read-only hosts / missing engine),
         so pre-existing snapshots and manual PowerBI exports get the fast path
         on their second load.

    Returns the raw frame with the PowerBI column names ``_clean`` expects
    (``'Demand'[DisplaySKU]``, ``Custnmbr``, ``WeekDate``, ``POS``,
    ``Projection``, optionally ``Sum of Quantity``). Discovery stays keyed on the
    ``.xlsx`` (see ``discover_raw_files``); only the physical read is swapped.
    """
    sidecar = _parquet_sidecar_path(path)
    try:
        if os.path.exists(sidecar) and (
            not os.path.exists(path)
            or os.path.getmtime(sidecar) >= os.path.getmtime(path)
        ):
            return pd.read_parquet(sidecar)
    except Exception:  # corrupt/unreadable sidecar or engine missing -> xlsx
        pass

    raw = pd.read_excel(path, header=2)
    # Lazy backfill so the next load hits the fast path. Never let a failed
    # sidecar write break the read.
    try:
        tmp = sidecar + ".tmp"
        raw.to_parquet(tmp, index=False)
        os.replace(tmp, sidecar)
    except Exception:
        try:
            if os.path.exists(sidecar + ".tmp"):
                os.remove(sidecar + ".tmp")
        except OSError:
            pass
    return raw


def load_raw(path, P=None):
    """Read + clean a raw demand workbook from disk.

    Reads via ``read_raw_frame`` (Parquet sidecar when available, else the
    ``header=2`` PowerBI workbook), then applies the shared ``_clean`` — the
    same result dashboard.py's ``load_raw_from_path`` produces.
    """
    P = _resolve_pipeline(P)
    raw = read_raw_frame(path)
    return _clean(raw, P)


# --------------------------------------------------------------------------- #
# Plytix-based SKU exclusions (single source of truth, shared by dashboard.py  #
# and the agent's ingest node). A SKU must never be forecast — or flagged by   #
# the agent — when it is discontinued/inactive, or when it appears in a region #
# it is not "Active in" per the Plytix export. Ported verbatim from            #
# dashboard.py so both paths drop the exact same rows before forecasting.      #
#                                                                              #
# The Plytix export doubles as the list-price file, so ``read_plytix`` reads   #
# the same file ``discover_price_file`` returns.                               #
# --------------------------------------------------------------------------- #

# The warehouse regions we check "Active in" against. A SKU should only be
# projected in a region it is "Active in" (per Plytix); a projection in any
# other region is flagged and excluded from the forecast.
WAREHOUSE_REGIONS = ["AU", "CA", "EU", "JP", "US"]

INACTIVE_COLS = [
    "SKU", "Region Code", "Region", "Active in", "Customer Grouping",
    "Customer", "First_WeekDate", "Last_WeekDate", "Original_Projection", "Source",
]

DISCONTINUED_COLS = [
    "SKU", "SKU Status", "Region", "Region Code", "Customer Grouping", "Customer",
    "First_WeekDate", "Last_WeekDate", "Original_Projection",
]

MISSING_POS_COLS = [
    "SKU", "Region Code", "Region", "Active in", "Customer",
    "Data Source", "Last Value",
    "First Missing Week", "Last Missing Week", "Missing Weeks",
]


def _read_csv_url(url, timeout=30):
    """Fetch a CSV feed over HTTP(S) into a DataFrame, with a timeout.

    Uses stdlib urllib (no extra dependency) so a hung endpoint can't stall the
    dashboard/agent indefinitely."""
    with urllib.request.urlopen(url, timeout=timeout) as resp:
        raw = resp.read()
    return pd.read_csv(BytesIO(raw))


def read_plytix(src):
    """Read the raw Plytix export (for the 'Active in' / discontinued checks).

    ``src`` may be:
      - an ``http(s)`` URL to the Plytix channel feed (CSV) — fetched with a
        timeout and parsed as CSV;
      - a ``.csv`` filesystem path — read as CSV;
      - a filesystem path or file-like object (e.g. a BytesIO of an uploaded
        workbook) — read as an Excel workbook, as before.

    Column names are stripped so the feed's trailing-space headers don't break
    the exact-name column checks downstream. Shared by the dashboard's on-disk,
    upload, and feed paths and the agent's ingest node."""
    if isinstance(src, str) and src.lower().startswith("http"):
        df = _read_csv_url(src)
    elif isinstance(src, str) and src.lower().endswith(".csv"):
        df = pd.read_csv(src)
    else:
        df = pd.read_excel(src)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def prices_from_plytix(plytix_df):
    """SKU -> List Price (USD) Series from an already-read Plytix frame.

    Mirrors the tail of each model's ``load_list_prices`` so the CSV/feed path
    yields the exact same lookup without re-reading through ``pd.read_excel``.
    SKUs with a blank price are dropped so they map to NaN downstream (an
    unknown price is left blank rather than treated as $0). Returns None if the
    frame lacks the required columns."""
    if plytix_df is None or not {"SKU", "List Price USD"}.issubset(plytix_df.columns):
        return None
    prices = plytix_df[["SKU", "List Price USD"]].dropna(subset=["SKU"]).copy()
    prices["SKU"] = prices["SKU"].astype(str).str.strip()
    prices["List Price USD"] = pd.to_numeric(prices["List Price USD"], errors="coerce")
    prices = prices.dropna(subset=["List Price USD"]).drop_duplicates("SKU", keep="last")
    return prices.set_index("SKU")["List Price USD"]


def _this_week_start():
    """Sunday-anchored start of the current week as a Timestamp.

    "This week" is deliberately the real current week (not the snapshot's
    anchor): the excluded tables report which future projections are being
    dropped going forward from now."""
    today = pd.Timestamp.today().normalize()
    return today - pd.Timedelta(days=(today.weekday() + 1) % 7)


def _active_in_list(sku_active_in, sku):
    """The list of regions a SKU is 'Active in' (e.g. ['US', 'CA', 'EU'])."""
    return [x.strip() for x in str(sku_active_in.get(sku, "")).split(",")]


def _region_code(P, grouping):
    """Two-letter region code for a customer grouping (US/CA/EU/JP/AU), or None.

    The pipeline's region_for_group returns labels like "JP (NETDEPOT)" or
    "US (LBC+NJ)"; the leading two letters are the region code we match against
    Plytix 'Active in'. Anything else (e.g. "Other") returns None.
    """
    try:
        label = P.region_for_group(grouping)
    except Exception:
        return None
    code = str(label)[:2].upper()
    return code if code in WAREHOUSE_REGIONS else None


def compute_active_products(plytix_df):
    """From the Plytix export, the set of active-product SKUs and a
    SKU -> 'Active in' string lookup.

    "Active product" mirrors inactive_projections.ipynb: SKU Status == Active,
    SKU Type == Product, and SKUs starting LS/AS excluded. Trailing '*' markers
    are stripped so SKUs line up with the demand file.

    Returns (active_sku_set, sku_active_in) or (None, None) if the Plytix export
    lacks the columns the check needs (an older list-price file).
    """
    required = {"SKU", "SKU Status", "SKU Type", "Active in"}
    if plytix_df is None or not required.issubset(plytix_df.columns):
        return None, None
    p = plytix_df.copy()
    p["SKU"] = p["SKU"].astype(str).str.rstrip("*")
    act = p[(p["SKU Status"] == "Active") & (p["SKU Type"] == "Product")]
    act = act[~act["SKU"].str.startswith(("LS", "AS"))]
    active_sku_set = set(act["SKU"])
    # Full (un-exploded) Active-in string per SKU, e.g. "US,CA,UK,SG,EU,AU".
    sku_active_in = dict(zip(p["SKU"], p["Active in"].astype(str)))
    return active_sku_set, sku_active_in


def compute_inactive_projections(df, active_sku_set, sku_active_in, P,
                                 anchors=None):
    """Active products showing up in a region they are not 'Active in'.

    This is the fix for cases like ST1082 (active in US/CA/UK/SG/EU/AU but not
    JP), which still appeared in the JP (NETDEPOT) summary: the dashboard builds
    a forward forecast for any SKU with demand history in a region. We look at
    the *demand file itself* — the same data the dashboard forecasts — map each
    customer to its region via the pipeline's region_for_group, and flag any
    active product whose region is not in its Plytix 'Active in' list.

    Returns a table (columns = INACTIVE_COLS) of the flagged
    SKU x customer x region combinations, empty if none or inputs are missing.
    """
    if not active_sku_set or not sku_active_in:
        return pd.DataFrame(columns=INACTIVE_COLS)

    frames = []

    # ----- Primary: the demand file, by customer-group region ---------------
    if df is not None and not df.empty:
        m = df.copy()
        m["SKU"] = m["SKU"].astype(str).str.rstrip("*")
        m = m[m["SKU"].isin(active_sku_set)]
        if not m.empty:
            m["Region"] = m["Customer Grouping"].map(
                lambda g: P.region_for_group(g)
            )
            m["Region Code"] = m["Customer Grouping"].map(
                lambda g: _region_code(P, g)
            )
            m = m[m["Region Code"].notna()]
            keep = [
                loc not in _active_in_list(sku_active_in, sku)
                for sku, loc in zip(m["SKU"], m["Region Code"])
            ]
            m = m[keep]
            m["WeekDate"] = pd.to_datetime(m["WeekDate"])
            # Only flag pairs the dashboard would actually forecast — i.e. that
            # carry a POS/Orders demand signal in the historical window (that is
            # exactly what puts a SKU in a region's summary). Without anchors we
            # fall back to any presence.
            if anchors is not None and not m.empty:
                lb, lcw, _ = anchors
                sig = (
                    (m["WeekDate"] >= lb) & (m["WeekDate"] <= lcw)
                    & (m["POS"].notna() | m.get("Orders", pd.Series(index=m.index)).notna())
                )
                live = m.loc[sig, ["SKU", "Customer", "Region Code"]].drop_duplicates()
                m = m.merge(live, on=["SKU", "Customer", "Region Code"], how="inner")
            if not m.empty:
                # Original projection over future weeks (this week onward) —
                # averaged per week — the projected weekly volume being excluded
                # going forward. Uses the same week boundary as the excluded
                # table's "future only" toggle.
                m["_future_proj"] = pd.to_numeric(
                    m["Projection"], errors="coerce"
                ).where(m["WeekDate"] >= _this_week_start())
                g = m.groupby(
                    ["SKU", "Region Code", "Region", "Customer Grouping", "Customer"],
                    as_index=False,
                ).agg(
                    First_WeekDate=("WeekDate", "min"),
                    Last_WeekDate=("WeekDate", "max"),
                    Original_Projection=("_future_proj", "mean"),
                )
                g["Active in"] = g["SKU"].map(lambda s: sku_active_in.get(s))
                g["Source"] = "Demand file"
                frames.append(g)

    if not frames:
        return pd.DataFrame(columns=INACTIVE_COLS)

    out = pd.concat(frames, ignore_index=True)[INACTIVE_COLS]
    out = out.drop_duplicates(subset=["SKU", "Region Code", "Customer"], keep="first")
    return out.sort_values(["SKU", "Region Code", "Customer"]).reset_index(drop=True)


def compute_discontinued_products(plytix_df):
    """SKU -> 'SKU Status' lookup for Discontinued/Inactive products.

    Mirrors discontinued_with_projections.ipynb: keep rows whose SKU Status is
    'Discontinued' or 'Inactive'. Trailing '*' markers are stripped so SKUs line
    up with the demand file. Returns None if the Plytix export lacks the columns
    the check needs (an older list-price file).
    """
    required = {"SKU", "SKU Status"}
    if plytix_df is None or not required.issubset(plytix_df.columns):
        return None
    p = plytix_df.copy()
    p["SKU"] = p["SKU"].astype(str).str.rstrip("*")
    disc = p[p["SKU Status"].isin(["Discontinued", "Inactive"])]
    return dict(zip(disc["SKU"], disc["SKU Status"]))


def compute_discontinued_projections(df, disc_status, P):
    """Discontinued/inactive products that still carry future projections.

    Ported from discontinued_with_projections.ipynb: intersect the demand file
    with the discontinued/inactive SKU set, keep only future projection weeks
    (WeekDate after today), and aggregate to one row per SKU x customer with the
    first/last projected week. A Region column (via the pipeline's
    region_for_group) is added so a by-customer-group view can be scoped to its
    own region.

    Returns a table (columns = DISCONTINUED_COLS), empty if none or inputs are
    missing.
    """
    if not disc_status or df is None or df.empty:
        return pd.DataFrame(columns=DISCONTINUED_COLS)

    m = df.copy()
    m["SKU"] = m["SKU"].astype(str).str.rstrip("*")
    m = m[m["SKU"].isin(disc_status)]
    if m.empty:
        return pd.DataFrame(columns=DISCONTINUED_COLS)

    # Future projections only. Like the active-in table, "future" starts at the
    # beginning of the current week (Sunday-anchored via _this_week_start), so
    # the in-progress week is included — e.g. 7/5 counts while the 7/7 week is
    # not yet over.
    m["WeekDate"] = pd.to_datetime(m["WeekDate"])
    week_start = _this_week_start()
    m = m[m["WeekDate"] >= week_start]
    if m.empty:
        return pd.DataFrame(columns=DISCONTINUED_COLS)

    m["Region"] = m["Customer Grouping"].map(lambda g: P.region_for_group(g))
    m["Region Code"] = m["Customer Grouping"].map(lambda g: _region_code(P, g))
    m["_future_proj"] = pd.to_numeric(m["Projection"], errors="coerce")
    # dropna=False: Region Code is None for non-warehouse regions (e.g. "Other");
    # without it those discontinued SKUs would be silently dropped from the table.
    g = m.groupby(
        ["SKU", "Region", "Region Code", "Customer Grouping", "Customer"],
        as_index=False, dropna=False,
    ).agg(
        First_WeekDate=("WeekDate", "min"),
        Last_WeekDate=("WeekDate", "max"),
        Original_Projection=("_future_proj", "mean"),
    )
    g["SKU Status"] = g["SKU"].map(lambda s: disc_status.get(s))
    out = g[DISCONTINUED_COLS]
    return out.sort_values(["SKU", "Customer"]).reset_index(drop=True)


def compute_missing_pos_orders(df, plytix_df, P, anchors=None):
    """Active SKUs (incl. Parts) with no recent POS/Orders where they're active.

    Ported from missing_pos.ipynb. For every ACTIVE SKU (Plytix SKU Status ==
    Active — any SKU Type, so Parts like PD6306 are kept, unlike
    compute_active_products), look at each customer whose warehouse region is one
    the SKU is "Active in", and flag the SKU x customer combos with no POS/Orders
    signal at or after the last completed week. This surfaces prolonged stockouts
    and gone-silent channels the forecast would otherwise carry quietly.

    Scope is the TRAILING 3 MONTHS: a combo only surfaces if it sold within the
    past 3 months (last_complete_week - 3mo) and has SINCE gone silent. Two
    categories are deliberately excluded because there is no recent demand to be
    "missing" — combos that never had any POS/Orders (never part of the assortment)
    and long-dead combos whose last sale is more than 3 months old. The gap is
    measured from the week after the combo's last data week through the last
    completed week.

    Returns a table (columns = MISSING_POS_COLS), empty if none. Returns None if
    the Plytix export lacks the columns the check needs (an older list-price file)
    or there are no anchors / no demand frame to work from.
    """
    required = {"SKU", "SKU Status", "Active in"}
    if plytix_df is None or not required.issubset(plytix_df.columns):
        return None
    if df is None or df.empty or anchors is None:
        return None

    # --- Plytix: active SKUs (any Type) and their "Active in" regions ---------
    p = plytix_df.copy()
    p["SKU"] = p["SKU"].astype(str).str.rstrip("*").str.strip()
    active = p[p["SKU Status"] == "Active"]
    active_skus = set(active["SKU"])
    sku_active_in = dict(zip(active["SKU"], active["Active in"].astype(str)))

    _, last_complete_week, _ = anchors

    # --- Candidate combos: active SKU x customer, restricted to active regions -
    m = df.copy()
    m["SKU"] = m["SKU"].astype(str).str.rstrip("*").str.strip()
    m["WeekDate"] = pd.to_datetime(m["WeekDate"])
    m = m[m["SKU"].isin(active_skus)]
    if m.empty:
        return pd.DataFrame(columns=MISSING_POS_COLS)

    combos = m[["SKU", "Customer", "Customer Grouping"]].drop_duplicates().copy()
    combos["Region Code"] = combos["Customer Grouping"].map(lambda g: _region_code(P, g))
    combos["Region"] = combos["Customer Grouping"].map(lambda g: P.region_for_group(g))
    combos = combos[combos["Region Code"].notna()]
    combos = combos[[
        loc in _active_in_list(sku_active_in, sku)
        for sku, loc in zip(combos["SKU"], combos["Region Code"])
    ]]
    if combos.empty:
        return pd.DataFrame(columns=MISSING_POS_COLS)

    # --- Each combo's last data-bearing week + the POS/Orders recorded on it ----
    # A recorded 0 counts as data; only NaN/absent is "no data". (SKU, Customer,
    # WeekDate) isn't unique, so sum the duplicate-grain rows within each week
    # (min_count=1 keeps an all-NaN week as NaN, not 0), then take each combo's
    # latest data-bearing week and the values booked on it. Future weeks carry
    # booked orders, so a combo with data at/after the reference week is NOT missing.
    weekly = (
        m.groupby(["SKU", "Customer", "WeekDate"], as_index=False)[["POS", "Orders"]]
        .sum(min_count=1)
    )
    weekly = weekly[weekly["POS"].notna() | weekly["Orders"].notna()]
    last_rows = (
        weekly.loc[weekly.groupby(["SKU", "Customer"])["WeekDate"].idxmax()]
        .set_index(["SKU", "Customer"])
    )
    key = pd.MultiIndex.from_frame(combos[["SKU", "Customer"]])
    combos["Last Data Week"] = key.map(last_rows["WeekDate"])
    combos["_last_pos"] = key.map(last_rows["POS"])
    combos["_last_orders"] = key.map(last_rows["Orders"])

    # --- Currently missing, but sold within the past 3 months ------------------
    # Kept: combos whose most recent data is BEFORE last_complete_week AND no older
    # than 3 months. The trailing-3-month floor drops long-dead combos and, since
    # NaT >= three_months_ago is False, never-had-data combos too — both are combos
    # with no recent demand to be "missing", not actionable gone-silent channels.
    three_months_ago = last_complete_week - pd.DateOffset(months=3)
    missing = combos[
        ~(combos["Last Data Week"] >= last_complete_week)   # currently silent
        & (combos["Last Data Week"] >= three_months_ago)     # but sold within the past 3 months
    ].copy()
    if missing.empty:
        return pd.DataFrame(columns=MISSING_POS_COLS)

    # --- Build the report -----------------------------------------------------
    # Gap runs from the week AFTER the last data week through the last completed
    # week. Every surviving combo has a real (non-NaT) last data week within the
    # past 3 months, so no earliest-week fallback is needed.
    missing["Active in"] = missing["SKU"].map(lambda s: sku_active_in.get(s))
    # POS-then-Orders: the last data point's source is POS if it recorded any POS,
    # else Orders; Last Value is the quantity booked on that final data week.
    missing["Data Source"] = missing["_last_pos"].notna().map({True: "POS", False: "Orders"})
    missing["Last Value"] = (
        missing["_last_pos"].where(missing["_last_pos"].notna(), missing["_last_orders"])
        .round(0).astype("Int64")
    )
    missing["First Missing Week"] = missing["Last Data Week"] + pd.Timedelta(weeks=1)
    missing["Last Missing Week"] = last_complete_week
    missing["Missing Weeks"] = (
        (last_complete_week - missing["First Missing Week"]).dt.days // 7 + 1
    ).astype("Int64")

    out = missing[MISSING_POS_COLS]
    return out.sort_values(["SKU", "Customer"]).reset_index(drop=True)


class ExclusionResult(NamedTuple):
    """Outcome of ``apply_exclusions``: the filtered demand frame plus the two
    "excluded" tables and counts the dashboard surfaces in its own sections."""

    df: pd.DataFrame                    # demand frame with excluded rows removed
    inactive_df: pd.DataFrame           # active-SKU-in-wrong-region exclusions
    discontinued_df: pd.DataFrame       # discontinued/inactive with projections
    active_check_ran: bool              # Plytix had the 'Active in' columns
    disc_check_ran: bool                # Plytix had the 'SKU Status' column
    n_excluded_rows: int                # demand rows dropped by the active-in check
    excluded_counts_by_key: pd.Series   # per SKU||Customer dropped-row counts
    n_disc_rows: int                    # demand rows dropped as discontinued/inactive
    n_disc_skus: int                    # distinct SKUs dropped as discontinued/inactive


def apply_exclusions(df, plytix_df, P, anchors=None):
    """Drop SKUs that must never be forecast, mirroring dashboard.main() exactly.

    Two independent Plytix-driven filters, applied to the demand frame BEFORE
    forecasting so neither the dashboard nor the agent projects or flags them:

    1. Active-in region check: an *active* product forecast in a region it is
       not "Active in" (per Plytix) has those SKU x customer rows dropped.
    2. Discontinued/inactive drop: a SKU marked Discontinued/Inactive in Plytix,
       OR carrying a trailing '*' in the demand file, is dropped entirely (the
       status is SKU-level, so every row of the SKU goes).

    With no Plytix export both checks degrade to no-ops, except the trailing-'*'
    drop, which needs no Plytix. Returns an ``ExclusionResult``; callers do the
    logging so this stays free of any logging/Streamlit dependency.
    """
    active_sku_set, sku_active_in = compute_active_products(plytix_df)
    active_check_ran = active_sku_set is not None
    inactive_df = compute_inactive_projections(
        df, active_sku_set, sku_active_in, P, anchors=anchors
    )

    disc_status = compute_discontinued_products(plytix_df)
    disc_check_ran = disc_status is not None
    discontinued_df = compute_discontinued_projections(df, disc_status, P)

    # ----- Drop the active-in exclusions (per SKU x customer) --------------
    n_excluded_rows = 0
    excluded_counts_by_key = pd.Series(dtype="int64")
    if not inactive_df.empty:
        exclude_keys = {
            f"{str(s)}||{str(c)}"
            for s, c in zip(inactive_df["SKU"], inactive_df["Customer"])
        }
        key = df["SKU"].astype(str).str.rstrip("*") + "||" + df["Customer"].astype(str)
        drop_mask = key.isin(exclude_keys)
        n_excluded_rows = int(drop_mask.sum())
        # Per SKU||Customer demand-row counts, so a region-scoped excluded table
        # can report accurate row totals.
        excluded_counts_by_key = key[drop_mask].value_counts()
        if n_excluded_rows:
            df = df[~drop_mask].reset_index(drop=True)

    # ----- Drop discontinued/inactive SKUs entirely ------------------------
    # Two independent signals: a trailing '*' on the SKU code in the demand file,
    # and a Plytix 'SKU Status' of Discontinued/Inactive. Either one drops the
    # whole SKU (match at SKU level so every row goes, even rows omitting the '*').
    sku_raw = df["SKU"].astype(str)
    sku_base = sku_raw.str.rstrip("*")
    disc_bases = set(sku_base[sku_raw.str.endswith("*")])
    if disc_status:
        disc_bases |= set(disc_status)
    disc_mask = sku_base.isin(disc_bases)
    n_disc_rows = int(disc_mask.sum())
    n_disc_skus = 0
    if n_disc_rows:
        n_disc_skus = int(sku_base[disc_mask].nunique())
        df = df[~disc_mask].reset_index(drop=True)

    return ExclusionResult(
        df=df,
        inactive_df=inactive_df,
        discontinued_df=discontinued_df,
        active_check_ran=active_check_ran,
        disc_check_ran=disc_check_ran,
        n_excluded_rows=n_excluded_rows,
        excluded_counts_by_key=excluded_counts_by_key,
        n_disc_rows=n_disc_rows,
        n_disc_skus=n_disc_skus,
    )


# --------------------------------------------------------------------------- #
# Warehouse projection exports -> "missing future projections" table.          #
#                                                                              #
# Ported from active_missing_projections.py. This uses a DIFFERENT data source #
# than everything above: the warehouse projection exports (one wide grid per   #
# region, raw_inputs/warehouse_projections), NOT the demand-projection frame   #
# the dashboard forecasts. The demand file only carries SKU×customer combos    #
# that already have projections, so a *missing* projection is only visible in   #
# the warehouse grid, which lists every active-in SKU×customer×week with a NaN  #
# cell where no projection exists.                                             #
# --------------------------------------------------------------------------- #

# Region prefix on a warehouse filename (e.g. "AU_warehouse_projections_*.xlsx").
REGION_PREFIXES = ("AU", "CA", "EU", "JP", "US")

WAREHOUSE_DIRNAME = "raw_inputs/warehouse_projections"

# Long-format columns produced when a wide warehouse grid is melted.
WAREHOUSE_LONG_COLS = ["SKU", "Customer", "WeekDate", "Projection", "Region Code"]

MISSING_COLS = [
    "SKU", "Region Code", "Region", "Active in", "Customer",
    "First_WeekDate", "Last_WeekDate",
]


def _warehouse_dir(warehouse_dir=None):
    """Resolve the folder holding the warehouse projection exports.

    Honours WAREHOUSE_RAW_DIR if set; otherwise uses the standard location,
    resolved relative to the repo root when it is a relative path."""
    folder = warehouse_dir or os.environ.get("WAREHOUSE_RAW_DIR") or WAREHOUSE_DIRNAME
    if not os.path.isabs(folder):
        folder = os.path.join(HERE, folder)
    return folder


def warehouse_glob(warehouse_dir=None):
    """Glob matching every warehouse export in the warehouse folder."""
    return os.path.join(_warehouse_dir(warehouse_dir), "*.xlsx")


def discover_warehouse_files(warehouse_dir=None):
    """Return {snapshot_date: [paths]} for warehouse exports, newest date first.

    Each snapshot date normally has one file per region (AU/CA/EU/JP/US), so we
    group by the date embedded in the filename (files without a date land under
    "undated")."""
    groups = {}
    for path in glob.glob(warehouse_glob(warehouse_dir)):
        d = _date_from_name(path) or "undated"
        groups.setdefault(d, []).append(path)
    for paths in groups.values():
        paths.sort()
    return dict(sorted(groups.items(), reverse=True))


def _warehouse_region(name):
    """Region code (AU/CA/EU/JP/US) from a warehouse filename, or None."""
    base = os.path.basename(str(name))
    return next((p for p in REGION_PREFIXES if base.startswith(p)), None)


def _long_export_header_row(source):
    """Header row index of a long-format export, or None if the file is wide.

    The long PowerBI export ("export data" of the underlying table) carries a
    literal ``WeekDate`` column header within the first few rows (banner row,
    blank row, then headers). The legacy wide matrix export never does — its
    week dates are *column values* on the title row. That makes ``WeekDate``
    a reliable format discriminator.
    """
    probe = pd.read_excel(source, header=None, nrows=6)
    for i in range(min(len(probe), 4)):
        row = probe.iloc[i].map(lambda v: str(v).strip())
        if (row == "WeekDate").any():
            return i
    return None


def _warehouse_long_export_to_long(source, location, header_row):
    """Clean one long-format export (PowerBI table export, or the file written
    by extract_warehouse_projections.py) into the WAREHOUSE_LONG_COLS shape.

    Unlike the wide grid, a long file has no blank cells: a missing projection
    is an *absent row*. The frame returned here is therefore only the observed
    values — combine_warehouse_projections reconstructs the full
    pairs × weeks grid (reintroducing NaN = missing) across the snapshot.
    Explicit zero rows are kept as real values: a 0 cell in the wide grid
    rendered as 0, i.e. "has a projection", not "missing".

    Keys are normalized here (whitespace-stripped — GP CHAR columns are
    space-padded — and trailing '*' display markers dropped) so SKUs line up
    with the Plytix side, which compute_active_products strips the same way.
    """
    df = pd.read_excel(source, header=header_row)

    renames = {}
    for c in df.columns:
        s = str(c).strip()
        if "DisplaySKU" in s or s == "SKU":
            renames[c] = "SKU"
        elif s == "CUSTNMBR":
            renames[c] = "Customer"
        elif s == "WeekDate":
            renames[c] = "WeekDate"
        elif "Proj" in s:
            renames[c] = "Projection"
    df = df.rename(columns=renames)
    needed = ["SKU", "Customer", "WeekDate", "Projection"]
    missing_cols = [c for c in needed if c not in df.columns]
    if missing_cols:
        raise ValueError(
            f"Long-format warehouse export is missing column(s) {missing_cols} "
            f"(found: {list(df.columns)})"
        )
    df = df[needed].copy()

    df["SKU"] = df["SKU"].astype(str).str.strip().str.rstrip("*")
    df["Customer"] = df["Customer"].astype(str).str.strip()
    # Timestamps *before* any cross-file week union — a mixed wide/long
    # snapshot must not end up with string-vs-Timestamp duplicate week keys.
    df["WeekDate"] = pd.to_datetime(df["WeekDate"], errors="coerce")
    df["Projection"] = pd.to_numeric(df["Projection"], errors="coerce")

    # Footer notes / stray rows: no usable key -> not data.
    bad_key = (
        df["SKU"].isin(["", "nan", "None"])
        | df["Customer"].isin(["", "nan", "None"])
        | df["WeekDate"].isna()
    )
    df = df[~bad_key]

    long_df = df.sort_values(["SKU", "Customer", "WeekDate"]).reset_index(drop=True)
    long_df["Region Code"] = location
    long_df = long_df[WAREHOUSE_LONG_COLS]
    # Tag for combine_warehouse_projections: this frame still needs its
    # missing cells reconstructed (attrs survive because combine reads the
    # flag before doing anything to the frame).
    long_df.attrs["needs_grid_reconstruction"] = True
    return long_df, location


def warehouse_wide_to_long(source, name=None):
    """Clean one warehouse export into a long frame with a Region Code column.

    ``source`` is a filesystem path or a file-like object (e.g. a BytesIO of an
    uploaded workbook); ``name`` supplies the filename used to detect the region
    prefix (defaults to ``source`` when it is a path). Returns (long_df,
    location), or (None, None) if the name has no known region prefix.

    Despite the name (kept for compatibility), this now sniffs the layout and
    dispatches: legacy *wide* matrix exports go through the original
    unmerge-and-melt path below; *long* table exports (what PowerBI's
    "export data" produces today, and what extract_warehouse_projections.py
    writes from the data warehouse) go through _warehouse_long_export_to_long.
    Pointing the wide parser at a long file used to melt the banner row into
    garbage — silently emptying the missing-projections table.
    """
    name = name if name is not None else source
    location = _warehouse_region(name)
    if location is None:
        return None, None

    # Sniff the layout. ``source`` may be a BytesIO from an upload, which the
    # probe read consumes — rewind before and after so the real parse (and a
    # caller retry) starts from the top.
    if hasattr(source, "seek"):
        source.seek(0)
    header_row = _long_export_header_row(source)
    if hasattr(source, "seek"):
        source.seek(0)
    if header_row is not None:
        return _warehouse_long_export_to_long(source, location, header_row)

    # 1) Load with openpyxl so we can unmerge the SKU column and propagate its value
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]

    for merged_range in list(ws.merged_cells.ranges):
        top_left_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in ws.iter_rows(
            min_row=merged_range.min_row, max_row=merged_range.max_row,
            min_col=merged_range.min_col, max_col=merged_range.max_col,
        ):
            for cell in row:
                cell.value = top_left_value

    raw = pd.DataFrame(ws.values)

    # 2) Row 0 = title row with week dates starting at column index 2
    #    Row 1 = "SKU" / "Customer" / "Proj..." labels
    #    Row 2+ = actual data, until a blank row (end of data / start of footer)
    week_dates = raw.iloc[0, 2:].tolist()
    data = raw.iloc[2:].reset_index(drop=True)

    # Stop at the first row with no customer AND no projection values at all
    # (cuts off the blank separator row + the "Applied filters..." footer note).
    def is_end_row(row):
        return pd.isna(row[1]) and row[2:].isna().all()

    end_idx = len(data)
    for i, row in data.iterrows():
        if is_end_row(row):
            end_idx = i
            break
    data = data.iloc[:end_idx]

    # Forward-fill SKU since it only appears on the first row of each merged block
    data[0] = data[0].ffill()

    # 3) Melt wide -> long
    data.columns = ["SKU", "Customer"] + week_dates
    long_df = data.melt(
        id_vars=["SKU", "Customer"],
        value_vars=week_dates,
        var_name="WeekDate",
        value_name="Projection",
    )
    long_df = long_df.sort_values(["SKU", "Customer", "WeekDate"]).reset_index(drop=True)
    long_df["Region Code"] = location
    return long_df, location


def _reconstruct_missing_cells(long_df, weeks):
    """Rebuild the full pairs × weeks grid for one long-format export.

    A long export only lists observed values, so a missing projection is an
    absent row. Recreate what the wide matrix showed: every (SKU, Customer)
    pair in the file gets a cell for every week in ``weeks``, NaN where the
    file had no row — NaN being exactly the "missing" signal downstream.
    Duplicate keys (e.g. a starred and unstarred variant of the same SKU that
    normalization collapsed) are summed.
    """
    pairs = long_df[["SKU", "Customer"]].drop_duplicates()
    values = (
        long_df.dropna(subset=["Projection"])
        .groupby(["SKU", "Customer", "WeekDate"], as_index=False)["Projection"]
        .sum()
    )
    grid = pairs.merge(pd.DataFrame({"WeekDate": weeks}), how="cross")
    grid = grid.merge(values, on=["SKU", "Customer", "WeekDate"], how="left")
    grid["Region Code"] = long_df["Region Code"].iloc[0]
    return grid[WAREHOUSE_LONG_COLS]


def combine_warehouse_projections(sources):
    """Clean and concatenate many warehouse exports into one long frame.

    ``sources`` is an iterable of (source, name) pairs, where ``source`` is a
    path or file-like object and ``name`` is the filename (for region
    detection). Files whose name lacks a region prefix are skipped. Returns an
    empty (but correctly-columned) frame when nothing usable is provided.

    Long-format exports get their missing cells reconstructed against the
    union of week dates across ALL long files in the snapshot — per-file weeks
    are not enough: a week in which a small region (e.g. JP) has no
    projections at all vanishes from that region's file entirely, and its
    missing cells would otherwise never be flagged. The big regions (US/EU)
    anchor the union. Wide-grid frames already carry their NaN cells and pass
    through untouched.
    """
    frames, to_reconstruct = [], []
    for source, name in sources:
        long_df, _ = warehouse_wide_to_long(source, name)
        if long_df is None:
            continue
        if long_df.attrs.get("needs_grid_reconstruction") and not long_df.empty:
            to_reconstruct.append(long_df)
        elif not long_df.attrs.get("needs_grid_reconstruction"):
            frames.append(long_df)
        # empty long-format frame (headers-only region file): nothing to add.

    if to_reconstruct:
        weeks = sorted(
            pd.concat([f["WeekDate"] for f in to_reconstruct]).dropna().unique()
        )
        frames.extend(_reconstruct_missing_cells(f, weeks) for f in to_reconstruct)

    if not frames:
        return pd.DataFrame(columns=WAREHOUSE_LONG_COLS)
    return pd.concat(frames, ignore_index=True)


def compute_missing_projections(projections, plytix_df, df, P):
    """Active SKUs missing future projections in regions they ARE 'Active in'.

    Mirrors active_missing_projections.py: from the combined warehouse grid
    (``projections``), keep the NaN projection cells, intersect with active
    products in a region that IS in their Plytix 'Active in' list, restrict to
    the coming 15-week window, and roll up to one row per SKU×Region Code×customer
    with the first/last missing week. A Region label (via the pipeline's
    region_for_group) is added from ``df``'s customer groupings so a
    by-customer-group view can scope to its own region, matching the sibling
    excluded tables.

    Returns a table (columns = MISSING_COLS), empty if none or inputs missing.
    """
    active_sku_set, sku_active_in = compute_active_products(plytix_df)
    if not active_sku_set or projections is None or projections.empty:
        return pd.DataFrame(columns=MISSING_COLS)

    # Missing projection cells only.
    missing = projections[projections["Projection"].isna()].copy()
    if missing.empty:
        return pd.DataFrame(columns=MISSING_COLS)

    # Active (SKU, Region Code) pairs, restricted to the warehouse regions.
    pairs = [
        (sku, loc)
        for sku in active_sku_set
        for loc in _active_in_list(sku_active_in, sku)
        if loc in WAREHOUSE_REGIONS
    ]
    if not pairs:
        return pd.DataFrame(columns=MISSING_COLS)
    active_pairs = pd.DataFrame(pairs, columns=["SKU", "Region Code"]).drop_duplicates()

    m = active_pairs.merge(missing, on=["SKU", "Region Code"], how="inner")
    if m.empty:
        return pd.DataFrame(columns=MISSING_COLS)

    # Coming 15-week window (today < WeekDate <= today + 15 weeks), per notebook.
    m["WeekDate"] = pd.to_datetime(m["WeekDate"])
    today = pd.Timestamp.today().normalize()
    cutoff = today + pd.Timedelta(weeks=15)
    m = m[(m["WeekDate"] > today) & (m["WeekDate"] <= cutoff)]
    if m.empty:
        return pd.DataFrame(columns=MISSING_COLS)

    g = m.groupby(["SKU", "Region Code", "Customer"], as_index=False).agg(
        First_WeekDate=("WeekDate", "min"),
        Last_WeekDate=("WeekDate", "max"),
    )
    # Full (un-exploded) 'Active in' string per SKU, e.g. "US,CA,UK,SG,EU,AU".
    g["Active in"] = g["SKU"].map(lambda s: sku_active_in.get(s))

    # Region label from the region code, consistent with the sibling tables
    # (e.g. code "JP" -> "JP (NETDEPOT)"). Fall back to the raw code when a
    # region has no customer group in the demand frame.
    code_to_label = {}
    if df is not None and not df.empty:
        for grp in df["Customer Grouping"].dropna().unique():
            code = _region_code(P, grp)
            if code is not None:
                code_to_label[code] = P.region_for_group(grp)
    g["Region"] = g["Region Code"].map(code_to_label).fillna(g["Region Code"])

    return g[MISSING_COLS].sort_values(
        ["SKU", "Region Code", "Customer"]
    ).reset_index(drop=True)
