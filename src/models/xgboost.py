"""
Updates projections for all SKUs based on combined demand projections for all companies.

Uses *all* completed weeks of historical POS data to calculate a 15-week
projection. The in-progress week is excluded so a partially-elapsed week's POS
never drags the model (see week_anchors). The amount of history is controlled
by LOOKBACK_WEEKS (None = all history; set an int to cap it).

This is the XGBoost (gradient-boosted trees) counterpart of the
exponential-smoothing pipeline. Instead of smoothing a level and a trend, it
learns a supervised mapping from each week's recent history to the next week's
demand, pooled across every SKU in the group:

    features(SKU, week t) = [ lag_1 .. lag_N, rolling means, weeks since first
                              sale, position in history, week-of-year sin/cos ]
    target(SKU, week t)   = demand at week t   (scaled by the SKU's own mean,
                                                so big and small SKUs share one
                                                model without the big ones
                                                dominating the loss)

One model is trained per group (per fit_xgboost call) on all of its SKUs'
week-rows together, then each SKU is forecast 15 weeks ahead RECURSIVELY: the
week-1 prediction is appended to the history to build week-2's features, and so
on. Pooling matters because individual SKU series here are short (often 10-40
weeks) -- far too little to train a tree ensemble per SKU -- while the pooled
group easily has hundreds or thousands of training rows.

Hyperparameters live in XGB_PARAMS / N_LAGS / ROLL_WINDOWS below. They are
deliberately NOT exposed to the dashboard: fit_xgboost's signature carries no
alpha/beta/phi, so the dashboard's smoothing sliders hide themselves (it
inspects the signature), while ``min_weeks_for_trend`` IS accepted so the
min-weeks slider keeps working.

Short-history handling matches the exponential-smoothing pipeline: SKUs with
fewer completed weeks than MIN_WEEKS_FOR_TREND are forecast FLAT at their mean
(no model), which stops one-off orders (e.g. "Others - AU") from producing
runaway projections. If a group is too small to train on at all (fewer than
MIN_TRAIN_ROWS pooled samples), every SKU falls back to a flat-mean forecast.

Before fitting, each SKU's series is cleansed of one-off promo spikes (e.g.
Amazon Prime Day) and stockout dips exactly as in the smoothing pipeline:
abnormal weeks are detected automatically (rolling median +/- OUTLIER_K MADs)
and/or named explicitly (PROMO_WEEKS), then replaced with a local baseline.
Every cleaned week is written to a ``cleaned_outliers_<date>.txt`` record.

Promos also lift future demand, so any PROMO_WEEKS date that lands inside the
15-week horizon has its projection scaled up by a promo uplift factor (fixed,
or estimated per SKU from its own past promos). Uplifted weeks are recorded to
a ``promo_uplift_<date>.txt`` file (see PROMO_UPLIFT).

Two kinds of output are produced:
    1) Per-group     : one file per Customer Grouping (see COMBINED_GROUPING),
                       summing POS across the customers that make up the group.
    2) All-customers : a single file forecasting each SKU "as a whole", built by
                       summing POS (and the existing Projection) across every
                       customer for each SKU-week before fitting.
"""

import os
import re
import glob
import traceback
import numpy as np
import pandas as pd
from pathlib import Path

# --- XGBoost model parameters ----------------------------------------------- #
# One gradient-boosted-tree model is trained per group, pooled across its SKUs.
# These are module-level constants on purpose: the dashboard does NOT expose
# them (its smoothing sliders key off alpha/beta/phi in the fit signature,
# which this pipeline deliberately omits). Tune them here.
XGB_PARAMS = dict(
    n_estimators=300,        # boosting rounds
    learning_rate=0.05,      # shrinkage per round (lower = smoother, needs more rounds)
    max_depth=4,             # tree depth -- shallow trees generalise better on small data
    min_child_weight=5,      # min sum of instance weight per leaf (regularisation)
    subsample=0.9,           # row subsampling per tree
    colsample_bytree=0.9,    # feature subsampling per tree
    reg_lambda=1.0,          # L2 regularisation
    objective="reg:squarederror",
    random_state=42,
    n_jobs=-1,
    verbosity=0,
)

N_LAGS = 6                   # lagged weeks fed as features (lag_1 .. lag_6)
ROLL_WINDOWS = (4, 8)        # trailing rolling-mean windows added as features
MIN_TRAIN_ROWS = 30          # pooled samples needed to train at all; below this
                             # every SKU in the group is forecast flat at its mean

# Backwards-compatibility caption for the dashboard (it shows this string in the
# header instead of the smoothing-specific blurb).
DASHBOARD_CAPTION = (
    "15-week XGBoost (gradient-boosted trees) forecast from the historical "
    "demand window (POS where available, else Orders). One model per view, "
    "pooled across its SKUs; hyperparameters are fixed in the pipeline file."
)

# Minimum completed weeks of history required before the model is used for a
# SKU. A pattern learned from only a week or two is essentially noise -- so a
# single spike in a 2-week series would otherwise become a runaway ramp (e.g.
# "Others - AU"). Series shorter than this are instead forecast FLAT at their
# mean, which is far more sensible when there isn't enough history.
# Raise it to be more conservative (more series held flat); 2 disables the guard.
MIN_WEEKS_FOR_TREND = 4

# --- History window --------------------------------------------------------- #
# How many of the most-recent *completed* weeks to fit the model on.
#   None -> use ALL available history (every completed week up to last week)
#   int  -> use only the most recent N completed weeks
# Tree models benefit from more rows, so all history is the sensible default.
# The in-progress week is always excluded (see week_anchors).
LOOKBACK_WEEKS = None

# When LOOKBACK_WEEKS is None the lower bound is set to this many years before
# the run date -- effectively "all history" for our data, while still guarding
# against a stray ancient row dragging the fit. See week_anchors.
HISTORY_YEARS = 3

# Display label for the descriptive-average column. Reflects LOOKBACK_WEEKS so
# the header never claims "8 Week" when it's actually averaging all history
# (or some other window) -- see DISPLAY_NAMES / SUMMARY_COLUMNS below.
AVG_COL_LABEL = (
    "All-History POS/Orders Average"
    if LOOKBACK_WEEKS is None
    else f"{LOOKBACK_WEEKS} Week POS/Orders Average"
)

# --- Intermittent / lumpy demand -------------------------------------------- #
# Same rationale as the smoothing pipeline: a missing completed week almost
# always means "nothing sold" = 0, so each SKU's series is reindexed to every
# completed week from its FIRST observation through the last completed week and
# the gaps filled with 0 before cleansing/fitting. Leading weeks BEFORE a SKU's
# first-ever order are NOT zero-filled. Set False to restore observed-weeks-only.
FILL_GAPS_WITH_ZERO = True

# --- Outlier / promo cleansing ---------------------------------------------- #
# One-off spikes (e.g. Amazon Prime Day) and dips (e.g. stockouts) distort the
# lag features the trees learn from, so abnormal weeks are replaced with a local
# baseline (the median of nearby normal weeks) before fitting. Each cleaned week
# is recorded to a .txt audit file (see __main__).
#
# Two sources of flags are unioned:
#   1. Manual  -- PROMO_WEEKS below (always applied; you know these are promos).
#   2. Auto    -- a week more than OUTLIER_K MADs from a centred rolling median
#                 (symmetric, so it catches both spikes and stockout dips).
CLEANSE_OUTLIERS = True       # master switch for the automatic MAD detector
OUTLIER_K = 3.0               # MADs from the local median to count as an outlier
OUTLIER_WINDOW = 7            # weeks in the rolling median/MAD (odd -> centred)
OUTLIER_MIN_WEEKS = 5         # don't auto-detect on series shorter than this

# Manual promo calendar: weeks you KNOW are abnormal. Any date inside the week
# works -- it is snapped to that week's start (the Sunday WeekDate) before
# matching. Always cleansed, regardless of CLEANSE_OUTLIERS, recorded as "manual".
PROMO_WEEKS = [
    # "2026-06-26",   # Amazon Prime Day week
    # # "2026-07-?"   # Friends and Family
    # "2026-08-15",   # Back to school shopping
    # "2026-11-27",   # Black Friday week
    # "2026-11-30",   # Cyber Monday
    # "2026-12-1",    # Christmas shopping
    # "2026-12-8",    # Christmas shopping
    # "2026-12-15"     # Christmas shopping
]

# --- Future promo uplift ----------------------------------------------------- #
# Promos lift sales, so any PROMO_WEEKS date that lands inside the 15-week
# forecast horizon has its projection scaled UP. (Historical copies of those
# weeks are still cleansed for the baseline fit above -- this only re-adds the
# expected lift onto FUTURE promo weeks: baseline + uplift.)
#
# PROMO_UPLIFT controls the peak multiplier on a promo week:
#   a number (e.g. 1.25) -> fixed: multiply promo-week projections by this factor
#   "auto"               -> estimate the factor per SKU from its OWN historical
#                           promo weeks, falling back to PROMO_UPLIFT_DEFAULT
PROMO_UPLIFT = 1.25            # "a bit higher"; set "auto" for a data-driven lift
PROMO_UPLIFT_DEFAULT = 1.25    # fallback factor for "auto" when no SKU history
PROMO_UPLIFT_MAX = 4.0         # clamp so a noisy estimate can't explode a week
PROMO_HALO_WEEKS = 0           # also lift this many weeks either side (0 = off),
                               # tapering linearly to the edge of the halo

RAW_INPUTS_FOLDER = "raw_inputs/demand_projections"
INPUT_GLOB = os.path.join(RAW_INPUTS_FOLDER, "all_demand_projections_*.xlsx")

# List-price workbook (SKU -> List Price USD), used to value the change in
# forecast as a revenue risk. See load_list_prices().
LIST_PRICE_GLOB = os.path.join("raw_inputs/list_prices", "list_prices_*.xlsx")

CUSTOMERS_TO_IGNORE = [
    "Others - HK", "Others - KR", "Others - MALDIV", "Others - MX",
    "Others - UAE", "Others - UK", "Others - ZZ",
]

# Maps each raw CUSTNMBR onto a consolidated customer group. Several customers
# fold into one group (e.g. the three Amazon-DC channels), so forecasts are built
# per group rather than per raw customer. Customers not listed here fall back to
# their own name as a single-member group.
US_GROUPING = {
    'AMAZON-DC': 'AMAZON-DC',
    'AMAZON-DS': 'AMAZON-DC',
    'MARVAL-FBM': 'AMAZON-DC',
    'TARGET-HQ': 'TARGET-HQ',
    'TARGET-DS': 'TARGET-DS',
    'TARGET-PLUS': 'TARGET-DS',
    'HOMDEP': 'HOMDEP',
    'LOWES-HQ': 'LOWES-HQ',
    'LOWES-SOS': 'LOWES-SOS',
    'Web Sales - US': 'Web Sales + Warranty US',
    'Warranty - US': 'Web Sales + Warranty US',
    'CONSTO': 'CONSTO',
    'COSTCO.COM': 'COSTCO.COM',
    'ULINE': 'ULINE',
    'AAFES': 'AAFES',
    'STAPADV.COM': 'STAPADV.COM',
    'STAPLES.COM': 'STAPADV.COM',
    'DILLARDS': 'DILLARDS',
    'NORSTR': 'NORSTR',
    'NORCOM-DS': 'NORSTR',
    'NORCOM': 'NORSTR',
    'Others - US': 'Others - US'
}

EU_GROUPING = {
    'AMAZON-EU': 'AMAZON-EU',
    'Web Sales - EU': 'Web Sales + Warranty EU',
    'CASTOR': 'CASTOR',
    'EBREUNINGER': 'EBREUNINGER',
    'FSKUSTERMANN': 'FSKUSTERMANN',
    'GALLAF': 'GALLAF',
    'LERMER-FR': 'LERMER-FR',
    'SANIKAL-KG': 'SANIKAL-KG',
    'Others - EU': 'Others - EU',
    'Others - IT': 'Others - EU'
}

AU_GROUPING = {
    'Web Sales - AU': 'Web Sales - AU',
    'Others - AU': 'Others - AU'
}

CA_GROUPING = {
    'AMAZON-DCCA': 'AMAZON-DCCA',
    'CANTIR': 'CANTIR',
    'COSTCO-CAN': 'COSTCO-CAN',
    'HOMEDEP-CA.COM': 'HOMEDEP-CA.COM',
    'HOMEDEPOT-CA': 'HOMEDEPOT-CA',
    'Others - CA': 'Others - CA',
    'RONA-HQ': 'RONA-HQ',
    'Web Sales - CA': 'Web Sales - CA',
    'WINMER': 'WINMER'
}

JP_GROUPING = {
    'Web Sales - JP': 'Web Sales - JP',
    'Others - JP': 'Others - JP'
}

COMBINED_GROUPING = {**US_GROUPING, **EU_GROUPING, **AU_GROUPING, **CA_GROUPING, **JP_GROUPING}

# Group labels belonging to each region (the *values* of the maps above), used to
# file each group's output under US / EU / NON-US-EU. A group that came from
# neither map (an unmapped customer kept as its own group) lands in NON-US-EU.
US_GROUPS = set(US_GROUPING.values())
EU_GROUPS = set(EU_GROUPING.values())
AU_GROUPS = set(AU_GROUPING.values())
CA_GROUPS = set(CA_GROUPING.values())
JP_GROUPS = set(JP_GROUPING.values())

def region_for_group(group):
    if group in US_GROUPS:
        return "US (LBC+NJ)"
    elif group in EU_GROUPS:
        return "EU (SH-CTS)"
    elif group in AU_GROUPS:
        return "AU (ACR)"
    elif group in CA_GROUPS:
        return "CA (YYZ5)"
    elif group in JP_GROUPS:
        return "JP (NETDEPOT)"
    return "Other"

# Label used for the all-customers (combined) output.
ALL_SKUS_LABEL = "ALL SKUS"

# Internal (snake_case) column names -> display names used in the output sheets.
DISPLAY_NAMES = {
    "weeks_with_data": "Weeks with data",
    "outlier_weeks_cleaned": "Outlier Weeks Cleaned",
    "promo_weeks_uplifted": "Promo Weeks Uplifted",
    "8_week_pos_avg": AVG_COL_LABEL,
    "initial_projection_avg": "Initial Projection Average",
    "updated_projection_avg": "Updated Projection Average",
    "projection_difference": "Projection Difference",
    "list_price_usd": "List Price (USD)",
    "revenue_risk_usd": "Revenue Risk (USD)",
}

# Final column order for every summary sheet
SUMMARY_COLUMNS = [
    "SKU",
    "Description",
    "Customer Grouping",
    "Data Source",
    "Weeks with data",
    "Outlier Weeks Cleaned",
    "Promo Weeks Uplifted",
    AVG_COL_LABEL,
    "Initial Projection Average",
    "Updated Projection Average",
    "Projection Difference",
    "List Price (USD)",
    "Revenue Risk (USD)",
]


# --------------------------------------------------------------------------- #
# Model backend: XGBoost, with an sklearn fallback so the dashboard still     #
# runs on a host without the xgboost wheel (e.g. a fresh Streamlit Cloud env) #
# --------------------------------------------------------------------------- #
def _make_regressor():
    """Return an unfitted boosted-tree regressor.

    Prefers xgboost.XGBRegressor with XGB_PARAMS. If xgboost is not installed,
    falls back to sklearn's HistGradientBoostingRegressor (also NaN-tolerant,
    also gradient-boosted trees) so the pipeline keeps working; a note is
    printed so the substitution is visible.
    """
    try:
        from xgboost import XGBRegressor
        return XGBRegressor(**XGB_PARAMS)
    except ImportError:
        from sklearn.ensemble import HistGradientBoostingRegressor
        print(
            "xgboost is not installed -- falling back to sklearn's "
            "HistGradientBoostingRegressor (pip install xgboost to use XGBoost)."
        )
        return HistGradientBoostingRegressor(
            max_iter=XGB_PARAMS.get("n_estimators", 300),
            learning_rate=XGB_PARAMS.get("learning_rate", 0.05),
            max_depth=XGB_PARAMS.get("max_depth", 4),
            l2_regularization=XGB_PARAMS.get("reg_lambda", 1.0),
            random_state=XGB_PARAMS.get("random_state", 42),
        )


def resolve_input_file():
    """Pick the most recent raw data file and anchor the forecast to its snapshot date.

    Anchoring TODAY to the date in the filename (rather than the wall clock) keeps
    the historical lookback and the 15-week forecast aligned with the data
    snapshot, even if the script is run a day or two after the file was generated.
    """
    candidates = []
    for path in glob.glob(INPUT_GLOB):
        m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
        if m:
            candidates.append((m.group(1), path))

    if not candidates:
        raise FileNotFoundError(
            f"No input files matching {INPUT_GLOB}. "
            f"Expected e.g. {RAW_INPUTS_FOLDER}/all_demand_projections_YYYY-MM-DD.xlsx"
        )

    today_str, input_file = max(candidates)  # latest date wins
    return input_file, today_str, pd.Timestamp(today_str)


def load_list_prices(path=None):
    """Load a SKU -> List Price (USD) lookup from a price workbook.

    If ``path`` is given, that workbook is read directly. Otherwise the most
    recently modified file matching LIST_PRICE_GLOB is used. The sheet carries
    two columns, "SKU" and "List Price USD". SKUs with a blank price are dropped
    from the lookup so they map to NaN downstream -- their revenue risk is left
    blank rather than treated as $0 (an unknown price is not zero risk).

    Returns a pandas Series of price indexed by SKU (str), or None if no price
    file is found (in which case revenue risk is left blank for every SKU).
    """
    if path is None:
        candidates = glob.glob(LIST_PRICE_GLOB)
        if not candidates:
            print(
                f"No list-price file matching {LIST_PRICE_GLOB}; "
                f"revenue risk will be left blank.\n"
            )
            return None
        path = max(candidates, key=os.path.getmtime)  # most recently modified

    prices = pd.read_excel(path)
    prices.columns = [str(c).strip() for c in prices.columns]
    prices = prices[["SKU", "List Price USD"]].dropna(subset=["SKU"])
    prices["SKU"] = prices["SKU"].astype(str).str.strip()
    prices["List Price USD"] = pd.to_numeric(prices["List Price USD"], errors="coerce")
    prices = prices.dropna(subset=["List Price USD"]).drop_duplicates("SKU", keep="last")
    lookup = prices.set_index("SKU")["List Price USD"]

    print(f"Loaded {len(lookup)} list prices from {os.path.basename(path)}\n")
    return lookup


def week_anchors(today):
    """Resolve the week boundaries for the forecast.

    WeekDate is the Sunday that *starts* each 7-day week, so the week labelled
    W covers [W, W+6]. A week is only "completely over" once today is past its
    Saturday. To avoid feeding a partially-elapsed week's POS into the model,
    the historical (training) window ends at the last fully-completed week.

    The forecast, however, STARTS at the current in-progress week rather than
    skipping to the next full week: that week is deliberately kept OUT of the
    training history (its partial POS would distort the fit), but it still
    needs a projection -- otherwise it would fall into a gap, plotted as
    neither an actual nor a forecast. Step h = 1 of the 15-step horizon is
    exactly this in-progress week.

    The window *start* depends on LOOKBACK_WEEKS: with the default ``None`` the
    lower bound is HISTORY_YEARS years before the run date (i.e. all available
    history in practice); set LOOKBACK_WEEKS to an int to use only that many
    most-recent completed weeks instead.

    Returns (lookback_start, last_complete_week, first_forecast_week).
    """
    days_since_sunday = (today.weekday() + 1) % 7          # Sun=0 ... Sat=6
    current_week_start = today - pd.Timedelta(days=days_since_sunday)
    last_complete_week = current_week_start - pd.Timedelta(weeks=1)
    if LOOKBACK_WEEKS is None:
        lookback_start = today - pd.DateOffset(years=HISTORY_YEARS)  # all history
    else:
        # N weeks inclusive -> step back N-1 from the last completed week.
        lookback_start = last_complete_week - pd.Timedelta(weeks=LOOKBACK_WEEKS - 1)
    first_forecast_week = current_week_start
    return lookback_start, last_complete_week, first_forecast_week


def aggregate_to_sku_week(df):
    """Collapse rows to one row per SKU-week, summing POS and Projection.

    For a single customer this is an identity (the data already has one row per
    SKU-week), so the per-customer logic is unchanged. For the all-customers view
    it sums each SKU's POS / Projection across every CUSTNMBR, giving the SKU's
    total demand. ``min_count=1`` keeps an all-NaN group as NaN (no data) rather
    than collapsing it to a real 0.
    """
    grp = df.groupby(["SKU", "WeekDate"])
    pos = grp["POS"].sum(min_count=1)
    orders = grp["Orders"].sum(min_count=1)
    proj = grp["Projection"].sum(min_count=1)
    agg = pd.concat([pos, orders, proj], axis=1).reset_index()

    # Description is consistent per SKU; attach the first non-null one.
    desc = df.dropna(subset=["Description"]).groupby("SKU")["Description"].first()
    agg["Description"] = agg["SKU"].map(desc)
    return agg


def top_volume_customers(df, today, top_n=3):
    """For each SKU, rank the customer groups driving the most volume.

    Uses the same completed-week window the forecast is built from (all history
    by default; see LOOKBACK_WEEKS). Volume is measured with POS where the SKU
    has any, otherwise with Orders (mirroring the forecast's POS-then-Orders
    fallback), summed per Customer Grouping and turned into a share of the SKU
    total. The top ``top_n`` are returned as one string, e.g.
    "AMAZON-DC (61%); Web Sales + Warranty US (31%); AMAZON-EU (3%)".

    Returns a DataFrame [SKU, Top Volume Customer Groups].
    """
    lookback_start, last_complete_week, _ = week_anchors(today)
    win = df[
        (df["WeekDate"] >= lookback_start)
        & (df["WeekDate"] <= last_complete_week)
        & (df["POS"].notna() | df["Orders"].notna())
        & ~df["SKU"].astype(str).str.endswith("*")
    ].copy()

    # Per SKU: rank by POS volume if the SKU has any POS, else by Orders volume.
    has_pos = win.groupby("SKU")["POS"].transform(lambda s: s.notna().any())
    win["vol"] = np.where(has_pos, win["POS"], win["Orders"])

    by_grp = win.groupby(["SKU", "Customer Grouping"], as_index=False)["vol"].sum()
    by_grp = by_grp[by_grp["vol"] > 0]
    by_grp["sku_total"] = by_grp.groupby("SKU")["vol"].transform("sum")
    by_grp = by_grp[by_grp["sku_total"] > 0]
    by_grp["share"] = by_grp["vol"] / by_grp["sku_total"]
    by_grp = by_grp.sort_values(["SKU", "vol"], ascending=[True, False])

    rows = []
    for sku, grp in by_grp.groupby("SKU"):
        head = grp.head(top_n)
        label = "; ".join(
            f"{name} ({share * 100:.0f}%)"
            for name, share in zip(head["Customer Grouping"], head["share"])
        )
        rows.append({"SKU": sku, "Top Volume Customer Groups": label})

    return pd.DataFrame(rows, columns=["SKU", "Top Volume Customer Groups"])


def _week_start(date):
    """Snap any date to the Sunday that starts its week (matching WeekDate)."""
    ts = pd.Timestamp(date).normalize()
    return ts - pd.Timedelta(days=(ts.weekday() + 1) % 7)


def _mad(a):
    """Median absolute deviation of a 1-D array (ignoring NaNs)."""
    a = np.asarray(a, dtype="float64")
    a = a[~np.isnan(a)]
    if a.size == 0:
        return np.nan
    return np.median(np.abs(a - np.median(a)))


def cleanse_series(week_dates, y, promo_week_starts=None, detect=None,
                   k=OUTLIER_K, window=OUTLIER_WINDOW, min_weeks=OUTLIER_MIN_WEEKS):
    """Replace promo / outlier weeks with a local baseline before fitting.

    Spikes (Prime Day) and dips (stockouts) corrupt the lag features the trees
    learn from. This returns a cleaned copy of ``y`` in which abnormal weeks are
    swapped for the median of nearby *normal* weeks, so the model sees
    underlying demand rather than the event.

    Weeks are flagged from two sources, unioned:
      * Manual -- WeekDate falls in ``promo_week_starts`` (defaults to the module
        PROMO_WEEKS); always applied, recorded as method "manual".
      * Auto   -- |y - rolling_median| > k * 1.4826 * rolling_MAD over a centred
        ``window``-week window (1.4826 scales MAD to a normal std, so k is ~k-sigma).
        Skipped when ``detect`` is False or the series is shorter than ``min_weeks``.
        Recorded as method "auto".

    Detection uses the raw series (the median/MAD are robust to a minority of
    outliers); replacement uses only non-flagged neighbours so a multi-week promo
    can't baseline itself. Returns (cleaned_y, flags, method) as numpy arrays the
    same length as ``y`` (``method`` holds "manual"/"auto"/"").
    """
    y = np.asarray(y, dtype="float64")
    n = len(y)
    flags = np.zeros(n, dtype=bool)
    method = np.array([""] * n, dtype=object)
    if n == 0:
        return y.copy(), flags, method

    if detect is None:
        detect = CLEANSE_OUTLIERS
    if promo_week_starts is None:
        promo_week_starts = PROMO_WEEKS

    # 1) Manual promo weeks (always applied).
    promo_set = {_week_start(d) for d in promo_week_starts} if promo_week_starts else set()
    if promo_set:
        wk = pd.DatetimeIndex(pd.to_datetime(week_dates)).normalize()
        wk_start = wk - pd.to_timedelta((wk.weekday + 1) % 7, unit="D")
        manual = np.array([ws in promo_set for ws in wk_start])
        flags |= manual
        method[manual] = "manual"

    # 2) Automatic MAD detection on the raw series.
    if detect and n >= min_weeks:
        s = pd.Series(y)
        med = s.rolling(window, center=True, min_periods=3).median().to_numpy()
        mad = s.rolling(window, center=True, min_periods=3).apply(_mad, raw=True).to_numpy()
        med = np.where(np.isnan(med), np.nanmedian(y), med)
        global_mad = _mad(y)
        scale = np.where((np.isnan(mad)) | (mad == 0), global_mad, mad)
        with np.errstate(invalid="ignore"):
            auto = (scale > 0) & (np.abs(y - med) > k * 1.4826 * scale)
        newly = auto & ~flags
        flags |= auto
        method[newly] = "auto"

    if not flags.any():
        return y.copy(), flags, method

    # Replace each flagged week with the median of nearby NON-flagged weeks.
    half = max(window // 2, 1)
    cleaned = y.copy()
    global_baseline = np.median(y[~flags]) if (~flags).any() else np.nan
    for i in np.where(flags)[0]:
        lo, hi = max(0, i - half), min(n, i + half + 1)
        local = [y[j] for j in range(lo, hi) if not flags[j]]
        if local:
            cleaned[i] = float(np.median(local))
        elif not np.isnan(global_baseline):
            cleaned[i] = float(global_baseline)
        # else: every week flagged -> leave the original value untouched
    return cleaned, flags, method


def estimate_promo_uplift(y_raw, baseline, method):
    """Per-SKU promo uplift factor = mean(raw / baseline) over its promo weeks.

    Uses only manually-flagged promo weeks (``method == "manual"``) with a
    positive baseline -- i.e. how much the SKU's real sales exceeded the cleansed
    baseline on past promos. Returns None when there is nothing to estimate from
    (no historical promo weeks for this SKU in the fitting window).
    """
    ratios = [
        y_raw[j] / baseline[j]
        for j in range(len(method))
        if method[j] == "manual" and baseline[j] > 0
    ]
    if not ratios:
        return None
    return float(np.mean(ratios))


def promo_week_multipliers(forecast_weeks, factor, promo_set, halo=PROMO_HALO_WEEKS):
    """Per-forecast-week uplift multipliers (1.0 = no change).

    ``forecast_weeks`` is the sequence of 15 projected week-start dates,
    ``promo_set`` a set of promo week-start Timestamps (see ``_week_start``), and
    ``factor`` the peak multiplier applied on a promo week. With ``halo`` > 0 the
    ``halo`` weeks on each side also lift, tapering linearly to the halo edge;
    overlapping halos keep the strongest lift. Returns a numpy array.
    """
    n = len(forecast_weeks)
    mult = np.ones(n, dtype="float64")
    if factor <= 1.0 or not promo_set:
        return mult

    fw = pd.DatetimeIndex(pd.to_datetime(forecast_weeks)).normalize()
    is_promo = np.array([ts in promo_set for ts in fw])
    for i in np.where(is_promo)[0]:
        for d in range(-halo, halo + 1):
            j = i + d
            if 0 <= j < n:
                taper = 1.0 if d == 0 else (halo + 1 - abs(d)) / (halo + 1)
                mult[j] = max(mult[j], 1.0 + (factor - 1.0) * taper)
    return mult


# --------------------------------------------------------------------------- #
# XGBoost feature engineering + forecasting                                   #
# --------------------------------------------------------------------------- #
def _feature_row(z_hist, t, n_total, week_date):
    """Feature vector for one (SKU, week) sample.

    ``z_hist`` is the SKU's scale-normalised demand up to (but excluding) week
    index ``t``; ``n_total`` its full training length; ``week_date`` the
    calendar Sunday of the target week. Missing lags (early in a series) are
    left as NaN -- both XGBoost and HistGradientBoosting handle NaN natively,
    so short histories still produce usable rows without imputation.

    Features:
      * lag_1 .. lag_N_LAGS      -- the previous weeks' demand
      * roll_mean_w (ROLL_WINDOWS) -- trailing means over the last w weeks
      * t_index                  -- weeks since the SKU's first observation
      * t_frac                   -- position within the SKU's history (0..1)
      * woy_sin / woy_cos        -- week-of-year seasonality encoding
    """
    feats = []
    for lag in range(1, N_LAGS + 1):
        feats.append(z_hist[t - lag] if t - lag >= 0 else np.nan)
    for w in ROLL_WINDOWS:
        lo = max(0, t - w)
        feats.append(float(np.mean(z_hist[lo:t])) if t > lo else np.nan)
    feats.append(float(t))
    feats.append(float(t) / max(n_total - 1, 1))
    woy = pd.Timestamp(week_date).isocalendar().week
    feats.append(np.sin(2 * np.pi * woy / 52.0))
    feats.append(np.cos(2 * np.pi * woy / 52.0))
    return feats


def build_training_set(series_by_sku, min_weeks_for_trend=MIN_WEEKS_FOR_TREND):
    """Pool supervised (features, target) rows across every SKU in the group.

    ``series_by_sku`` maps sku -> (week_dates, cleaned_y). Each SKU's demand is
    divided by its own mean (its "scale") so big and small SKUs train one model
    together without the big ones dominating the squared-error loss; forecasts
    are multiplied back by the scale afterwards. SKUs shorter than
    ``min_weeks_for_trend`` contribute no training rows (they are forecast flat
    at their mean instead -- see ``xgboost_forecast``).

    Returns (X, y, scales) where ``scales`` maps sku -> its mean level.
    """
    X, y = [], []
    scales = {}
    for sku, (week_dates, cy) in series_by_sku.items():
        cy = np.asarray(cy, dtype="float64")
        n = len(cy)
        scale = float(np.mean(cy))
        scales[sku] = scale
        if n < min_weeks_for_trend or scale <= 0:
            continue
        z = cy / scale
        wd = pd.to_datetime(pd.Series(week_dates)).to_numpy()
        for t in range(1, n):          # first week has no lag -> start at t=1
            X.append(_feature_row(z, t, n, wd[t]))
            y.append(z[t])
    if not X:
        return np.empty((0, 0)), np.empty(0), scales
    return np.asarray(X, dtype="float64"), np.asarray(y, dtype="float64"), scales


def xgboost_forecast(model, z_hist, n_total, forecast_weeks, scale,
                     mean_val, min_weeks_for_trend=MIN_WEEKS_FOR_TREND):
    """Recursive 15-step forecast for one SKU using the pooled group model.

    Each step builds a feature row from the trailing (normalised) history,
    predicts the next week, clamps it at zero, and appends the prediction to
    the history so the following step can use it as a lag -- standard recursive
    multi-step forecasting.

    Short series are handled exactly like the smoothing pipeline:
      * 0 weeks  -> flat at zero.
      * 1 week   -> flat at that value.
      * < ``min_weeks_for_trend`` weeks -> flat at the mean (no model), which
        stops tiny-sample groups (e.g. "Others - AU") from producing runaway
        forecasts.
    A ``model`` of None (group too small to train) also falls back to the flat
    mean. Returns a list of ``len(forecast_weeks)`` values (not rounded; the
    caller floors at zero and rounds for consistency with the other pipelines).
    """
    horizon = len(forecast_weeks)
    n = len(z_hist)
    if n == 0:
        return [0.0] * horizon
    if n == 1:
        return [float(z_hist[0]) * scale] * horizon
    if model is None or n < min_weeks_for_trend or scale <= 0:
        return [float(mean_val)] * horizon

    z = list(np.asarray(z_hist, dtype="float64"))
    out = []
    for h in range(horizon):
        t = len(z)
        row = np.asarray(
            [_feature_row(np.asarray(z), t, n_total + h + 1, forecast_weeks[h])],
            dtype="float64",
        )
        pred = float(model.predict(row)[0])
        pred = max(pred, 0.0)          # demand can't be negative
        z.append(pred)
        out.append(pred * scale)
    return out


def fit_xgboost(df, today, grouping_label, breakdown_df=None,
                list_prices=None, cleansing_log=None, uplift_log=None,
                min_weeks_for_trend=MIN_WEEKS_FOR_TREND):
    """Build a 15-week forecast from the historical demand window with XGBoost.

    The fitting window is all completed weeks by default (LOOKBACK_WEEKS=None),
    or the most recent N completed weeks if LOOKBACK_WEEKS is set; the in-progress
    week is always excluded (see ``week_anchors``).

    Before fitting, each SKU's series is run through ``cleanse_series`` so promo
    spikes (e.g. Prime Day) and stockout dips are replaced by a local baseline and
    don't distort the fit. The count per SKU is reported in "Outlier Weeks Cleaned".

    ONE boosted-tree model is trained per call, pooled across every SKU in
    ``df`` (see ``build_training_set``), then each SKU is forecast recursively
    15 weeks ahead (see ``xgboost_forecast``). Hyperparameters come from the
    module-level XGB_PARAMS -- deliberately NOT function arguments, so the
    dashboard's smoothing sliders hide themselves (it inspects this signature
    for alpha/beta/phi and finds none).

    ``df`` must be at SKU-week granularity (see ``aggregate_to_sku_week``).
    For each SKU the forecast is built from POS where available; if a SKU has no
    POS in the window, it falls back to the Orders signal. The "Data Source"
    column records which one was used. SKUs with neither are skipped.
    ``grouping_label`` is written into the "Customer Grouping" column.
    If ``breakdown_df`` (rows that still carry "Customer Grouping") is provided,
    a "Top Volume Customer Groups" column is appended.
    If ``list_prices`` (a SKU -> List Price USD Series) is provided, "List Price
    (USD)" and "Revenue Risk (USD)" = projection_difference * list price are
    added. SKUs without a known price are left blank.
    If ``cleansing_log`` / ``uplift_log`` (lists) are supplied, audit rows are
    appended to them; the dashboard omits them, so the return signature is
    unchanged.
    ``min_weeks_for_trend`` overrides MIN_WEEKS_FOR_TREND for this call (the
    dashboard passes its live slider value here): SKUs with fewer completed
    weeks than this are forecast flat at their mean instead of using the model.
    Returns (summary_df, weekly_df), or (None, None) if no SKU has POS or Orders
    in the historical window (nothing to forecast from).
    """
    lookback_start, last_complete_week, first_forecast_week = week_anchors(today)

    # All completed weeks in the window (default: all history; the in-progress
    # week, whose data is only partial, is always excluded -- see week_anchors).
    # Discontinued items (SKU ends in '*') are dropped entirely. Rows are kept if
    # they carry POS OR Orders, so an orders-only SKU survives.
    window = df[
        (df["WeekDate"] >= lookback_start)
        & (df["WeekDate"] <= last_complete_week)
        & (df["POS"].notna() | df["Orders"].notna())
        & ~df["SKU"].astype(str).str.endswith("*")
    ].sort_values(["SKU", "WeekDate"])

    if window.empty:
        return None, None

    # Project 15 weeks forward starting from the current in-progress week
    forecast_weeks = pd.date_range(start=first_forecast_week, periods=15, freq="W-SUN")

    # Promo week-starts (for re-adding expected lift onto future promo weeks).
    promo_set = {_week_start(d) for d in PROMO_WEEKS} if PROMO_WEEKS else set()

    # ------------------------------------------------------------------ #
    # Pass 1: per SKU -- pick the signal, densify, cleanse. Collect the  #
    # cleaned series so ONE pooled model can be trained across all SKUs. #
    # ------------------------------------------------------------------ #
    prepared = {}          # sku -> dict of everything pass 2 needs
    series_by_sku = {}     # sku -> (week_dates, cleaned_y) for the training pool

    for (sku, desc), grp in window.groupby(["SKU", "Description"]):
        # Prefer POS; fall back to Orders only when the SKU has no POS at all.
        pos_grp = grp[grp["POS"].notna()]
        if not pos_grp.empty:
            source, src_grp = "POS", pos_grp
        else:
            orders_grp = grp[grp["Orders"].notna()]
            if orders_grp.empty:
                continue  # no POS and no Orders -> nothing to forecast
            source, src_grp = "Orders", orders_grp

        src_grp = src_grp.sort_values("WeekDate").reset_index(drop=True)

        # Densify: fill weeks with no data inside the SKU's active span with 0,
        # so a sparse/one-off series isn't fit as if every observed week were
        # consecutive. Span = first observation .. last completed week (no
        # leading zeros before the SKU first appears). See FILL_GAPS_WITH_ZERO.
        if FILL_GAPS_WITH_ZERO and not src_grp.empty:
            full_weeks = pd.date_range(
                start=src_grp["WeekDate"].min(),
                end=last_complete_week,
                freq="W-SUN",
            )
            src_grp = (
                src_grp.set_index("WeekDate")
                .reindex(full_weeks)
                .rename_axis("WeekDate")
                .reset_index()
            )
            src_grp[source] = src_grp[source].fillna(0.0)

        y_raw = src_grp[source].to_numpy(dtype="float64")
        week_dates = src_grp["WeekDate"]
        n = len(src_grp)

        # Replace promo spikes / stockout dips with a local baseline before
        # fitting, so the model learns underlying demand, not the event.
        y, flags, method = cleanse_series(week_dates.to_numpy(), y_raw)

        prepared[sku] = dict(
            desc=desc, source=source, week_dates=week_dates,
            y_raw=y_raw, y=y, flags=flags, method=method, n=n,
        )
        series_by_sku[sku] = (week_dates.to_numpy(), y)

    if not prepared:
        return None, None

    # ------------------------------------------------------------------ #
    # Train ONE pooled model for the whole group. If there's too little  #
    # to train on (or the fit fails), every SKU falls back to flat-mean. #
    # ------------------------------------------------------------------ #
    X, y_train, scales = build_training_set(
        series_by_sku, min_weeks_for_trend=min_weeks_for_trend
    )
    model = None
    if len(y_train) >= MIN_TRAIN_ROWS:
        try:
            model = _make_regressor()
            model.fit(X, y_train)
        except Exception:
            print(traceback.format_exc())
            print(
                f"[{grouping_label}] model fit failed on {len(y_train)} samples; "
                "falling back to flat-mean forecasts."
            )
            model = None
    else:
        print(
            f"[{grouping_label}] only {len(y_train)} pooled training rows "
            f"(< {MIN_TRAIN_ROWS}); forecasting every SKU flat at its mean."
        )

    # ------------------------------------------------------------------ #
    # Pass 2: forecast each SKU with the pooled model (or its fallback), #
    # apply promo uplift, and assemble the summary / weekly rows.        #
    # ------------------------------------------------------------------ #
    summary_rows = []
    weekly_rows = []

    for sku, p in prepared.items():
        y, y_raw, flags, method = p["y"], p["y_raw"], p["flags"], p["method"]
        week_dates, n = p["week_dates"], p["n"]
        desc, source = p["desc"], p["source"]

        # Descriptive average over the (cleaned) fitting window. The column's
        # display label (AVG_COL_LABEL) already reflects LOOKBACK_WEEKS, so it
        # reads "All-History..." rather than a hardcoded "8 Week..." when the
        # window isn't actually 8 weeks.
        mean_val = y.mean()

        scale = scales.get(sku, float(mean_val))
        z_hist = (y / scale) if scale > 0 else y
        raw_forecast = xgboost_forecast(
            model, z_hist, n, forecast_weeks, scale, mean_val,
            min_weeks_for_trend=min_weeks_for_trend,
        )

        # Flat forecast: hold the first week's prediction across all 15 weeks.
        # The app re-runs weekly and only the first projection is ever used, so
        # every week repeats it. Promo uplifts are intentionally dropped
        # (multiplier held at 1.0) so the line is truly flat.
        mult = np.ones(len(forecast_weeks), dtype="float64")
        base = max(round(float(raw_forecast[0]), 1), 0)
        projected_15 = [base] * 15
        n_uplifted = 0

        # Audit record: one row per future week that received a promo uplift.
        if uplift_log is not None and n_uplifted:
            for h in np.where(mult > 1.0)[0]:
                uplift_log.append(
                    {
                        "Customer Grouping": grouping_label,
                        "SKU": sku,
                        "Description": desc,
                        "Data Source": source,
                        "WeekDate": forecast_weeks[h].date(),
                        "Uplift": round(float(mult[h]), 3),
                        "Baseline": max(round(float(raw_forecast[h]), 1), 0),
                        "Projected": projected_15[h],
                    }
                )

        # Audit record: one row per cleaned week (raw value -> baseline used).
        if cleansing_log is not None and flags.any():
            for j in np.where(flags)[0]:
                cleansing_log.append(
                    {
                        "Customer Grouping": grouping_label,
                        "SKU": sku,
                        "Description": desc,
                        "Data Source": source,
                        "WeekDate": pd.Timestamp(week_dates.iloc[j]).date(),
                        "Method": method[j],
                        "Raw": round(float(y_raw[j]), 1),
                        "Baseline": round(float(y[j]), 1),
                    }
                )

        summary_rows.append(
            {
                "SKU": sku,
                "Description": desc,
                "Data Source": source,
                "weeks_with_data": n,
                "outlier_weeks_cleaned": int(flags.sum()),
                "promo_weeks_uplifted": n_uplifted,
                "8_week_pos_avg": round(mean_val, 1),
                "updated_projection_avg": int(round(np.mean(projected_15))),
            }
        )

        for week, projected, m in zip(forecast_weeks, projected_15, mult):
            weekly_rows.append(
                {
                    "SKU": sku,
                    "Description": desc,
                    "WeekDate": week.date(),
                    "projected_pos": projected,
                    "promo_uplift": round(float(m), 3),
                }
            )

    if not summary_rows:
        return None, None

    # Report the actual span of data used (the lower bound is a far-past floor
    # when LOOKBACK_WEEKS is None, so show the earliest week present).
    actual_start = window["WeekDate"].min()
    span_label = (
        "all completed weeks"
        if LOOKBACK_WEEKS is None
        else f"{LOOKBACK_WEEKS} completed weeks"
    )
    print(f"  Historical window: {actual_start.date()} -> {last_complete_week.date()} ({span_label})")
    print(f"  Forecast window:   {forecast_weeks[0].date()} -> {forecast_weeks[-1].date()}")
    print(f"  SKUs projected:    {len(summary_rows)} "
          f"({'pooled XGBoost model' if model is not None else 'flat-mean fallback'}, "
          f"{len(y_train)} training rows)")

    # initial_projection_avg: average of the existing system Projection over the
    # SAME 15 forecast weeks the updated average uses -- from the current
    # in-progress week (first_forecast_week) through the 15th forecast week
    # (forecast_weeks[-1]). Capping the window here (rather than averaging the
    # original projection over its entire future span) makes the two averages
    # apples-to-apples, so "Projection Difference" / "Revenue Risk" agree in sign
    # with the forecast-vs-original comparison shown on the chart. Weeks with a
    # missing projection are still excluded (mean() skips NaN), so a SKU whose
    # projection runs out mid-horizon is not penalised for the blank weeks.
    avg_initial = (
        df[
            (df["WeekDate"] >= first_forecast_week)
            & (df["WeekDate"] <= forecast_weeks[-1])
        ]
        .dropna(subset=["Projection"])
        .groupby("SKU")["Projection"]
        .mean()
        .reset_index()
        .rename(columns={"Projection": "initial_projection_avg"})
    )

    summary_df = (
        pd.DataFrame(summary_rows)
        .merge(avg_initial, on="SKU", how="left")
        .assign(
            initial_projection_avg=lambda d: d["initial_projection_avg"]
            .round()
            .astype("Int64")
        )
        .sort_values("SKU")
        .reset_index(drop=True)
    )

    summary_df['projection_difference'] = summary_df['updated_projection_avg'] - summary_df['initial_projection_avg']

    # Revenue risk: value the change in forecast at each SKU's list price.
    #   revenue_risk = projection_difference * list_price
    # SKUs absent from the price list (or with a blank price) map to NaN, so
    # their revenue risk is left blank rather than treated as $0.
    if list_prices is not None:
        summary_df["list_price_usd"] = summary_df["SKU"].astype(str).map(list_prices)
    else:
        summary_df["list_price_usd"] = np.nan
    summary_df["revenue_risk_usd"] = (
        summary_df["projection_difference"].astype("float64")
        * summary_df["list_price_usd"]
    ).round(2)

    # Stamp the group this file represents, then switch to display column names.
    summary_df["Customer Grouping"] = grouping_label
    summary_df = summary_df.rename(columns=DISPLAY_NAMES)
    summary_df = summary_df[SUMMARY_COLUMNS]

    # All-customers view only: identify which groups drive each SKU's volume.
    if breakdown_df is not None:
        contributors = top_volume_customers(breakdown_df, today)
        summary_df = summary_df.merge(contributors, on="SKU", how="left")
        # SKUs whose POS is present but all zero have no volume to attribute.
        summary_df["Top Volume Customer Groups"] = summary_df[
            "Top Volume Customer Groups"
        ].fillna("(no volume)")

    weekly_df = (
        pd.DataFrame(weekly_rows).sort_values(["SKU", "WeekDate"]).reset_index(drop=True)
    )

    return summary_df, weekly_df


# Backwards-compatibility alias. dashboard.py loads this module by path and calls
# ``P.fit_regression`` (and inspects its signature). Because fit_xgboost accepts
# ``min_weeks_for_trend`` but NOT alpha/beta/phi, the dashboard automatically
# keeps the min-weeks slider and hides the smoothing sliders -- only
# DEMAND_PIPELINE / PIPELINE_PATH needs to change.
fit_regression = fit_xgboost


def _autosize_to_headers(worksheet, df, padding=2):
    """Widen each column so its title is fully visible (title length + padding)."""
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = 'A2'

    for i, col in enumerate(df.columns, start=1):
        if i == 1:
            worksheet.column_dimensions[get_column_letter(i)].width = len(str(col)) + 10
        else:
            worksheet.column_dimensions[get_column_letter(i)].width = len(str(col)) + padding


def write_forecast(summary_df, weekly_df, out_path):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        weekly_df.to_excel(writer, sheet_name="weekly_forecast", index=False)
        _autosize_to_headers(writer.sheets["summary"], summary_df)
        _autosize_to_headers(writer.sheets["weekly_forecast"], weekly_df)


if __name__ == "__main__":
    INPUT_FILE, today_str, TODAY = resolve_input_file()
    OUTPUT_FOLDER = f"outputs/demand_projections/xgboost/{today_str}"
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    print(f"Snapshot date (anchor): {today_str}\n")

    LIST_PRICES = load_list_prices()

    df = pd.read_excel(INPUT_FILE, header=2)
    df = df.rename(
        columns={
            "'Demand'[DisplaySKU]": "SKU",
            "Custnmbr": "CUSTNMBR",
            "Sum of Quantity": "Orders",
        }
    )
    df = df[["SKU", "Description", "CUSTNMBR", "WeekDate", "POS", "Orders", "Projection"]]
    # The fixed-width export space-pads SKU/CUSTNMBR; strip before any key-based
    # lookup so SKUs match the list-price index and CUSTNMBRs fold via
    # COMBINED_GROUPING. Kept in sync with agent/data_io._clean (shared by the
    # dashboard + agent), which this __main__ block mirrors.
    df["SKU"] = df["SKU"].astype(str).str.strip()
    df["CUSTNMBR"] = df["CUSTNMBR"].astype(str).str.strip()
    df = df[~df['CUSTNMBR'].isin(CUSTOMERS_TO_IGNORE)]
    df["WeekDate"] = pd.to_datetime(df["WeekDate"])

    # Consolidated customer group. Customers absent from COMBINED_GROUPING fall
    # back to their own name (single-member group), so nothing is dropped.
    df["Customer Grouping"] = df["CUSTNMBR"].map(COMBINED_GROUPING).fillna(df["CUSTNMBR"])
    ungrouped = sorted(
        df.loc[~df["CUSTNMBR"].isin(COMBINED_GROUPING), "CUSTNMBR"].dropna().unique()
    )
    if ungrouped:
        print(
            f"{len(ungrouped)} customers not in COMBINED_GROUPING "
            f"(kept as their own group): {ungrouped}\n"
        )

    # ------------------------------------------------------------------ #
    # 1) Per-group forecasts: one file per Customer Grouping             #
    #    (member customers' POS are summed together)                     #
    # ------------------------------------------------------------------ #
    print("=== Per-group forecasts ===")
    groups = df["Customer Grouping"].dropna().unique().tolist()
    lb, lc, _ = week_anchors(TODAY)
    succeeded, no_pos, no_data, errors = [], [], [], []
    cleansing_log = []   # one row per cleaned promo/outlier week (written below)
    uplift_log = []      # one row per future promo week given an uplift (below)

    for group in groups:
        try:
            group_df = aggregate_to_sku_week(df[df["Customer Grouping"] == group])

            # Group-level data presence in the window (discontinued excluded).
            gw = group_df[
                (group_df["WeekDate"] >= lb)
                & (group_df["WeekDate"] <= lc)
                & ~group_df["SKU"].astype(str).str.endswith("*")
            ]
            has_pos = gw["POS"].notna().any()
            has_orders = gw["Orders"].notna().any()
            if not has_pos:
                no_pos.append(group)            # forecast (if any) came from Orders
            if not has_pos and not has_orders:
                no_data.append(group)           # nothing to forecast at all

            summary_df, weekly_df = fit_xgboost(
                group_df, TODAY, grouping_label=group, list_prices=LIST_PRICES,
                cleansing_log=cleansing_log, uplift_log=uplift_log,
            )
            if summary_df is None:
                continue

            print(f"[{group}] ok")
            region = region_for_group(group)
            safe_group = group.replace("/", "-")
            out_path = (
                f"{OUTPUT_FOLDER}/{region}/{safe_group}/"
                f"{safe_group}_demand_projections_{today_str}.xlsx"
            )
            write_forecast(summary_df, weekly_df, out_path)
            succeeded.append(group)
        except Exception:
            print(traceback.format_exc())
            errors.append(group)

    # ------------------------------------------------------------------ #
    # 2) All-customers combined: one SKU-level file (POS summed across    #
    #    every customer group)                                            #
    # ------------------------------------------------------------------ #
    print("\n=== All-customers (combined SKU) forecast ===")
    combined_path = f"{OUTPUT_FOLDER}/ALL_SKUS_demand_projections_{today_str}.xlsx"
    try:
        combined_df = aggregate_to_sku_week(df)
        combined_summary, combined_weekly = fit_xgboost(
            combined_df, TODAY, grouping_label=ALL_SKUS_LABEL,
            breakdown_df=df, list_prices=LIST_PRICES,
            cleansing_log=cleansing_log, uplift_log=uplift_log,
        )
        if combined_summary is None:
            print("No POS data in the historical window for any SKU; combined file skipped.")
        else:
            write_forecast(combined_summary, combined_weekly, combined_path)
            print(f"[ALL_SKUS] ok -> {combined_path}")
    except Exception:
        print(traceback.format_exc())

    # Concatenate every per-group summary sheet into one ALL_CUSTOMERS workbook.
    # Region folders are derived from region_for_group (plus "Other") rather
    # than hardcoded, so no region's output can be silently skipped.
    region_folders = sorted({region_for_group(g) for g in groups} | {"Other"})
    customer_dfs = []
    for folder in region_folders:
        for xlsx_file in (Path(OUTPUT_FOLDER) / folder).rglob("*.xlsx"):
            try:
                customer_dfs.append(pd.read_excel(xlsx_file))
            except Exception as e:
                print(f"Failed: {xlsx_file.name} — {e}")
    if customer_dfs:
        combined = pd.concat(customer_dfs, ignore_index=True)
        combined.to_excel(
            f"{OUTPUT_FOLDER}/ALL_CUSTOMERS_demand_projections_{today_str}.xlsx",
            index=False,
        )
    else:
        print("No per-group workbooks found; ALL_CUSTOMERS file skipped.")

    # ------------------------------------------------------------------ #
    print("\n=== Summary ===")
    print(f"Per-group forecasts written: {len(succeeded)}/{len(groups)}")
    print(f"No POS, forecast from Orders (or skipped): {len(no_pos)} -> {no_pos}")
    print(f"Skipped, no POS and no Orders: {len(no_data)} -> {no_data}")
    print(f"Errors: {len(errors)}/{len(groups)} -> {errors}")
    print(f"Input file: {INPUT_FILE}")
    print(f"Output folder: {OUTPUT_FOLDER}")

    # Groups with no POS in the window (their forecast, if any, used Orders).
    if no_pos:
        no_pos_path = os.path.join(OUTPUT_FOLDER, f"no_pos_{today_str}.txt")
        with open(no_pos_path, "w") as f:
            f.write("\n".join(no_pos))
        print(f"No-POS customers saved to: {no_pos_path}")

    # Groups with neither POS nor Orders (no forecast produced).
    if no_data:
        no_data_path = os.path.join(OUTPUT_FOLDER, f"no_pos_or_orders_{today_str}.txt")
        with open(no_data_path, "w") as f:
            f.write("\n".join(no_data))
        print(f"No-POS-or-Orders customers saved to: {no_data_path}")

    # Promo / outlier weeks that were cleansed before fitting (both the manual
    # PROMO_WEEKS and the automatically MAD-detected ones), for the record.
    cleaned_path = os.path.join(OUTPUT_FOLDER, f"cleaned_outliers_{today_str}.txt")
    promo_weeks_resolved = sorted({str(_week_start(d).date()) for d in PROMO_WEEKS}) or ["(none)"]
    with open(cleaned_path, "w") as f:
        f.write(f"Outlier / promo weeks cleaned before fitting (anchor {today_str})\n")
        f.write(
            f"Auto-detect (MAD): {'on' if CLEANSE_OUTLIERS else 'off'} "
            f"(K={OUTLIER_K}, window={OUTLIER_WINDOW}, min_weeks={OUTLIER_MIN_WEEKS})\n"
        )
        f.write(f"Manual promo weeks: {', '.join(promo_weeks_resolved)}\n")
        f.write(f"Total weeks cleaned: {len(cleansing_log)}\n\n")
        if cleansing_log:
            log_df = (
                pd.DataFrame(cleansing_log)[
                    ["Customer Grouping", "SKU", "Description",
                     "Data Source", "WeekDate", "Method", "Raw", "Baseline"]
                ]
                .sort_values(["Customer Grouping", "SKU", "WeekDate"])
                .reset_index(drop=True)
            )
            f.write(log_df.to_string(index=False))
            f.write("\n")
        else:
            f.write("No weeks were flagged.\n")
    print(f"Cleaned-outlier record saved to: {cleaned_path} ({len(cleansing_log)} weeks)")

    # Future promo weeks that received an uplift (baseline -> projected), for the
    # record. Mirrors the cleansing log: which upcoming weeks were bumped and by
    # how much, so the promo lift in the forecast is auditable.
    uplift_path = os.path.join(OUTPUT_FOLDER, f"promo_uplift_{today_str}.txt")
    uplift_mode = (
        f'auto (per-SKU, default {PROMO_UPLIFT_DEFAULT}x, max {PROMO_UPLIFT_MAX}x)'
        if isinstance(PROMO_UPLIFT, str) and PROMO_UPLIFT.lower() == "auto"
        else f"fixed {PROMO_UPLIFT}x"
    )
    with open(uplift_path, "w") as f:
        f.write(f"Future promo weeks uplifted in the forecast (anchor {today_str})\n")
        f.write(f"Uplift mode: {uplift_mode}; halo: {PROMO_HALO_WEEKS} week(s) each side\n")
        f.write(f"Manual promo weeks: {', '.join(promo_weeks_resolved)}\n")
        f.write(f"Total week-rows uplifted: {len(uplift_log)}\n\n")
        if uplift_log:
            up_df = (
                pd.DataFrame(uplift_log)[
                    ["Customer Grouping", "SKU", "Description",
                     "Data Source", "WeekDate", "Uplift", "Baseline", "Projected"]
                ]
                .sort_values(["Customer Grouping", "SKU", "WeekDate"])
                .reset_index(drop=True)
            )
            f.write(up_df.to_string(index=False))
            f.write("\n")
        else:
            f.write(
                "No promo weeks fell inside the 15-week horizon for this snapshot "
                "(or PROMO_UPLIFT <= 1.0).\n"
            )
    print(f"Promo-uplift record saved to: {uplift_path} ({len(uplift_log)} week-rows)")
