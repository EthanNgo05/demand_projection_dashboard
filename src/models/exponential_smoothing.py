"""
Updates projections for all SKUs based on combined demand projections for all companies.

Uses *all* completed weeks of historical POS data to calculate a 15-week
projection. The in-progress week is excluded so a partially-elapsed week's POS
never drags the smoothed level (see week_anchors). The amount of history is
controlled by LOOKBACK_WEEKS (None = all history; set an int to cap it).

This is the exponential-smoothing counterpart of the linear-regression pipeline.
Instead of fitting a straight line to the history and projecting the slope, it
runs Holt's linear method with a damped trend (double exponential smoothing):

    level_t    = ALPHA * y_t + (1 - ALPHA) * (level_{t-1} + PHI * trend_{t-1})
    trend_t    = BETA  * (level_t - level_{t-1}) + (1 - BETA) * PHI * trend_{t-1}
    projected_pos(week_h) = level_T + (PHI + PHI^2 + ... + PHI^h) * trend_T   (h = 1 .. 15)

    - ALPHA : level smoothing  (0..1) -- how fast the level tracks recent demand
    - BETA  : trend smoothing  (0..1) -- how fast the slope adapts (the ES analogue
                                         of the old TREND_WEIGHT; exposed as
                                         TREND_WEIGHT for dashboard compatibility)
    - PHI   : trend damping    (0..1) -- < 1 flattens the trend the further out we
                                         forecast, so a short-run slope is not
                                         extrapolated indefinitely (PHI = 1 -> plain
                                         Holt; PHI = 0 -> flat at the level)
    - updated_projection_avg (per SKU) = mean of the 15 weekly projected_pos values

Compared with the regression version this reacts more smoothly to noise, weights
recent weeks more heavily than older ones, and -- thanks to the damping -- tapers
an unsustainable trend rather than letting it run away over the 15-week horizon.

Before fitting, each SKU's series is cleansed of one-off promo spikes (e.g. Amazon
Prime Day) and stockout dips: abnormal weeks are detected automatically (rolling
median +/- OUTLIER_K MADs) and/or named explicitly (PROMO_WEEKS), then replaced
with a local baseline so they don't inflate the smoothed level. Every cleaned week
is written to a ``cleaned_outliers_<date>.txt`` record (see CLEANSE_OUTLIERS and
``cleanse_series``).

Promos also lift future demand, so any PROMO_WEEKS date that lands inside the
15-week horizon has its projection scaled up by a promo uplift factor (fixed, or
estimated per SKU from its own past promos). Cleansing removes the historical
spike to get an honest baseline; the uplift re-adds the expected lift onto the
upcoming promo weeks (baseline + uplift). Uplifted weeks are recorded to a
``promo_uplift_<date>.txt`` file (see PROMO_UPLIFT).

Two kinds of output are produced:
    1) Per-group     : one file per Customer Grouping (see COMBINED_GROUPING),
                       summing POS across the customers that make up the group
                       (e.g. the three Amazon-DC channels become one AMAZON-DC
                       forecast). Customers not in the mapping become their own
                       single-member group.
    2) All-customers : a single file forecasting each SKU "as a whole", built by
                       summing POS (and the existing Projection) across every
                       customer for each SKU-week before fitting.
"""

import os
import re
import glob
import traceback
import numpy as np
import pandas as pd
from pathlib import Path

# --- Exponential smoothing (Holt's damped-trend) parameters ---------------- #
# All in [0, 1]. Defaults are deliberately conservative: a moderately responsive
# level, a gently adapting trend, and damping so the trend tapers rather than
# runs away over the 15-week horizon.
ALPHA = 0.15   # level smoothing  -- higher = track recent demand faster
BETA = 0.05    # trend smoothing  -- higher = let the slope change faster
PHI = 0.6    # trend damping    -- < 1 flattens the trend further out (1 = plain Holt)

# Backwards-compatibility alias. The dashboard reads ``TREND_WEIGHT`` for a
# caption; in Holt's method BETA *is* the weight given to the trend, so it is the
# natural analogue. Nothing in this file's math reads TREND_WEIGHT directly.
TREND_WEIGHT = BETA

# Minimum completed weeks of history required before a TREND is fitted. A slope
# estimated from only a week or two is essentially noise, and Holt's method then
# extrapolates it across the whole 15-week horizon -- so a single spike in a
# 2-week series becomes a runaway ramp (e.g. "Others - AU"). Series shorter than
# this are instead forecast FLAT at their mean (level only, no extrapolation),
# which is far more sensible when there isn't enough history to see a trend.
# Raise it to be more conservative (more series held flat); 2 disables the guard.
MIN_WEEKS_FOR_TREND = 4

# --- History window -------------------------------------------------------- #
# How many of the most-recent *completed* weeks to fit the model on.
#   None -> use ALL available history (every completed week up to last week)
#   int  -> use only the most recent N completed weeks (e.g. 8 for the old window)
# Exponential smoothing weights recent weeks more heavily anyway, so feeding it
# all history lets long-run level/trend inform the forecast without old weeks
# dominating. The in-progress week is always excluded (see week_anchors).
LOOKBACK_WEEKS = None

# When LOOKBACK_WEEKS is None the lower bound is set to this many years before
# the run date -- effectively "all history" for our data, while still guarding
# against a stray ancient row dragging the fit. See week_anchors.
HISTORY_YEARS = 3

# Display label for the descriptive-average column. Reflects LOOKBACK_WEEKS so
# the header never claims "8 Week" when it's actually averaging all history
# (or some other window) -- see DISPLAY_NAMES / SUMMARY_COLUMNS below.
AVG_COL_LABEL = (
    "All-History POS/Orders Average"
    if LOOKBACK_WEEKS is None
    else f"{LOOKBACK_WEEKS} Week POS/Orders Average"
)

# --- Intermittent / lumpy demand ------------------------------------------- #
# The historical series for a SKU only carries rows for weeks that actually had
# POS (or Orders). Weeks with none are ABSENT, not zero. For sell-through / order
# demand a missing completed week almost always means "nothing sold/ordered that
# week" = 0, not "unknown", so dropping those weeks massively inflates a sparse
# SKU: a customer that ordered 3160 once and nothing for the next two months is
# otherwise fit as a one-point series and forecast at 3160 EVERY week (this is
# what blows up "Others - AU", which is entirely one-off orders).
#
# With this on, each SKU's series is reindexed to every completed week from its
# FIRST observation through the last completed week and the gaps filled with 0
# before cleansing/fitting. Two effects:
#   * the flat-mean fallback becomes a true weekly average (total / weeks-in-span)
#     instead of the mean of only the non-zero weeks, and
#   * "Weeks with data" / MIN_WEEKS_FOR_TREND now count the real elapsed span, so
#     a single spike is correctly seen as 1 order across many weeks (held flat),
#     not as a full multi-week history.
# Leading weeks BEFORE a SKU's first-ever order are NOT zero-filled, so a newly
# introduced SKU isn't penalised for weeks it didn't yet exist. Set False to
# restore the old observed-weeks-only behaviour.
FILL_GAPS_WITH_ZERO = True

# --- Outlier / promo cleansing --------------------------------------------- #
# One-off spikes (e.g. Amazon Prime Day) and dips (e.g. stockouts) distort an
# exponential-smoothing fit: with ALPHA=0.5 the level chases the most recent
# weeks, so a single promotional week can inflate the whole 15-week forecast.
# Before fitting, abnormal weeks are replaced with a local baseline (the median
# of nearby normal weeks) so the model learns underlying demand, not the event.
# Each cleaned week is recorded to a .txt audit file (see __main__).
#
# Two sources of flags are unioned:
#   1. Manual  -- PROMO_WEEKS below (always applied; you know these are promos).
#   2. Auto    -- a week more than OUTLIER_K MADs from a centred rolling median
#                 (symmetric, so it catches both spikes and stockout dips).
CLEANSE_OUTLIERS = True       # master switch for the automatic MAD detector
OUTLIER_K = 3.0               # MADs from the local median to count as an outlier
OUTLIER_WINDOW = 7            # weeks in the rolling median/MAD (odd -> centred)
OUTLIER_MIN_WEEKS = 5         # don't auto-detect on series shorter than this

# Manual promo calendar: weeks you KNOW are abnormal. Any date inside the week
# works -- it is snapped to that week's start (the Sunday WeekDate) before
# matching, so you don't have to look up the exact Sunday. Always cleansed,
# regardless of CLEANSE_OUTLIERS, and recorded with method "manual".
PROMO_WEEKS = [
    # "2026-06-26",   # Amazon Prime Day week
    # # "2026-07-?"   # Friends and Family
    # "2026-08-15",   # Back to school shopping
    # "2026-11-27",   # Black Friday week
    # "2026-11-30",   # Cyber Monday
    # "2026-12-1",    # Christmas shopping
    # "2026-12-8",    # Christmas shopping
    # "2026-12-15"     # Christmas shopping
]

# --- Future promo uplift --------------------------------------------------- #
# Promos lift sales, so any PROMO_WEEKS date that lands inside the 15-week
# forecast horizon has its projection scaled UP. (Historical copies of those
# weeks are still cleansed for the baseline fit above -- this only re-adds the
# expected lift onto FUTURE promo weeks: baseline + uplift.)
#
# PROMO_UPLIFT controls the peak multiplier on a promo week:
#   a number (e.g. 1.25) -> fixed: multiply promo-week projections by this factor
#   "auto"               -> estimate the factor per SKU from its OWN historical
#                           promo weeks (raw / cleansed baseline), falling back to
#                           PROMO_UPLIFT_DEFAULT when that SKU has no promo history
PROMO_UPLIFT = 1.25            # "a bit higher"; set "auto" for a data-driven lift
PROMO_UPLIFT_DEFAULT = 1.25    # fallback factor for "auto" when no SKU history
PROMO_UPLIFT_MAX = 4.0         # clamp so a noisy estimate can't explode a week
PROMO_HALO_WEEKS = 0           # also lift this many weeks either side (0 = off),
                               # tapering linearly to the edge of the halo

# --- Autofit (grid-search backtest of ALPHA / BETA / PHI) ------------------- #
# ``autofit_smoothing`` searches a grid of (alpha, beta, phi) combinations and
# scores each one by BACKTESTING: the last AUTOFIT_HOLDOUT_WEEKS completed weeks
# of every SKU's (cleansed) history are hidden from the model, forecast with the
# candidate parameters, and compared against what actually happened. This is
# repeated from AUTOFIT_FOLDS rolling origins (each fold slides the cut-off back
# a few more weeks) so the winner isn't tuned to a single lucky window. The
# combination with the lowest total absolute error across every SKU, fold and
# holdout week wins.
#
# Notes on the scoring choices:
#   * Errors are summed in UNITS (not percentages), so high-volume SKUs weigh
#     more -- the fit minimises total units of forecast error, which is what
#     drives inventory dollars.
#   * Both training and holdout weeks use the CLEANSED series (promo spikes /
#     stockout dips replaced by baseline), so the search optimises for
#     underlying demand rather than chasing one-off events.
#   * Forecasts are floored at zero before scoring, matching the pipeline's
#     output convention.
#   * Ties break toward the earliest grid entry, i.e. the LOWEST alpha/beta/phi
#     (grids below are ascending), so "equally good" defaults stay conservative.
AUTOFIT_HOLDOUT_WEEKS = 6      # weeks hidden at the end of history for scoring
AUTOFIT_FOLDS = 3              # rolling origins (1 = single holdout)
AUTOFIT_FOLD_STEP = 3          # weeks the cut-off slides back per extra fold
AUTOFIT_MIN_TRAIN_WEEKS = 8    # min completed weeks a fold must train on to count
AUTOFIT_ALPHA_GRID = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50, 0.70, 0.90]
AUTOFIT_BETA_GRID = [0.00, 0.02, 0.05, 0.10, 0.15, 0.20, 0.30, 0.50]
AUTOFIT_PHI_GRID = [0.20, 0.40, 0.60, 0.80, 0.90, 0.95, 1.00]

RAW_INPUTS_FOLDER = "raw_inputs/demand_projections"
INPUT_GLOB = os.path.join(RAW_INPUTS_FOLDER, "all_demand_projections_*.xlsx")

# List-price workbook (SKU -> List Price USD), used to value the change in
# forecast as a revenue risk. See load_list_prices().
LIST_PRICE_GLOB = os.path.join("raw_inputs/list_prices", "list_prices_*.xlsx")

CUSTOMERS_TO_IGNORE = [
    "Others - HK", "Others - KR", "Others - MALDIV", "Others - MX",
    "Others - UAE", "Others - UK", "Others - ZZ",
]

# Maps each raw CUSTNMBR onto a consolidated customer group. Several customers
# fold into one group (e.g. the three Amazon-DC channels), so forecasts are built
# per group rather than per raw customer. Customers not listed here fall back to
# their own name as a single-member group.
US_GROUPING = {
    'AMAZON-DC': 'AMAZON-DC',
    'AMAZON-DS': 'AMAZON-DC',
    'MARVAL-FBM': 'AMAZON-DC',
    'TARGET-HQ': 'TARGET-HQ',
    'TARGET-DS': 'TARGET-DS',
    'TARGET-PLUS': 'TARGET-DS',
    'HOMDEP': 'HOMDEP',
    'LOWES-HQ': 'LOWES-HQ',
    'LOWES-SOS': 'LOWES-SOS',
    'Web Sales - US': 'Web Sales + Warranty US',
    'Warranty - US': 'Web Sales + Warranty US',
    'CONSTO': 'CONSTO',
    'COSTCO.COM': 'COSTCO.COM',
    'ULINE': 'ULINE',
    'AAFES': 'AAFES',
    'STAPADV.COM': 'STAPADV.COM',
    'STAPLES.COM': 'STAPADV.COM',
    'DILLARDS': 'DILLARDS',
    'NORSTR': 'NORSTR',
    'NORCOM-DS': 'NORSTR',
    'NORCOM': 'NORSTR',
    'Others - US': 'Others - US'
}

EU_GROUPING = {
    'AMAZON-EU': 'AMAZON-EU',
    'Web Sales - EU': 'Web Sales + Warranty EU',
    'CASTOR': 'CASTOR',
    'EBREUNINGER': 'EBREUNINGER',
    'FSKUSTERMANN': 'FSKUSTERMANN',
    'GALLAF': 'GALLAF',
    'LERMER-FR': 'LERMER-FR',
    'SANIKAL-KG': 'SANIKAL-KG',
    'Others - EU': 'Others - EU',
    'Others - IT': 'Others - EU'
}

AU_GROUPING = {
    'Web Sales - AU': 'Web Sales - AU',
    'Others - AU': 'Others - AU'
}

CA_GROUPING = {
    'AMAZON-DCCA': 'AMAZON-DCCA',
    'CANTIR': 'CANTIR',
    'COSTCO-CAN': 'COSTCO-CAN',
    'HOMEDEP-CA.COM': 'HOMEDEP-CA.COM',
    'HOMEDEPOT-CA': 'HOMEDEPOT-CA',
    'Others - CA': 'Others - CA',
    'RONA-HQ': 'RONA-HQ',
    'Web Sales - CA': 'Web Sales - CA',
    'WINMER': 'WINMER'
}

JP_GROUPING = {
    'Web Sales - JP': 'Web Sales - JP',
    'Others - JP': 'Others - JP'
}

COMBINED_GROUPING = {**US_GROUPING, **EU_GROUPING, **AU_GROUPING, **CA_GROUPING, **JP_GROUPING}

# Group labels belonging to each region (the *values* of the maps above), used to
# file each group's output under US / EU / NON-US-EU. A group that came from
# neither map (an unmapped customer kept as its own group) lands in NON-US-EU.
US_GROUPS = set(US_GROUPING.values())
EU_GROUPS = set(EU_GROUPING.values())
AU_GROUPS = set(AU_GROUPING.values())
CA_GROUPS = set(CA_GROUPING.values())
JP_GROUPS = set(JP_GROUPING.values())

def region_for_group(group):
    if group in US_GROUPS:
        return "US (LBC+NJ)"
    elif group in EU_GROUPS:
        return "EU (SH-CTS)"
    elif group in AU_GROUPS:
        return "AU (ACR)"
    elif group in CA_GROUPS:
        return "CA (YYZ5)"
    elif group in JP_GROUPS:
        return "JP (NETDEPOT)"
    return "Other"

# Label used for the all-customers (combined) output.
ALL_SKUS_LABEL = "ALL SKUS"

# Internal (snake_case) column names -> display names used in the output sheets.
DISPLAY_NAMES = {
    "weeks_with_data": "Weeks with data",
    "outlier_weeks_cleaned": "Outlier Weeks Cleaned",
    "promo_weeks_uplifted": "Promo Weeks Uplifted",
    "8_week_pos_avg": AVG_COL_LABEL,
    "initial_projection_avg": "Current Projection Average",
    "updated_projection_avg": "Updated Projection Average",
    "projection_difference": "Projection Difference",
    "list_price_usd": "List Price (USD)",
    "revenue_risk_usd": "Revenue Risk (avg/wk)",
}

# Final column order for every summary sheet
SUMMARY_COLUMNS = [
    "SKU",
    "Description",
    "Customer Grouping",
    "Data Source",
    "Weeks with data",
    AVG_COL_LABEL,
    "Updated Projection Average",
    "Current Projection Average",
    "Projection Difference",
    "List Price (USD)",
    "Revenue Risk (avg/wk)",
]


def resolve_input_file():
    """Pick the most recent raw data file and anchor the forecast to its snapshot date.

    Anchoring TODAY to the date in the filename (rather than the wall clock) keeps
    the historical lookback and the 15-week forecast aligned with the data
    snapshot, even if the script is run a day or two after the file was generated.
    """
    candidates = []
    for path in glob.glob(INPUT_GLOB):
        m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
        if m:
            candidates.append((m.group(1), path))

    if not candidates:
        raise FileNotFoundError(
            f"No input files matching {INPUT_GLOB}. "
            f"Expected e.g. {RAW_INPUTS_FOLDER}/all_demand_projections_YYYY-MM-DD.xlsx"
        )

    today_str, input_file = max(candidates)  # latest date wins
    return input_file, today_str, pd.Timestamp(today_str)


def load_list_prices(path=None):
    """Load a SKU -> List Price (USD) lookup from a price workbook.

    If ``path`` is given, that workbook is read directly. Otherwise the most
    recently modified file matching LIST_PRICE_GLOB is used. The sheet carries
    two columns, "SKU" and "List Price USD". SKUs with a blank price are dropped
    from the lookup so they map to NaN downstream -- their revenue risk is left
    blank rather than treated as $0 (an unknown price is not zero risk).

    Returns a pandas Series of price indexed by SKU (str), or None if no price
    file is found (in which case revenue risk is left blank for every SKU).
    """
    if path is None:
        candidates = glob.glob(LIST_PRICE_GLOB)
        if not candidates:
            print(
                f"No list-price file matching {LIST_PRICE_GLOB}; "
                f"revenue risk will be left blank.\n"
            )
            return None
        path = max(candidates, key=os.path.getmtime)  # most recently modified

    prices = pd.read_excel(path)
    prices.columns = [str(c).strip() for c in prices.columns]
    prices = prices[["SKU", "List Price USD"]].dropna(subset=["SKU"])
    prices["SKU"] = prices["SKU"].astype(str).str.strip()
    prices["List Price USD"] = pd.to_numeric(prices["List Price USD"], errors="coerce")
    prices = prices.dropna(subset=["List Price USD"]).drop_duplicates("SKU", keep="last")
    lookup = prices.set_index("SKU")["List Price USD"]

    print(f"Loaded {len(lookup)} list prices from {os.path.basename(path)}\n")
    return lookup


def week_anchors(today):
    """Resolve the week boundaries for the forecast.

    WeekDate is the Sunday that *starts* each 7-day week, so the week labelled
    W covers [W, W+6]. A week is only "completely over" once today is past its
    Saturday. To avoid feeding a partially-elapsed week's POS into the model,
    the historical (training) window ends at the last fully-completed week.

    The forecast, however, STARTS at the current in-progress week rather than
    skipping to the next full week. That week is only partly elapsed, so it is
    deliberately kept OUT of the training history (its partial POS would drag
    the smoothed level), but it still needs a projection -- otherwise it would
    fall into a gap, plotted as neither an actual nor a forecast. Because Holt's
    method projects h = 1..15 steps ahead of the last trained week, step h = 1 is
    exactly this in-progress week, so first_forecast_week = current_week_start
    also keeps the horizon date labels aligned with the projection steps.

    The window *start* depends on LOOKBACK_WEEKS: with the default ``None`` the
    lower bound is HISTORY_YEARS years before the run date (i.e. all available
    history in practice); set LOOKBACK_WEEKS to an int to use only that many
    most-recent completed weeks instead.

    Example (run on Thu 2026-06-25): the week of 2026-06-21 is still in progress
    (runs through Sat 2026-06-27), so the training window ends at 2026-06-14.
    With LOOKBACK_WEEKS=None it starts HISTORY_YEARS years back (all history);
    with LOOKBACK_WEEKS=8 it would start at 2026-04-26. The forecast starts at the
    in-progress week (2026-06-21): that week is projected but is NOT used to fit
    the model, so its partial POS never distorts the level/trend.

    Returns (lookback_start, last_complete_week, first_forecast_week).
    """
    days_since_sunday = (today.weekday() + 1) % 7          # Sun=0 ... Sat=6
    current_week_start = today - pd.Timedelta(days=days_since_sunday)
    last_complete_week = current_week_start - pd.Timedelta(weeks=1)
    if LOOKBACK_WEEKS is None:
        lookback_start = today - pd.DateOffset(years=HISTORY_YEARS)  # all history
    else:
        # N weeks inclusive -> step back N-1 from the last completed week.
        lookback_start = last_complete_week - pd.Timedelta(weeks=LOOKBACK_WEEKS - 1)
    # Forecast starts at the current in-progress week (not the next full week),
    # so that partly-elapsed week gets a projection instead of falling into a gap.
    # It is still excluded from the training window above (which ends at
    # last_complete_week), so its partial POS never drags the fit.
    first_forecast_week = current_week_start
    return lookback_start, last_complete_week, first_forecast_week


def aggregate_to_sku_week(df):
    """Collapse rows to one row per SKU-week, summing POS and Projection.

    For a single customer this is an identity (the data already has one row per
    SKU-week), so the per-customer logic is unchanged. For the all-customers view
    it sums each SKU's POS / Projection across every CUSTNMBR, giving the SKU's
    total demand. ``min_count=1`` keeps an all-NaN group as NaN (no data) rather
    than collapsing it to a real 0.
    """
    grp = df.groupby(["SKU", "WeekDate"])
    pos = grp["POS"].sum(min_count=1)
    orders = grp["Orders"].sum(min_count=1)
    proj = grp["Projection"].sum(min_count=1)
    agg = pd.concat([pos, orders, proj], axis=1).reset_index()

    # Description is consistent per SKU; attach the first non-null one.
    desc = df.dropna(subset=["Description"]).groupby("SKU")["Description"].first()
    agg["Description"] = agg["SKU"].map(desc)
    return agg


def top_volume_customers(df, today, top_n=3):
    """For each SKU, rank the customer groups driving the most volume.

    Uses the same completed-week window the forecast is built from (all history
    by default; see LOOKBACK_WEEKS). Volume is measured with POS where the SKU
    has any, otherwise with Orders (mirroring the forecast's POS-then-Orders
    fallback), summed per Customer Grouping and turned into a share of the SKU
    total. The top ``top_n`` are returned as one string, e.g.
    "AMAZON-DC (61%); Web Sales + Warranty US (31%); AMAZON-EU (3%)".

    Returns a DataFrame [SKU, Top Volume Customer Groups].
    """
    lookback_start, last_complete_week, _ = week_anchors(today)
    win = df[
        (df["WeekDate"] >= lookback_start)
        & (df["WeekDate"] <= last_complete_week)
        & (df["POS"].notna() | df["Orders"].notna())
        & ~df["SKU"].astype(str).str.endswith("*")
    ].copy()

    # Per SKU: rank by POS volume if the SKU has any POS, else by Orders volume.
    has_pos = win.groupby("SKU")["POS"].transform(lambda s: s.notna().any())
    win["vol"] = np.where(has_pos, win["POS"], win["Orders"])

    by_grp = win.groupby(["SKU", "Customer Grouping"], as_index=False)["vol"].sum()
    by_grp = by_grp[by_grp["vol"] > 0]
    by_grp["sku_total"] = by_grp.groupby("SKU")["vol"].transform("sum")
    by_grp = by_grp[by_grp["sku_total"] > 0]
    by_grp["share"] = by_grp["vol"] / by_grp["sku_total"]
    by_grp = by_grp.sort_values(["SKU", "vol"], ascending=[True, False])

    rows = []
    for sku, grp in by_grp.groupby("SKU"):
        head = grp.head(top_n)
        label = "; ".join(
            f"{name} ({share * 100:.0f}%)"
            for name, share in zip(head["Customer Grouping"], head["share"])
        )
        rows.append({"SKU": sku, "Top Volume Customer Groups": label})

    return pd.DataFrame(rows, columns=["SKU", "Top Volume Customer Groups"])


def _week_start(date):
    """Snap any date to the Sunday that starts its week (matching WeekDate)."""
    ts = pd.Timestamp(date).normalize()
    return ts - pd.Timedelta(days=(ts.weekday() + 1) % 7)


def _mad(a):
    """Median absolute deviation of a 1-D array (ignoring NaNs)."""
    a = np.asarray(a, dtype="float64")
    a = a[~np.isnan(a)]
    if a.size == 0:
        return np.nan
    return np.median(np.abs(a - np.median(a)))


def cleanse_series(week_dates, y, promo_week_starts=None, detect=None,
                   k=OUTLIER_K, window=OUTLIER_WINDOW, min_weeks=OUTLIER_MIN_WEEKS):
    """Replace promo / outlier weeks with a local baseline before smoothing.

    Spikes (Prime Day) and dips (stockouts) corrupt an exponential-smoothing fit
    because the level chases recent weeks. This returns a cleaned copy of ``y``
    in which abnormal weeks are swapped for the median of nearby *normal* weeks,
    so the model sees underlying demand rather than the event.

    Weeks are flagged from two sources, unioned:
      * Manual -- WeekDate falls in ``promo_week_starts`` (defaults to the module
        PROMO_WEEKS); always applied, recorded as method "manual".
      * Auto   -- |y - rolling_median| > k * 1.4826 * rolling_MAD over a centred
        ``window``-week window (1.4826 scales MAD to a normal std, so k is ~k-sigma).
        Skipped when ``detect`` is False or the series is shorter than ``min_weeks``.
        Recorded as method "auto".

    Detection uses the raw series (the median/MAD are robust to a minority of
    outliers); replacement uses only non-flagged neighbours so a multi-week promo
    can't baseline itself. Returns (cleaned_y, flags, method) as numpy arrays the
    same length as ``y`` (``method`` holds "manual"/"auto"/"").
    """
    y = np.asarray(y, dtype="float64")
    n = len(y)
    flags = np.zeros(n, dtype=bool)
    method = np.array([""] * n, dtype=object)
    if n == 0:
        return y.copy(), flags, method

    if detect is None:
        detect = CLEANSE_OUTLIERS
    if promo_week_starts is None:
        promo_week_starts = PROMO_WEEKS

    # 1) Manual promo weeks (always applied).
    promo_set = {_week_start(d) for d in promo_week_starts} if promo_week_starts else set()
    if promo_set:
        wk = pd.DatetimeIndex(pd.to_datetime(week_dates)).normalize()
        wk_start = wk - pd.to_timedelta((wk.weekday + 1) % 7, unit="D")
        manual = np.array([ws in promo_set for ws in wk_start])
        flags |= manual
        method[manual] = "manual"

    # 2) Automatic MAD detection on the raw series.
    if detect and n >= min_weeks:
        s = pd.Series(y)
        med = s.rolling(window, center=True, min_periods=3).median().to_numpy()
        mad = s.rolling(window, center=True, min_periods=3).apply(_mad, raw=True).to_numpy()
        med = np.where(np.isnan(med), np.nanmedian(y), med)
        global_mad = _mad(y)
        scale = np.where((np.isnan(mad)) | (mad == 0), global_mad, mad)
        with np.errstate(invalid="ignore"):
            auto = (scale > 0) & (np.abs(y - med) > k * 1.4826 * scale)
        newly = auto & ~flags
        flags |= auto
        method[newly] = "auto"

    if not flags.any():
        return y.copy(), flags, method

    # Replace each flagged week with the median of nearby NON-flagged weeks.
    half = max(window // 2, 1)
    cleaned = y.copy()
    global_baseline = np.median(y[~flags]) if (~flags).any() else np.nan
    for i in np.where(flags)[0]:
        lo, hi = max(0, i - half), min(n, i + half + 1)
        local = [y[j] for j in range(lo, hi) if not flags[j]]
        if local:
            cleaned[i] = float(np.median(local))
        elif not np.isnan(global_baseline):
            cleaned[i] = float(global_baseline)
        # else: every week flagged -> leave the original value untouched
    return cleaned, flags, method


def estimate_promo_uplift(y_raw, baseline, method):
    """Per-SKU promo uplift factor = mean(raw / baseline) over its promo weeks.

    Uses only manually-flagged promo weeks (``method == "manual"``) with a
    positive baseline -- i.e. how much the SKU's real sales exceeded the cleansed
    baseline on past promos. Returns None when there is nothing to estimate from
    (no historical promo weeks for this SKU in the fitting window).
    """
    ratios = [
        y_raw[j] / baseline[j]
        for j in range(len(method))
        if method[j] == "manual" and baseline[j] > 0
    ]
    if not ratios:
        return None
    return float(np.mean(ratios))


def promo_week_multipliers(forecast_weeks, factor, promo_set, halo=PROMO_HALO_WEEKS):
    """Per-forecast-week uplift multipliers (1.0 = no change).

    ``forecast_weeks`` is the sequence of 15 projected week-start dates,
    ``promo_set`` a set of promo week-start Timestamps (see ``_week_start``), and
    ``factor`` the peak multiplier applied on a promo week. With ``halo`` > 0 the
    ``halo`` weeks on each side also lift, tapering linearly to the halo edge;
    overlapping halos keep the strongest lift. Returns a numpy array.
    """
    n = len(forecast_weeks)
    mult = np.ones(n, dtype="float64")
    if factor <= 1.0 or not promo_set:
        return mult

    fw = pd.DatetimeIndex(pd.to_datetime(forecast_weeks)).normalize()
    is_promo = np.array([ts in promo_set for ts in fw])
    for i in np.where(is_promo)[0]:
        for d in range(-halo, halo + 1):
            j = i + d
            if 0 <= j < n:
                taper = 1.0 if d == 0 else (halo + 1 - abs(d)) / (halo + 1)
                mult[j] = max(mult[j], 1.0 + (factor - 1.0) * taper)
    return mult


def holt_damped_forecast(y, horizon, alpha=ALPHA, beta=BETA, phi=PHI,
                         min_weeks_for_trend=MIN_WEEKS_FOR_TREND):
    """Holt's linear method with a damped trend (double exponential smoothing).

    Smooths the ordered demand sequence ``y`` into a level and a trend, then
    projects ``horizon`` steps ahead, damping the trend by ``phi`` each step so a
    short-run slope is not extrapolated indefinitely:

        level_t = alpha * y_t + (1 - alpha) * (level_{t-1} + phi * trend_{t-1})
        trend_t = beta  * (level_t - level_{t-1}) + (1 - beta) * phi * trend_{t-1}
        forecast_h = level_T + (phi + phi**2 + ... + phi**h) * trend_T

    ``y`` is the chronological sequence of observed demand. Gaps in the weekly
    history are treated as consecutive observations (the original regression
    pipeline only used real week offsets for its slope; here the smoothing runs
    over the points it has).

    Short series are handled specially, because a trend estimated from only a
    couple of weeks is noise that would be extrapolated over the whole horizon:
      * 0 weeks  -> flat at zero.
      * 1 week   -> flat at that value.
      * < ``min_weeks_for_trend`` weeks -> flat at the mean of ``y`` (level only,
        no trend). This is what stops tiny-sample groups (e.g. "Others - AU")
        from producing runaway ramps.
    With at least ``min_weeks_for_trend`` weeks the full damped-trend model runs.

    Returns a list of ``horizon`` forecast values. Values are NOT clamped or
    rounded here -- the caller does that (so it can floor at zero consistently).
    """
    y = np.asarray(y, dtype="float64")
    n = len(y)
    if n == 0:
        return [0.0] * horizon
    if n == 1:
        return [float(y[0])] * horizon

    # Too few weeks to trust a trend: forecast FLAT at the mean rather than
    # extrapolating a slope fit to a handful of noisy points.
    if n < min_weeks_for_trend:
        return [float(np.mean(y))] * horizon

    # Initialise the level at the first observation and the trend at the first
    # observed step. With only ~8 points this is more stable than estimating a
    # separate warm-up regression.
    level = y[0]
    trend = y[1] - y[0]
    for t in range(1, n):
        prev_level = level
        level = alpha * y[t] + (1.0 - alpha) * (prev_level + phi * trend)
        trend = beta * (level - prev_level) + (1.0 - beta) * phi * trend

    # Damped-trend projection: each step adds one more phi power to the running
    # sum (phi + phi^2 + ... + phi^h), so the trend contribution saturates.
    forecasts = []
    phi_pow = 1.0
    damp_sum = 0.0
    for _ in range(horizon):
        phi_pow *= phi
        damp_sum += phi_pow
        forecasts.append(level + damp_sum * trend)
    return forecasts


def _series_for_fit(grp, last_complete_week):
    """Per-SKU demand series exactly as ``fit_exponential_smoothing`` builds it.

    Mirrors the fit's preprocessing so the autofit backtest scores the SAME
    series the real forecast is fit on: POS-then-Orders source selection,
    optional zero-densification of the active span (FILL_GAPS_WITH_ZERO), and
    promo/outlier cleansing. Returns (source, week_dates, y_cleaned) or
    (None, None, None) when the SKU has neither POS nor Orders.
    """
    pos_grp = grp[grp["POS"].notna()]
    if not pos_grp.empty:
        source, src_grp = "POS", pos_grp
    else:
        orders_grp = grp[grp["Orders"].notna()]
        if orders_grp.empty:
            return None, None, None
        source, src_grp = "Orders", orders_grp

    src_grp = src_grp.sort_values("WeekDate").reset_index(drop=True)

    if FILL_GAPS_WITH_ZERO and not src_grp.empty:
        full_weeks = pd.date_range(
            start=src_grp["WeekDate"].min(), end=last_complete_week, freq="W-SUN"
        )
        src_grp = (
            src_grp.set_index("WeekDate")
            .reindex(full_weeks)
            .rename_axis("WeekDate")
            .reset_index()
        )
        src_grp[source] = src_grp[source].fillna(0.0)

    y_raw = src_grp[source].to_numpy(dtype="float64")
    y, _, _ = cleanse_series(src_grp["WeekDate"].to_numpy(), y_raw)
    return source, src_grp["WeekDate"], y


def _holt_grid_forecast(y, horizon, alphas, betas, phis):
    """Run Holt's damped-trend recursion for MANY (alpha, beta, phi) combos at once.

    ``alphas`` / ``betas`` / ``phis`` are parallel 1-D arrays (one entry per
    combo). The smoothing recursion is identical to ``holt_damped_forecast``
    but the level/trend state is a vector over combos, so a whole parameter
    grid is evaluated in one pass over the series instead of one pass per
    combo. Requires len(y) >= 2. Returns an array of shape (horizon, n_combos).
    """
    m = alphas.shape[0]
    level = np.full(m, y[0], dtype="float64")
    trend = np.full(m, y[1] - y[0], dtype="float64")
    for t in range(1, len(y)):
        prev_level = level
        level = alphas * y[t] + (1.0 - alphas) * (prev_level + phis * trend)
        trend = betas * (level - prev_level) + (1.0 - betas) * phis * trend

    fc = np.empty((horizon, m), dtype="float64")
    phi_pow = np.ones(m, dtype="float64")
    damp_sum = np.zeros(m, dtype="float64")
    for h in range(horizon):
        phi_pow = phi_pow * phis
        damp_sum = damp_sum + phi_pow
        fc[h] = level + damp_sum * trend
    return fc


def autofit_smoothing(df, today, alpha_grid=None, beta_grid=None, phi_grid=None,
                      holdout_weeks=AUTOFIT_HOLDOUT_WEEKS, folds=AUTOFIT_FOLDS,
                      fold_step=AUTOFIT_FOLD_STEP,
                      min_train_weeks=AUTOFIT_MIN_TRAIN_WEEKS,
                      min_weeks_for_trend=MIN_WEEKS_FOR_TREND):
    """Grid-search the (alpha, beta, phi) that best predicts held-out history.

    ``df`` must be at SKU-week granularity (see ``aggregate_to_sku_week``) --
    the same frame ``fit_exponential_smoothing`` receives. For every SKU with
    enough history the model is fit on all weeks up to a rolling cut-off with
    each parameter combination, and the single next week is forecast and scored
    one-step-ahead (absolute error, in units, on the cleansed series, forecast
    floored at zero). One-step-ahead matches the published forecast, which is
    flat at the first week's value, so only next-week accuracy is used. ``folds``
    rolling origins slide the cut-off back ``fold_step`` weeks at a time so the
    winner generalises across recent windows rather than fitting one lucky one.

    A fold only counts when its training slice still has at least
    ``max(min_train_weeks, min_weeks_for_trend)`` weeks -- below
    ``min_weeks_for_trend`` the model forecasts flat at the mean regardless of
    parameters, so such folds carry no signal and are skipped for speed.

    The module defaults (ALPHA, BETA, PHI) are always scored too, so the
    result reports how much the winner improves on the file defaults.

    Returns a dict::

        {"alpha", "beta", "phi",          # best combination found
         "mae",                            # its backtest MAE (units/week)
         "baseline_mae",                   # MAE of the file defaults
         "baseline_params": (a, b, p),     # the file defaults scored
         "n_series", "n_points",           # how much data backed the score
         "holdout_weeks", "folds"}

    or None when no SKU has enough completed history to backtest.
    """
    alpha_grid = AUTOFIT_ALPHA_GRID if alpha_grid is None else list(alpha_grid)
    beta_grid = AUTOFIT_BETA_GRID if beta_grid is None else list(beta_grid)
    phi_grid = AUTOFIT_PHI_GRID if phi_grid is None else list(phi_grid)

    # Flatten the grid into parallel combo arrays; ascending order means a tie
    # resolves to the most conservative (lowest) parameters via argmin.
    combos = [
        (a, b, p)
        for a in sorted(alpha_grid)
        for b in sorted(beta_grid)
        for p in sorted(phi_grid)
    ]
    baseline = (float(ALPHA), float(BETA), float(PHI))
    if baseline not in combos:
        combos.append(baseline)          # always score the file defaults too
    baseline_idx = combos.index(baseline)

    A = np.array([c[0] for c in combos], dtype="float64")
    B = np.array([c[1] for c in combos], dtype="float64")
    PH = np.array([c[2] for c in combos], dtype="float64")

    lookback_start, last_complete_week, _ = week_anchors(today)
    window = df[
        (df["WeekDate"] >= lookback_start)
        & (df["WeekDate"] <= last_complete_week)
        & (df["POS"].notna() | df["Orders"].notna())
        & ~df["SKU"].astype(str).str.endswith("*")
    ].sort_values(["SKU", "WeekDate"])
    if window.empty:
        return None

    min_train = max(int(min_train_weeks), int(min_weeks_for_trend), 2)
    total_err = np.zeros(A.shape[0], dtype="float64")
    n_points = 0
    n_series = 0

    for (_sku, _desc), grp in window.groupby(["SKU", "Description"]):
        source, _weeks, y = _series_for_fit(grp, last_complete_week)
        if source is None:
            continue
        n = len(y)
        if n < min_train + 1:
            continue                     # nothing to hold out after training

        scored = False
        for k in range(max(int(folds), 1)):
            cut = n - int(holdout_weeks) - k * int(fold_step)
            if cut < min_train:
                # Not enough training weeks left; with < holdout remaining
                # weeks a shorter final fold is still allowed below.
                cut = min_train if (k == 0 and n - min_train >= 1) else None
                if cut is None:
                    break
            # One-step-ahead only: the published forecast is flat at the first
            # week's value, so tune params on how well they predict the NEXT
            # week, not a multi-week horizon we no longer use. Each rolling
            # fold contributes one one-step-ahead sample at a distinct origin.
            if n - cut < 1:
                break
            fc = np.maximum(_holt_grid_forecast(y[:cut], 1, A, B, PH), 0.0)  # (1, n_combos)
            total_err += np.abs(fc[0] - y[cut])
            n_points += 1
            scored = True
            if cut == min_train:
                break                    # can't slide the origin back further
        if scored:
            n_series += 1

    if n_points == 0:
        return None

    best = int(np.argmin(total_err))
    return {
        "alpha": float(A[best]),
        "beta": float(B[best]),
        "phi": float(PH[best]),
        "mae": float(total_err[best] / n_points),
        "baseline_mae": float(total_err[baseline_idx] / n_points),
        "baseline_params": baseline,
        "n_series": int(n_series),
        "n_points": int(n_points),
        "holdout_weeks": int(holdout_weeks),
        "folds": int(folds),
    }


def fit_exponential_smoothing(df, today, grouping_label, breakdown_df=None,
                              list_prices=None, cleansing_log=None, uplift_log=None,
                              alpha=ALPHA, beta=BETA, phi=PHI,
                              min_weeks_for_trend=MIN_WEEKS_FOR_TREND):
    """Build a 15-week forecast from the historical demand window.

    The fitting window is all completed weeks by default (LOOKBACK_WEEKS=None),
    or the most recent N completed weeks if LOOKBACK_WEEKS is set; the in-progress
    week is always excluded (see ``week_anchors``).

    Before smoothing, each SKU's series is run through ``cleanse_series`` so promo
    spikes (e.g. Prime Day) and stockout dips are replaced by a local baseline and
    don't distort the fit. The count per SKU is reported in "Outlier Weeks Cleaned".

    ``df`` must be at SKU-week granularity (see ``aggregate_to_sku_week``).
    For each SKU the forecast is built from POS where available; if a SKU has no
    POS in the window, it falls back to the Orders signal using the identical
    Holt damped-trend exponential smoothing (see ``holt_damped_forecast``). The
    "Data Source" column records which one was used. SKUs with neither POS nor
    Orders are skipped.
    ``grouping_label`` is written into the "Customer Grouping" column (the group
    name for a per-group file, or ALL_SKUS_LABEL for the combined file).
    If ``breakdown_df`` (rows that still carry "Customer Grouping") is provided, a
    "Top Volume Customer Groups" column is appended.
    If ``list_prices`` (a SKU -> List Price USD Series, see ``load_list_prices``)
    is provided, two columns are added: "List Price (USD)" and "Revenue Risk
    (USD)" = projection_difference * list price. A negative value means the
    updated forecast fell below the original (revenue at risk on the downside);
    a positive value is upside. SKUs without a known price are left blank.
    If ``cleansing_log`` (a list) is supplied, one dict per cleaned week is
    appended to it for the audit record; the dashboard omits it (no file is
    written during interactive use), so the return signature is unchanged.
    Future PROMO_WEEKS falling inside the 15-week horizon have their projection
    scaled up by the promo uplift factor (see PROMO_UPLIFT); if ``uplift_log`` (a
    list) is supplied, one dict per uplifted week is appended to it.
    ``alpha`` / ``beta`` / ``phi`` override the module-level smoothing constants
    for this call (the dashboard passes its live slider values here); when
    omitted they default to ALPHA / BETA / PHI, so the batch __main__ run is
    unchanged. ``min_weeks_for_trend`` likewise overrides MIN_WEEKS_FOR_TREND:
    SKUs with fewer completed weeks than this are forecast flat at their mean
    instead of extrapolating a trend from too little history.
    Returns (summary_df, weekly_df), or (None, None) if no SKU has POS or Orders
    in the historical window (nothing to forecast from).
    """
    lookback_start, last_complete_week, first_forecast_week = week_anchors(today)

    # All completed weeks in the window (default: all history; the in-progress
    # week, whose data is only partial, is always excluded -- see week_anchors).
    # Discontinued items (SKU ends in '*') are dropped entirely. Rows are kept if
    # they carry POS OR Orders, so an orders-only SKU survives.
    window = df[
        (df["WeekDate"] >= lookback_start)
        & (df["WeekDate"] <= last_complete_week)
        & (df["POS"].notna() | df["Orders"].notna())
        & ~df["SKU"].astype(str).str.endswith("*")
    ].sort_values(["SKU", "WeekDate"])

    if window.empty:
        return None, None

    # Project 15 weeks forward starting from the current in-progress week
    forecast_weeks = pd.date_range(start=first_forecast_week, periods=15, freq="W-SUN")

    # Promo week-starts (for re-adding expected lift onto future promo weeks).
    promo_set = {_week_start(d) for d in PROMO_WEEKS} if PROMO_WEEKS else set()

    summary_rows = []
    weekly_rows = []

    for (sku, desc), grp in window.groupby(["SKU", "Description"]):
        # Prefer POS; fall back to Orders only when the SKU has no POS at all.
        pos_grp = grp[grp["POS"].notna()]
        if not pos_grp.empty:
            source, src_grp = "POS", pos_grp
        else:
            orders_grp = grp[grp["Orders"].notna()]
            if orders_grp.empty:
                continue  # no POS and no Orders -> nothing to forecast
            source, src_grp = "Orders", orders_grp

        src_grp = src_grp.sort_values("WeekDate").reset_index(drop=True)

        # Densify: fill weeks with no data inside the SKU's active span with 0,
        # so a sparse/one-off series isn't fit as if every observed week were
        # consecutive. Span = first observation .. last completed week (no
        # leading zeros before the SKU first appears). See FILL_GAPS_WITH_ZERO.
        if FILL_GAPS_WITH_ZERO and not src_grp.empty:
            full_weeks = pd.date_range(
                start=src_grp["WeekDate"].min(),
                end=last_complete_week,
                freq="W-SUN",
            )
            src_grp = (
                src_grp.set_index("WeekDate")
                .reindex(full_weeks)
                .rename_axis("WeekDate")
                .reset_index()
            )
            src_grp[source] = src_grp[source].fillna(0.0)

        y_raw = src_grp[source].to_numpy(dtype="float64")
        week_dates = src_grp["WeekDate"]
        n = len(src_grp)

        # Replace promo spikes / stockout dips with a local baseline before
        # fitting, so the model learns underlying demand, not the event.
        y, flags, method = cleanse_series(week_dates.to_numpy(), y_raw)

        # Descriptive average over the (cleaned) fitting window. The column's
        # display label (AVG_COL_LABEL) already reflects LOOKBACK_WEEKS, so it
        # reads "All-History..." rather than a hardcoded "8 Week..." when the
        # window isn't actually 8 weeks.
        mean_val = y.mean()

        # Holt's damped-trend exponential smoothing over the cleaned weeks.
        raw_forecast = holt_damped_forecast(
            y, 15, alpha=alpha, beta=beta, phi=phi,
            min_weeks_for_trend=min_weeks_for_trend,
        )

        # Flat forecast: hold the first week's cleansed baseline across all 15
        # weeks. The app re-runs weekly and only the first projection is ever
        # used, so every week repeats it. Promo uplifts are intentionally
        # dropped (multiplier held at 1.0) so the line is truly flat.
        # Rounded to a whole number: projections are unit counts, not decimals.
        mult = np.ones(len(forecast_weeks), dtype="float64")
        base = max(int(round(float(raw_forecast[0]))), 0)
        projected_15 = [base] * 15
        n_uplifted = 0

        # Audit record: one row per future week that received a promo uplift.
        if uplift_log is not None and n_uplifted:
            for h in np.where(mult > 1.0)[0]:
                uplift_log.append(
                    {
                        "Customer Grouping": grouping_label,
                        "SKU": sku,
                        "Description": desc,
                        "Data Source": source,
                        "WeekDate": forecast_weeks[h].date(),
                        "Uplift": round(float(mult[h]), 3),
                        "Baseline": max(round(float(raw_forecast[h]), 1), 0),
                        "Projected": projected_15[h],
                    }
                )

        # Audit record: one row per cleaned week (raw value -> baseline used).
        if cleansing_log is not None and flags.any():
            for j in np.where(flags)[0]:
                cleansing_log.append(
                    {
                        "Customer Grouping": grouping_label,
                        "SKU": sku,
                        "Description": desc,
                        "Data Source": source,
                        "WeekDate": pd.Timestamp(week_dates.iloc[j]).date(),
                        "Method": method[j],
                        "Raw": round(float(y_raw[j]), 1),
                        "Baseline": round(float(y[j]), 1),
                    }
                )

        summary_rows.append(
            {
                "SKU": sku,
                "Description": desc,
                "Data Source": source,
                "weeks_with_data": n,
                "outlier_weeks_cleaned": int(flags.sum()),
                "promo_weeks_uplifted": n_uplifted,
                "8_week_pos_avg": round(mean_val, 1),
                "updated_projection_avg": int(round(np.mean(projected_15))),
            }
        )

        for week, projected, m in zip(forecast_weeks, projected_15, mult):
            weekly_rows.append(
                {
                    "SKU": sku,
                    "Description": desc,
                    "WeekDate": week.date(),
                    "projected_pos": projected,
                    "promo_uplift": round(float(m), 3),
                }
            )

    if not summary_rows:
        return None, None

    # Report the actual span of data used (the lower bound is a far-past
    # floor when LOOKBACK_WEEKS is None, so show the earliest week present).
    actual_start = window["WeekDate"].min()
    span_label = (
        "all completed weeks"
        if LOOKBACK_WEEKS is None
        else f"{LOOKBACK_WEEKS} completed weeks"
    )
    print(f"  Historical window: {actual_start.date()} -> {last_complete_week.date()} ({span_label})")
    print(f"  Forecast window:   {forecast_weeks[0].date()} -> {forecast_weeks[-1].date()}")
    print(f"  SKUs projected:    {len(summary_rows)}")

    # initial_projection_avg: average of the existing system Projection over the
    # SAME 15 forecast weeks the updated average uses -- from the current
    # in-progress week (first_forecast_week) through the 15th forecast week
    # (forecast_weeks[-1]). Capping the window here (rather than averaging the
    # original projection over its entire future span) makes the two averages
    # apples-to-apples, so "Projection Difference" / "Revenue Risk" agree in sign
    # with the forecast-vs-original comparison shown on the chart. Weeks with a
    # missing projection are still excluded (mean() skips NaN), so a SKU whose
    # projection runs out mid-horizon is not penalised for the blank weeks.
    avg_initial = (
        df[
            (df["WeekDate"] >= first_forecast_week)
            & (df["WeekDate"] <= forecast_weeks[-1])
        ]
        .dropna(subset=["Projection"])
        .groupby("SKU")["Projection"]
        .mean()
        .reset_index()
        .rename(columns={"Projection": "initial_projection_avg"})
    )

    summary_df = (
        pd.DataFrame(summary_rows)
        .merge(avg_initial, on="SKU", how="left")
        .assign(
            initial_projection_avg=lambda d: d["initial_projection_avg"]
            .round()
            .astype("Int64")
        )
        .sort_values("SKU")
        .reset_index(drop=True)
    )


    summary_df['projection_difference'] = summary_df['updated_projection_avg'] - summary_df['initial_projection_avg']

    # Revenue risk: value the change in forecast at each SKU's list price.
    #   revenue_risk = projection_difference * list_price
    # SKUs absent from the price list (or with a blank price) map to NaN, so
    # their revenue risk is left blank rather than treated as $0.
    if list_prices is not None:
        summary_df["list_price_usd"] = summary_df["SKU"].astype(str).map(list_prices)
    else:
        summary_df["list_price_usd"] = np.nan
    summary_df["revenue_risk_usd"] = (
        summary_df["projection_difference"].astype("float64")
        * summary_df["list_price_usd"]
    ).round(2)

    # Stamp the group this file represents, then switch to display column names.
    summary_df["Customer Grouping"] = grouping_label
    summary_df = summary_df.rename(columns=DISPLAY_NAMES)
    summary_df = summary_df[SUMMARY_COLUMNS]

    # All-customers view only: identify which groups drive each SKU's volume.
    if breakdown_df is not None:
        contributors = top_volume_customers(breakdown_df, today)
        summary_df = summary_df.merge(contributors, on="SKU", how="left")
        # SKUs whose POS is present but all zero have no volume to attribute.
        summary_df["Top Volume Customer Groups"] = summary_df[
            "Top Volume Customer Groups"
        ].fillna("(no volume)")

    weekly_df = (
        pd.DataFrame(weekly_rows).sort_values(["SKU", "WeekDate"]).reset_index(drop=True)
    )

    return summary_df, weekly_df


# Backwards-compatibility alias. dashboard.py loads this module by path and calls
# ``P.fit_regression`` (and inspects its signature for ``list_prices``). Pointing
# the old name at the exponential-smoothing implementation lets the dashboard run
# unchanged against this pipeline -- only DEMAND_PIPELINE / PIPELINE_PATH changes.
fit_regression = fit_exponential_smoothing


def _autosize_to_headers(worksheet, df, padding=2):
    """Widen each column so its title is fully visible (title length + padding)."""
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = 'A2'

    for i, col in enumerate(df.columns, start=1):
        if i == 1:
            worksheet.column_dimensions[get_column_letter(i)].width = len(str(col)) + 10
        else:
            worksheet.column_dimensions[get_column_letter(i)].width = len(str(col)) + padding


def write_forecast(summary_df, weekly_df, out_path):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        weekly_df.to_excel(writer, sheet_name="weekly_forecast", index=False)
        _autosize_to_headers(writer.sheets["summary"], summary_df)
        _autosize_to_headers(writer.sheets["weekly_forecast"], weekly_df)


if __name__ == "__main__":
    INPUT_FILE, today_str, TODAY = resolve_input_file()
    OUTPUT_FOLDER = f"outputs/demand_projections/exponential_smoothing/{today_str}"
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    print(f"Snapshot date (anchor): {today_str}\n")

    LIST_PRICES = load_list_prices()

    df = pd.read_excel(INPUT_FILE, header=2)
    df = df.rename(
        columns={
            "'Demand'[DisplaySKU]": "SKU",
            "Custnmbr": "Customer",
            "Sum of Quantity": "Orders",
        }
    )
    df = df[["SKU", "Description", "Customer", "WeekDate", "POS", "Orders", "Projection"]]
    # The fixed-width export space-pads SKU/Customer; strip before any key-based
    # lookup so SKUs match the list-price index and customers fold via
    # COMBINED_GROUPING. Kept in sync with agent/data_io._clean (shared by the
    # dashboard + agent), which this __main__ block mirrors.
    df["SKU"] = df["SKU"].astype(str).str.strip()
    df["Customer"] = df["Customer"].astype(str).str.strip()
    df = df[~df['Customer'].isin(CUSTOMERS_TO_IGNORE)]
    df["WeekDate"] = pd.to_datetime(df["WeekDate"])

    # Consolidated customer group. Customers absent from COMBINED_GROUPING fall
    # back to their own name (single-member group), so nothing is dropped.
    df["Customer Grouping"] = df["Customer"].map(COMBINED_GROUPING).fillna(df["Customer"])
    ungrouped = sorted(
        df.loc[~df["Customer"].isin(COMBINED_GROUPING), "Customer"].dropna().unique()
    )
    if ungrouped:
        print(
            f"{len(ungrouped)} customers not in COMBINED_GROUPING "
            f"(kept as their own group): {ungrouped}\n"
        )

    # ------------------------------------------------------------------ #
    # 1) Per-group forecasts: one file per Customer Grouping             #
    #    (member customers' POS are summed together)                     #
    # ------------------------------------------------------------------ #
    print("=== Per-group forecasts ===")
    groups = df["Customer Grouping"].dropna().unique().tolist()
    lb, lc, _ = week_anchors(TODAY)
    succeeded, no_pos, no_data, errors = [], [], [], []
    cleansing_log = []   # one row per cleaned promo/outlier week (written below)
    uplift_log = []      # one row per future promo week given an uplift (below)

    for group in groups:
        try:
            group_df = aggregate_to_sku_week(df[df["Customer Grouping"] == group])

            # Group-level data presence in the window (discontinued excluded).
            gw = group_df[
                (group_df["WeekDate"] >= lb)
                & (group_df["WeekDate"] <= lc)
                & ~group_df["SKU"].astype(str).str.endswith("*")
            ]
            has_pos = gw["POS"].notna().any()
            has_orders = gw["Orders"].notna().any()
            if not has_pos:
                no_pos.append(group)            # forecast (if any) came from Orders
            if not has_pos and not has_orders:
                no_data.append(group)           # nothing to forecast at all

            summary_df, weekly_df = fit_exponential_smoothing(
                group_df, TODAY, grouping_label=group, list_prices=LIST_PRICES,
                cleansing_log=cleansing_log, uplift_log=uplift_log,
            )
            if summary_df is None:
                continue

            print(f"[{group}] ok")
            region = region_for_group(group)
            safe_group = group.replace("/", "-")
            out_path = (
                f"{OUTPUT_FOLDER}/{region}/{safe_group}/"
                f"{safe_group}_demand_projections_{today_str}.xlsx"
            )
            write_forecast(summary_df, weekly_df, out_path)
            succeeded.append(group)
        except Exception:
            print(traceback.format_exc())
            errors.append(group)

    # ------------------------------------------------------------------ #
    # 2) All-customers combined: one SKU-level file (POS summed across    #
    #    every customer group)                                            #
    # ------------------------------------------------------------------ #
    print("\n=== All-customers (combined SKU) forecast ===")
    combined_path = f"{OUTPUT_FOLDER}/ALL_SKUS_demand_projections_{today_str}.xlsx"
    try:
        combined_df = aggregate_to_sku_week(df)
        combined_summary, combined_weekly = fit_exponential_smoothing(
            combined_df, TODAY, grouping_label=ALL_SKUS_LABEL,
            breakdown_df=df, list_prices=LIST_PRICES,
            cleansing_log=cleansing_log, uplift_log=uplift_log,
        )
        if combined_summary is None:
            print("No POS data in the historical window for any SKU; combined file skipped.")
        else:
            write_forecast(combined_summary, combined_weekly, combined_path)
            print(f"[ALL_SKUS] ok -> {combined_path}")
    except Exception:
        print(traceback.format_exc())

    # Concatenate every per-group summary sheet into one ALL_CUSTOMERS workbook.
    # Region folders are derived from region_for_group (plus "Other") rather
    # than hardcoded, so no region's output can be silently skipped.
    region_folders = sorted({region_for_group(g) for g in groups} | {"Other"})
    customer_dfs = []
    for folder in region_folders:
        for xlsx_file in (Path(OUTPUT_FOLDER) / folder).rglob("*.xlsx"):
            try:
                customer_dfs.append(pd.read_excel(xlsx_file))
            except Exception as e:
                print(f"Failed: {xlsx_file.name} — {e}")
    if customer_dfs:
        combined = pd.concat(customer_dfs, ignore_index=True)
        combined.to_excel(
            f"{OUTPUT_FOLDER}/ALL_CUSTOMERS_demand_projections_{today_str}.xlsx",
            index=False,
        )
    else:
        print("No per-group workbooks found; ALL_CUSTOMERS file skipped.")

    # ------------------------------------------------------------------ #
    print("\n=== Summary ===")
    print(f"Per-group forecasts written: {len(succeeded)}/{len(groups)}")
    print(f"No POS, forecast from Orders (or skipped): {len(no_pos)} -> {no_pos}")
    print(f"Skipped, no POS and no Orders: {len(no_data)} -> {no_data}")
    print(f"Errors: {len(errors)}/{len(groups)} -> {errors}")
    print(f"Input file: {INPUT_FILE}")
    print(f"Output folder: {OUTPUT_FOLDER}")

    # Groups with no POS in the window (their forecast, if any, used Orders).
    if no_pos:
        no_pos_path = os.path.join(OUTPUT_FOLDER, f"no_pos_{today_str}.txt")
        with open(no_pos_path, "w", encoding="utf-8") as f:
            f.write("\n".join(no_pos))
        print(f"No-POS customers saved to: {no_pos_path}")

    # Groups with neither POS nor Orders (no forecast produced).
    if no_data:
        no_data_path = os.path.join(OUTPUT_FOLDER, f"no_pos_or_orders_{today_str}.txt")
        with open(no_data_path, "w", encoding="utf-8") as f:
            f.write("\n".join(no_data))
        print(f"No-POS-or-Orders customers saved to: {no_data_path}")

    # Promo / outlier weeks that were cleansed before fitting (both the manual
    # PROMO_WEEKS and the automatically MAD-detected ones), for the record.
    cleaned_path = os.path.join(OUTPUT_FOLDER, f"cleaned_outliers_{today_str}.txt")
    promo_weeks_resolved = sorted({str(_week_start(d).date()) for d in PROMO_WEEKS}) or ["(none)"]
    with open(cleaned_path, "w", encoding="utf-8") as f:
        f.write(f"Outlier / promo weeks cleaned before fitting (anchor {today_str})\n")
        f.write(
            f"Auto-detect (MAD): {'on' if CLEANSE_OUTLIERS else 'off'} "
            f"(K={OUTLIER_K}, window={OUTLIER_WINDOW}, min_weeks={OUTLIER_MIN_WEEKS})\n"
        )
        f.write(f"Manual promo weeks: {', '.join(promo_weeks_resolved)}\n")
        f.write(f"Total weeks cleaned: {len(cleansing_log)}\n\n")
        if cleansing_log:
            log_df = (
                pd.DataFrame(cleansing_log)[
                    ["Customer Grouping", "SKU", "Description",
                     "Data Source", "WeekDate", "Method", "Raw", "Baseline"]
                ]
                .sort_values(["Customer Grouping", "SKU", "WeekDate"])
                .reset_index(drop=True)
            )
            f.write(log_df.to_string(index=False))
            f.write("\n")
        else:
            f.write("No weeks were flagged.\n")
    print(f"Cleaned-outlier record saved to: {cleaned_path} ({len(cleansing_log)} weeks)")

    # Future promo weeks that received an uplift (baseline -> projected), for the
    # record. Mirrors the cleansing log: which upcoming weeks were bumped and by
    # how much, so the promo lift in the forecast is auditable.
    uplift_path = os.path.join(OUTPUT_FOLDER, f"promo_uplift_{today_str}.txt")
    uplift_mode = (
        f'auto (per-SKU, default {PROMO_UPLIFT_DEFAULT}x, max {PROMO_UPLIFT_MAX}x)'
        if isinstance(PROMO_UPLIFT, str) and PROMO_UPLIFT.lower() == "auto"
        else f"fixed {PROMO_UPLIFT}x"
    )
    with open(uplift_path, "w", encoding="utf-8") as f:
        f.write(f"Future promo weeks uplifted in the forecast (anchor {today_str})\n")
        f.write(f"Uplift mode: {uplift_mode}; halo: {PROMO_HALO_WEEKS} week(s) each side\n")
        f.write(f"Manual promo weeks: {', '.join(promo_weeks_resolved)}\n")
        f.write(f"Total week-rows uplifted: {len(uplift_log)}\n\n")
        if uplift_log:
            up_df = (
                pd.DataFrame(uplift_log)[
                    ["Customer Grouping", "SKU", "Description",
                     "Data Source", "WeekDate", "Uplift", "Baseline", "Projected"]
                ]
                .sort_values(["Customer Grouping", "SKU", "WeekDate"])
                .reset_index(drop=True)
            )
            f.write(up_df.to_string(index=False))
            f.write("\n")
        else:
            f.write(
                "No promo weeks fell inside the 15-week horizon for this snapshot "
                "(or PROMO_UPLIFT <= 1.0).\n"
            )
    print(f"Promo-uplift record saved to: {uplift_path} ({len(uplift_log)} week-rows)")
