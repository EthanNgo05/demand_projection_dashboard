"""Forecasting compute core: view enumeration, per-view/per-group forecasts."""
import os
import glob
import json
import inspect
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st

from dashboard_app.config import (
    ALL_CUSTOMERS_VIEW, region_from_view, MODEL_OPTIONS, MODEL_USED_COL,
    model_display, REPO_ROOT,
)
from dashboard_app.pipeline import (
    load_pipeline, pipeline_path,
    _supports_prices, _supports_smoothing, _supports_min_weeks, _supports_autofit,
)
from dashboard_app import forecast_cache


def _region_frame(df, P, region):
    """Rows of ``df`` whose customer group belongs to ``region``.

    str() on region_for_group: a custom pipeline may return non-string labels
    (see the key=str note in the sidebar), and the view string the region was
    parsed from was built from the str form.
    """
    groups = df["Customer Grouping"].map(lambda g: str(P.region_for_group(g)))
    return df[groups == region]


def list_views(df):
    """Group views organised by region, plus the combined ALL CUSTOMERS view."""
    P = load_pipeline(pipeline_path())
    groups = sorted(df["Customer Grouping"].dropna().unique().tolist())
    by_region = {}
    for g in groups:
        by_region.setdefault(P.region_for_group(g), []).append(g)
    return by_region


def _cache_key(data_sig, view, model_path, today_ts, alpha, beta, phi,
               min_weeks, kind):
    """Disk-cache key, or None when we must not use the disk cache.

    ``data_sig`` is threaded down from dashboard.py (which is the only layer that
    knows the snapshot/price file identities). It is an explicit argument rather
    than module state on purpose: Streamlit runs each browser session's script in
    its own thread against these same module objects, so a shared global could
    let one session's snapshot signature key another session's forecast. Absent a
    signature we simply skip the disk tier — the in-process ``@st.cache_data``
    layer still applies.
    """
    if not data_sig or not forecast_cache.enabled():
        return None
    return forecast_cache.forecast_key(
        data_sig, view, model_path, pd.Timestamp(today_ts).date().isoformat(),
        alpha=alpha, beta=beta, phi=phi, min_weeks=min_weeks, kind=kind,
    )


@st.cache_data(show_spinner="Building forecast…")
def compute_view(df, view, today_ts, model_path, prices=None, alpha=None,
                 beta=None, phi=None, min_weeks=None, data_sig=None):
    """Recompute summary + weekly + per-week aggregate for the selected view.

    Returns (summary_df, weekly_df, agg_frame) where agg_frame is the SKU-week
    POS/Orders/Projection table (used to draw historical actuals and the original
    projection). For ALL CUSTOMERS the breakdown is included so the summary
    carries 'Top Volume Customer Groups'. When ``prices`` (a SKU -> price Series)
    is supplied and the pipeline supports it, the summary also carries
    'List Price (USD)' and 'Revenue Risk (avg/wk)'. ``alpha`` / ``beta`` / ``phi``,
    when given, override the pipeline's smoothing constants for this call, and
    ``min_weeks`` overrides MIN_WEEKS_FOR_TREND (all are part of the cache key, so
    moving a slider recomputes the forecast). ``model_path`` selects the
    pipeline and keys the cache, so toggling the model recomputes too.

    Two cache tiers sit in front of the fit: this ``@st.cache_data`` decorator
    (per process, per session) and — when ``data_sig`` is supplied — the on-disk
    ``forecast_cache``, which survives restarts and is shared across sessions and
    warmed by the nightly ``agent.batch``. On a disk hit the fit is skipped
    entirely and the frames are read back from Parquet.
    """
    ck = _cache_key(data_sig, view, model_path, today_ts, alpha, beta, phi,
                    min_weeks, "view")
    if ck:
        hit = forecast_cache.get(ck)
        if hit is not None:
            return hit["summary"], hit["weekly"], hit["agg"]

    P = load_pipeline(model_path)
    kwargs = {}
    if prices is not None and _supports_prices(P):
        kwargs["list_prices"] = prices
    if None not in (alpha, beta, phi) and _supports_smoothing(P):
        kwargs.update(alpha=alpha, beta=beta, phi=phi)
    if min_weeks is not None and _supports_min_weeks(P):
        kwargs["min_weeks_for_trend"] = min_weeks
    if view == ALL_CUSTOMERS_VIEW:
        combined_label = getattr(
            P, "ALL_CUSTOMERS_LABEL", getattr(P, "ALL_SKUS_LABEL", ALL_CUSTOMERS_VIEW)
        )
        agg = P.aggregate_to_sku_week(df)
        summary, weekly = P.fit_regression(
            agg, today_ts, grouping_label=combined_label,
            breakdown_df=df, **kwargs,
        )
    elif (region_all := region_from_view(view)) is not None:
        # Per-region rollup: every customer group in the region, combined.
        # breakdown_df mirrors the ALL CUSTOMERS branch so the summary carries
        # 'Top Volume Customer Groups' (here: the region's groups).
        sub = _region_frame(df, P, region_all)
        agg = P.aggregate_to_sku_week(sub)
        summary, weekly = P.fit_regression(
            agg, today_ts, grouping_label=view, breakdown_df=sub, **kwargs
        )
    else:
        sub = df[df["Customer Grouping"] == view]
        agg = P.aggregate_to_sku_week(sub)
        summary, weekly = P.fit_regression(
            agg, today_ts, grouping_label=view, **kwargs
        )
    if ck:
        forecast_cache.put(
            ck, {"summary": summary, "weekly": weekly, "agg": agg},
            {"view": view, "model": os.path.basename(str(model_path)), "kind": "view"},
        )
    return summary, weekly, agg


def _format_sheet(ws, df, max_width=60, padding=2):
    """Make a written worksheet easy to read: bold+centered header row,
    AutoFilter over the data range, a frozen header row, and columns wide
    enough that the header and every cell value show in full (capped at
    ``max_width`` so one long text column can't blow out the sheet)."""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align

    # AutoFilter across the whole used range so users can sort/filter in Excel.
    ws.auto_filter.ref = ws.dimensions
    # Keep the header row visible while scrolling.
    ws.freeze_panes = "A2"

    for i, col in enumerate(df.columns, start=1):
        header_len = len(str(col))
        cell_len = df[col].map(lambda v: len(str(v))).max() if len(df) else 0
        width = min(max(header_len, int(cell_len)) + padding, max_width)
        ws.column_dimensions[get_column_letter(i)].width = width


@st.cache_data(show_spinner=False)
def view_to_excel(summary_df, weekly_df):
    """Build an in-memory .xlsx (same two-sheet layout as the pipeline output).

    Cached because ``st.download_button``'s ``data=`` argument is EAGER — it takes
    bytes, not a callable, so the workbook is rebuilt on every rerun whether or
    not anyone clicks. ``_format_sheet``'s per-column width scan is the expensive
    part (a Python ``len(str(v))`` over every cell), which made this ~0.6s of
    pure waste per widget interaction. Keyed on the frames themselves, so the
    bytes rebuild exactly when the table changes.
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        summary_df.to_excel(w, sheet_name="summary", index=False)
        weekly_df.to_excel(w, sheet_name="weekly_forecast", index=False)
        _format_sheet(w.sheets["summary"], summary_df)
        _format_sheet(w.sheets["weekly_forecast"], weekly_df)
    buf.seek(0)
    return buf.getvalue()


@st.cache_data(show_spinner=False)
def summary_to_excel(summary_df, sheet_name="summary"):
    """Build an in-memory single-sheet .xlsx of a summary table.

    Used for the by-SKU-and-customer table, which mirrors the pipeline's
    ALL_CUSTOMERS_demand_projections file (a single concatenated summary sheet).

    Cached for the same reason as ``view_to_excel``: every ``st.download_button``
    evaluates its ``data=`` eagerly, and the Exceptions view alone renders ~15 of
    them per rerun (two Under/Over sections × two tabs, the spikes table × two
    tabs, and four data-quality sections × two tabs).
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        summary_df.to_excel(w, sheet_name=sheet_name, index=False)
        _format_sheet(w.sheets[sheet_name], summary_df)
    buf.seek(0)
    return buf.getvalue()


@st.cache_data(show_spinner=False)
def run_autofit(df, view, today_ts, model_path, min_weeks=None, data_sig=None):
    """Grid-search the best alpha/beta/phi for the selected view (cached).

    Builds the same SKU-week aggregate ``compute_view`` fits on, then delegates
    to the pipeline's ``autofit_smoothing`` backtest. Cached on
    (data, view, snapshot, model, min_weeks) so clicking Autofit twice — or
    returning to a view already fitted this session — is instant.

    Also persisted to the on-disk cache when ``data_sig`` is given: the search is
    ~5s per view and Optimized Projections runs it for every group whose winning
    model supports it, so without persistence that cost returned on every
    restart. The result is a handful of floats, so it is stored as JSON rather
    than Parquet (``forecast_cache.get_params`` / ``put_params``).
    """
    ck = _cache_key(data_sig, view, model_path, today_ts, None, None, None,
                    min_weeks, "autofit")
    if ck:
        hit = forecast_cache.get_params(ck)
        if hit is not None:
            # An empty dict is how we record "this pipeline has no autofit" so the
            # miss isn't retried; map it back to the None the callers expect.
            return hit or None

    P = load_pipeline(model_path)
    if not _supports_autofit(P):
        if ck:
            forecast_cache.put_params(ck, {}, {"view": view, "kind": "autofit"})
        return None
    if view == ALL_CUSTOMERS_VIEW:
        agg = P.aggregate_to_sku_week(df)
    elif (region_all := region_from_view(view)) is not None:
        agg = P.aggregate_to_sku_week(_region_frame(df, P, region_all))
    else:
        agg = P.aggregate_to_sku_week(df[df["Customer Grouping"] == view])
    kwargs = {}
    if min_weeks is not None and "min_weeks_for_trend" in inspect.signature(
        P.autofit_smoothing
    ).parameters:
        kwargs["min_weeks_for_trend"] = min_weeks
    fitted = P.autofit_smoothing(agg, today_ts, **kwargs)
    if ck:
        forecast_cache.put_params(
            ck, fitted if fitted is not None else {},
            {"view": view, "model": os.path.basename(str(model_path)),
             "kind": "autofit"},
        )
    return fitted


@st.cache_data(show_spinner=False)
def _forecast_one_group(df_group, today_ts, model_path, group_label,
                        prices=None, alpha=None, beta=None, phi=None,
                        min_weeks=None, data_sig=None):
    """Forecast a single customer group's SKUs. Cached; calls NO Streamlit
    element, so it is safe to replay on a cache hit. ``group_label`` is a
    normal (hashable) argument so distinct groups get distinct cache entries.

    Returns ``(summary, weekly, agg)`` — the same three frames ``compute_view``
    produces for a single view, so callers that stitch groups together (the
    Optimal Projections combined view) can build charts, not just the summary.

    Backed by the same two tiers as ``compute_view``: ``@st.cache_data`` in
    process, plus the on-disk ``forecast_cache`` when ``data_sig`` is supplied.
    The disk tier matters most here — the by-customer loop runs this ~64 times,
    which is where Optimized Projections' ~55s went.
    """
    ck = _cache_key(data_sig, group_label, model_path, today_ts, alpha, beta, phi,
                    min_weeks, "group")
    if ck:
        hit = forecast_cache.get(ck)
        if hit is not None:
            return hit["summary"], hit["weekly"], hit["agg"]

    P = load_pipeline(model_path)
    kwargs = {}
    if prices is not None and _supports_prices(P):
        kwargs["list_prices"] = prices
    if None not in (alpha, beta, phi) and _supports_smoothing(P):
        kwargs.update(alpha=alpha, beta=beta, phi=phi)
    if min_weeks is not None and _supports_min_weeks(P):
        kwargs["min_weeks_for_trend"] = min_weeks
    agg = P.aggregate_to_sku_week(df_group)
    summary, weekly = P.fit_regression(
        agg, today_ts, grouping_label=group_label, **kwargs
    )
    if ck:
        forecast_cache.put(
            ck, {"summary": summary, "weekly": weekly, "agg": agg},
            {"view": group_label, "model": os.path.basename(str(model_path)),
             "kind": "group"},
        )
    return summary, weekly, agg


def _narrow_group_frames(weekly, agg, group):
    """The chart-ready ``(weekly, agg)`` pair for one group, tagged with it.

    Models carry extra per-model columns (exponential_smoothing adds
    ``promo_uplift``) which differ across models and would break the concat
    downstream, so both frames are narrowed to the columns the charts actually
    read. ``WeekDate`` is coerced because the models emit ``datetime.date``
    objects, not Timestamps. ``.assign(**{...})`` rather than a keyword because
    "Customer Grouping" is not an identifier.
    """
    wk = weekly[["SKU", "WeekDate", "projected_pos"]].copy()
    wk["WeekDate"] = pd.to_datetime(wk["WeekDate"])
    ag = agg[["SKU", "WeekDate", "POS", "Orders", "Projection"]].copy()
    ag["WeekDate"] = pd.to_datetime(ag["WeekDate"])
    return (wk.assign(**{"Customer Grouping": group}),
            ag.assign(**{"Customer Grouping": group}))


def _by_customer_frames(df, groups, forecast_group, progress_cb=None):
    """Stitch a per-group forecast loop into ``(combined, weekly_by_group,
    agg_by_group)``.

    The shared engine behind ``compute_by_customer``, ``compute_by_customer_frames``
    and ``compute_by_customer_best`` — they differ only in WHICH model fits each
    group, which is what ``forecast_group(group, sub)`` decides (it returns that
    group's ``(summary, weekly, agg)``, and may stamp extra columns on the
    summary). Groups that produce no rows are skipped silently.

    The two returned per-group frames keep their ``Customer Grouping`` column and
    are deliberately NOT summed: callers that want one total per (SKU, WeekDate)
    sum them themselves, and callers that draw one group at a time (the Customer
    detail charts, the per-row detail cards) need them un-summed. Returns
    ``(None, None, None)`` when nothing was forecast.
    """
    frames = []
    weekly_by_group_frames = []
    agg_by_group_frames = []
    n_groups = len(groups)
    for i, group in enumerate(groups):
        sub = df[df["Customer Grouping"] == group]
        summary, weekly, agg = forecast_group(group, sub)
        if summary is not None and not summary.empty:
            frames.append(summary)
            wk, ag = _narrow_group_frames(weekly, agg, group)
            weekly_by_group_frames.append(wk)
            agg_by_group_frames.append(ag)
        if progress_cb is not None:
            progress_cb(i + 1, n_groups, group)

    if not frames:
        return None, None, None
    return (pd.concat(frames, ignore_index=True),
            pd.concat(weekly_by_group_frames, ignore_index=True),
            pd.concat(agg_by_group_frames, ignore_index=True))


def compute_by_customer(df, today_ts, model_path, prices=None, alpha=None,
                        beta=None, phi=None, min_weeks=None, progress_cb=None,
                        data_sig=None):
    """Per-(SKU, Customer Grouping) summary — the rows behind ALL_CUSTOMERS.

    The pipeline's ``ALL_CUSTOMERS_demand_projections`` file is just a
    concatenation of every per-customer-group summary sheet. This reproduces it
    live: for each Customer Grouping we run the identical per-group forecast via
    the cached ``_forecast_one_group`` helper, then stack the summaries.
    Recomputing rather than reading the saved workbook keeps this table on the
    same snapshot / prices / smoothing as the rest of the page.

    This orchestrator is intentionally NOT cached: it may call ``progress_cb``
    (which drives a progress bar), and Streamlit element calls are not allowed
    inside a cached function. Each group's forecast is cached instead, so the
    expensive work is still memoised. On plain reruns this function isn't called
    at all — the result is held in session_state (see main()).

    Returns a DataFrame in the pipeline's SUMMARY_COLUMNS order, or None if no
    group had anything to forecast. Deliberately the RAW stitched summary: the
    descriptive averages the dashboard wants are attached by
    ``compute_by_customer_frames``, so this stays the plain concatenation the
    golden master pins.
    """
    return _by_customer_frames(
        df, sorted(df["Customer Grouping"].dropna().unique().tolist()),
        lambda group, sub: _forecast_one_group(
            sub, today_ts, model_path, group,
            prices, alpha, beta, phi, min_weeks, data_sig,
        ),
        progress_cb,
    )[0]


def _agent_summaries_mtime():
    """Newest mtime among outputs/agent_summary_*.json, or 0.0 if none exist.

    Folded into the combined view's cache signature so the table rebuilds
    automatically as soon as a batch (the "Agent Summary (all views)" button, the
    nightly job, or `agent.batch`) writes fresh summaries — no manual reload."""
    paths = glob.glob(os.path.join(REPO_ROOT, "outputs", "agent_summary_*.json"))
    return max((os.path.getmtime(p) for p in paths), default=0.0)


def _agent_summaries_generated_at():
    """Latest ``generated_at`` stamped across outputs/agent_summary_*.json.

    Reflects when the batch last produced the per-group recommendations that the
    Optimal Projections (Combined) view is stitched from. Returns the ISO string,
    or None if no summary carries a parseable timestamp. The stamps share one
    format (``YYYY-MM-DDTHH:MM:SS``), so a lexical max is also the newest."""
    latest = None
    for p in glob.glob(os.path.join(REPO_ROOT, "outputs", "agent_summary_*.json")):
        try:
            with open(p, encoding="utf-8") as f:
                gen = json.load(f).get("generated_at")
        except (OSError, ValueError):
            continue
        if gen and (latest is None or str(gen) > latest):
            latest = str(gen)
    return latest


def _agent_summaries_oldest_at():
    """Oldest ``generated_at`` across outputs/agent_summary_*.json, or None.

    The companion to _agent_summaries_generated_at (which returns the NEWEST).
    While a batch rewrites the summaries one view at a time, the newest stamp
    races ahead of the run even though most files are still from the PREVIOUS
    run — so "last generated <newest>" reads as if the whole set just refreshed
    when it hasn't. The OLDEST stamp is the honest "as of" for the table: every
    view in it is at least this fresh. Same one-pass parse; the stamps share the
    ISO ``YYYY-MM-DDTHH:MM:SS`` format, so a lexical min is the oldest."""
    oldest = None
    for p in glob.glob(os.path.join(REPO_ROOT, "outputs", "agent_summary_*.json")):
        try:
            with open(p, encoding="utf-8") as f:
                gen = json.load(f).get("generated_at")
        except (OSError, ValueError):
            continue
        if gen and (oldest is None or str(gen) < oldest):
            oldest = str(gen)
    return oldest


def _best_model_for_group(group):
    """(label, model_path) for a group's backtest-winning model, or None.

    Reads the group's published agent summary (agent_summary_<group>.json) and
    maps its ``best_model`` label to a MODEL_OPTIONS file path. Returns None when
    the summary is missing, has no best model, or names a label this deployment
    doesn't offer — the caller treats all three as "no summary yet".
    """
    payload = _load_agent_summary(group)
    if not payload:
        return None
    label = payload.get("best_model")
    path = MODEL_OPTIONS.get(label)
    if not label or path is None:
        return None
    return label, path


# The two descriptive-average column names, and the ONLY spellings the UI shows.
# The model files' AVG_COL_LABEL / DISPLAY_NAMES match these exactly so
# ``attach_descriptive_averages`` replaces the model's column in place instead of
# leaving two differently-named copies of one figure on the same table.
ALL_TIME_AVG_COL = "All-Time POS/Orders Average"
EIGHT_WK_AVG_COL = "8-Week POS/Orders Average"


@st.cache_data(show_spinner=False)
def _descriptive_averages(agg_by_group, today_ts):
    """Per-(Customer Grouping, SKU) all-time and 8-week demand averages.

    These are the CANONICAL figures behind ``ALL_TIME_AVG_COL`` /
    ``EIGHT_WK_AVG_COL`` everywhere in the UI — observed demand, model-independent,
    so one SKU/customer row reads the same whichever model produced it (see
    ``attach_descriptive_averages``).

    Cached because several callers ask for it with the same inputs on one pass
    through a view — ``compute_exceptions`` and ``compute_spikes`` pass the same
    ``sku_week_by_group`` frame and the same ``today_ts``, so only the first
    pays. (The per-group aggregate the forecast paths pass is a differently
    shaped frame, so those are separate entries; at ~0.25s each that is fine.)

    Computed straight from the stitched per-group SKU-week aggregates so BOTH
    averages exist for every group regardless of which model won its backtest.
    Definitions mirror the model files so the numbers agree:

    * source per SKU = POS if the SKU has ANY POS in the window, else Orders
      (SKUs with neither are skipped) -- the POS-then-Orders fallback the models use;
    * average = total demand / weeks-in-span, where the span runs from the SKU's
      first observation in the window through the last completed week (post-launch
      gaps count as real zeros). Over the 8-week window this reproduces
      ``regression.fit_regression``'s ``mean_val = y.sum() / weeks_since_first``.

    Note these are RAW observations: unlike the models' own reported average, no
    outlier cleansing is applied. Window, source fallback and span are otherwise
    identical to the models', so cleansing is the only difference between the two.

    Discontinued SKUs (name ending in '*') are dropped, matching the models.
    Returns a frame: Customer Grouping, SKU, ALL_TIME_AVG_COL, EIGHT_WK_AVG_COL.
    """
    # In-progress week is excluded; the historical window ends last completed week.
    days_since_sunday = (today_ts.weekday() + 1) % 7          # Sun=0 ... Sat=6
    current_week_start = today_ts - pd.Timedelta(days=days_since_sunday)
    last_complete_week = current_week_start - pd.Timedelta(weeks=1)
    eight_wk_start = last_complete_week - pd.Timedelta(weeks=7)   # 8 weeks inclusive
    hist_start = today_ts - pd.DateOffset(years=3)   # matches HISTORY_YEARS (all history)

    A = agg_by_group.copy()
    A["SKU"] = A["SKU"].astype(str)
    A = A[~A["SKU"].str.endswith("*")]
    A["WeekDate"] = pd.to_datetime(A["WeekDate"])

    key = ["Customer Grouping", "SKU"]

    def _avg(start, out_col):
        """Vectorised equivalent of the original per-(group, SKU) Python loop.

        The loop it replaces sliced the window frame once per (group, SKU) pair —
        ~11,000 slices, ~10s on the live snapshot — and this runs on the critical
        path of the Exceptions scan, the spikes scan and both combined-forecast
        views. Vectorised it is ~0.25s.

        Same three decisions, now as whole-column operations:
          * source = POS when the pair has ANY non-null POS in the window, else
            Orders; a pair with neither is dropped (``keep`` below);
          * the span starts at the first week of the CHOSEN source, not of the
            pair, which is why the two sources are aggregated separately rather
            than coalesced first;
          * output row order is first-appearance order, matching the original's
            ``groupby(..., sort=False)`` iteration.

        The final ``round`` stays Python's, applied per value, rather than
        ``np.round``: the two use different algorithms (correctly-rounded decimal
        vs. scale-rint-unscale) and disagree on some halfway values, and these
        numbers are displayed. ~11k cheap C-level calls, not 11k frame slices.
        ``tests/test_perf_parity.py::test_descriptive_averages_golden`` holds the
        whole thing to exact equality with the original output.
        """
        win = A[(A["WeekDate"] >= start) & (A["WeekDate"] <= last_complete_week)]
        if win.empty:
            return pd.DataFrame(columns=key + [out_col])

        # First-appearance order of the pairs, so row order is unchanged.
        idx = pd.MultiIndex.from_frame(win[key].drop_duplicates())

        def _agg(col):
            """(sum, first week) per pair over the rows where ``col`` is present.

            The empty case must keep its dtypes: a plain ``pd.Series()`` defaults
            to float64, and reindexing that in place of the week-start column
            would leave np.where trying to promote float64 against datetime64
            (which has no common dtype) whenever a window happens to contain no
            POS rows at all, or no Orders rows at all.
            """
            src = win[win[col].notna()]
            if src.empty:
                empty_idx = pd.MultiIndex.from_arrays([[], []], names=key)
                return (pd.Series(dtype="float64", index=empty_idx),
                        pd.Series(dtype=win["WeekDate"].dtype, index=empty_idx))
            g = src.groupby(key, sort=False)
            return g[col].sum(), g["WeekDate"].min()

        pos_sum, pos_first = _agg("POS")
        ord_sum, ord_first = _agg("Orders")

        has_pos = idx.isin(pos_sum.index)
        keep = has_pos | idx.isin(ord_sum.index)
        if not keep.any():
            return pd.DataFrame(columns=key + [out_col])

        totals = np.where(has_pos, pos_sum.reindex(idx).to_numpy(),
                          ord_sum.reindex(idx).to_numpy())
        firsts = np.where(has_pos, pos_first.reindex(idx).to_numpy(),
                          ord_first.reindex(idx).to_numpy())

        # Floor-divide to whole days, mirroring Timedelta.days (both bounds are
        # week-start dates, and the window guarantees first <= last_complete_week,
        # so this is exact rather than a truncation).
        days = (last_complete_week.to_datetime64() - firsts[keep]) \
            // np.timedelta64(1, "D")
        spans = np.round(days / 7).astype("int64") + 1
        np.maximum(spans, 1, out=spans)

        kept = idx[keep]
        return pd.DataFrame({
            "Customer Grouping": kept.get_level_values(0),
            "SKU": kept.get_level_values(1),
            out_col: [round(v, 1) for v in totals[keep] / spans],
        })

    all_time = _avg(hist_start, ALL_TIME_AVG_COL)
    eight_wk = _avg(eight_wk_start, EIGHT_WK_AVG_COL)
    return all_time.merge(eight_wk, on=["Customer Grouping", "SKU"], how="outer")


def attach_descriptive_averages(summary, agg_by_group, today_ts):
    """Give every row of ``summary`` BOTH descriptive averages, centrally computed.

    Each model reports only ONE average, and reports it as the mean of the series
    it actually fit — which for four of the five is outlier-CLEANSED (promo spikes
    flattened, stockout dips lifted). Read off a table that figure is indefensible:
    it silently changes definition when the model changes, and it doesn't match the
    demand a planner can see on the chart next to it.

    So both columns come from ``_descriptive_averages`` — observed demand over the
    same window and span the models use, no cleansing — and the central value WINS
    over whatever the model reported. The model's own figure survives only where
    ``_descriptive_averages`` has no row at all (see below). That is what makes one
    number mean one thing across every table, card and Excel export in the UI, and
    what lets a recent run-rate sit beside a long-run average and be comparable.

    (Each model's standalone .xlsx output is untouched — it still reports the
    cleansed mean it fit on, which is the right figure in that context. So is
    ``compute_view``'s summary, which the agent parity tests hold to exact equality
    with ``fit_regression``; the one place that surfaces is the collapsed
    "Summary table by SKU (view total)", whose column dashboard.py suffixes
    "(model fit)" so it can't be read as this one.)

    Merges on ``["SKU", "Customer Grouping"]``, so ``summary["SKU"]`` must already
    be string-typed (``_descriptive_averages`` casts its own side). Discontinued
    ``*`` SKUs are dropped by ``_descriptive_averages``, so they keep the model's
    value if it reported one and otherwise come back with a blank All-Time and a
    0.0 8-Week — harmless, since the models drop them too.

    Stops deliberately after slotting the two columns in: callers that reorder
    other columns (``compute_by_customer_best`` moves ``Model Used``) depend on
    this ordering being final.
    """
    avgs = _descriptive_averages(agg_by_group, today_ts)
    # Legacy spelling from before the model files were aligned on EIGHT_WK_AVG_COL.
    # Harmless now, kept so a frame persisted by an older build still loads clean.
    summary = summary.drop(columns=["8 Week POS/Orders Average"], errors="ignore")
    summary = summary.merge(
        avgs.rename(columns={
            ALL_TIME_AVG_COL: "_central_all_time",
            EIGHT_WK_AVG_COL: "_central_8wk",
        }),
        on=["SKU", "Customer Grouping"], how="left",
    )
    # All-Time: the central observed value wins; the model's own (cleansed) figure
    # only fills pairs _descriptive_averages produced no row for, so such a row
    # shows *something* rather than a blank.
    summary[ALL_TIME_AVG_COL] = (
        summary["_central_all_time"].fillna(summary[ALL_TIME_AVG_COL])
        if ALL_TIME_AVG_COL in summary.columns
        else summary["_central_all_time"]
    )
    # 8-Week: the central value everywhere. A SKU with history but no POS/Orders in
    # the last 8 weeks has no run-rate to compute; its recent average is a genuine 0
    # (absent week = zero, matching the models' gap-fill), so fill rather than leave
    # a blank.
    summary[EIGHT_WK_AVG_COL] = summary["_central_8wk"].fillna(0.0)
    summary = summary.drop(columns=["_central_all_time", "_central_8wk"])

    # Slot both averages right after "Weeks with data" (All-Time then 8-Week),
    # immediately ahead of "Updated Projection Average", for a stable layout.
    if "Weeks with data" in summary.columns:
        cols = [c for c in summary.columns
                if c not in (ALL_TIME_AVG_COL, EIGHT_WK_AVG_COL)]
        pos = cols.index("Weeks with data") + 1
        cols[pos:pos] = [ALL_TIME_AVG_COL, EIGHT_WK_AVG_COL]
        summary = summary[cols]
    return summary


def compute_by_customer_frames(df, today_ts, model_path, prices=None, alpha=None,
                               beta=None, phi=None, min_weeks=None,
                               progress_cb=None, data_sig=None):
    """``compute_by_customer`` plus everything the Quick Projections view needs.

    Returns ``(combined, weekly_by_group, agg_by_group)``: the same per-(SKU,
    Customer Grouping) summary, with BOTH descriptive averages attached, alongside
    the un-summed per-group weekly-forecast and SKU-week frames that the Customer
    detail chart and the summary table's per-row detail cards draw from.

    Not cached, for the same reason ``compute_by_customer`` isn't: it drives a
    ``progress_cb``, and Streamlit element calls can't happen inside a cached
    function. The expensive part — each group's fit — is cached in
    ``_forecast_one_group`` (in process and on disk), so a repeat call pays only
    the stitching.
    """
    combined, weekly_by_group, agg_by_group = _by_customer_frames(
        df, sorted(df["Customer Grouping"].dropna().unique().tolist()),
        lambda group, sub: _forecast_one_group(
            sub, today_ts, model_path, group,
            prices, alpha, beta, phi, min_weeks, data_sig,
        ),
        progress_cb,
    )
    if combined is None:
        return None, None, None
    return (attach_descriptive_averages(combined, agg_by_group, today_ts),
            weekly_by_group, agg_by_group)


def single_group_frames(summary, weekly, agg, group, today_ts):
    """``compute_by_customer_frames``' 3-tuple for a view that IS one group.

    ``compute_view``'s single-group branch has already made exactly the calls the
    per-group loop would make — same slice, same ``aggregate_to_sku_week``, same
    ``fit_regression`` with the same ``grouping_label`` — so re-entering the loop
    would fit the group a second time purely because ``compute_view`` and
    ``_forecast_one_group`` are different cache entries (and different disk-cache
    kinds, "view" vs "group"). Reuse the frames instead.
    """
    wk, ag = _narrow_group_frames(weekly, agg, group)
    return attach_descriptive_averages(summary, ag, today_ts), wk, ag


def compute_by_customer_best(df, today_ts, prices=None, min_weeks=None,
                             progress_cb=None, data_sig=None):
    """Per-(SKU, Customer Grouping) summary using each group's BEST model.

    Like ``compute_by_customer``, but instead of one model for every group it
    forecasts each group with the model that won that group's backtest (from
    ``agent_summary_<group>.json``) and stamps a ``MODEL_USED_COL`` column. To
    match what the single-group view shows, groups whose best model supports
    autofit are tuned per group via ``run_autofit`` before forecasting.

    A group is only included if it has a resolvable best model. Groups with no
    published summary, or whose summary has no backtest winner (``best_model`` is
    null — history too short to score any model), are left OUT of the table and
    returned separately so the caller can list them.

    Returns ``(table, weekly_all, agg_all, weekly_by_group, agg_by_group,
    excluded)`` where ``table`` is a DataFrame (SUMMARY_COLUMNS + MODEL_USED_COL)
    or None when no group resolved / produced rows; ``weekly_all`` / ``agg_all``
    are the per-group forecast and SKU-week aggregate frames stitched together and
    summed by (SKU, WeekDate) so the view can draw the total-demand and per-SKU
    charts; ``weekly_by_group`` / ``agg_by_group`` are the SAME per-group frames
    stitched together but NOT summed — each row keeps its ``Customer Grouping`` so
    the view can draw one customer group's total on demand (all four frames are
    None alongside a None table); and ``excluded`` is the sorted list of group
    names with no best model. Groups are disjoint customer subsets, so summing by
    (SKU, WeekDate) is a plain total — no double counting — and the actuals match
    the Executive Overview.
    """
    groups = sorted(df["Customer Grouping"].dropna().unique().tolist())

    # First pass: split into groups with a resolvable best model vs. those without
    # (no summary file, or a summary whose best_model is null).
    resolved = {}
    excluded = []
    for group in groups:
        best = _best_model_for_group(group)
        if best is None:
            excluded.append(group)
        else:
            resolved[group] = best
    if not resolved:
        return None, None, None, None, None, excluded

    # Second pass: forecast each resolved group with its own model (autofit when
    # supported), stamping the winning model's name on each row.
    def _forecast_best(group, sub):
        label, path = resolved[group]
        alpha = beta = phi = None
        P = load_pipeline(path)
        if _supports_autofit(P):
            # Pass the group slice, not the whole frame: run_autofit's non-region
            # branch filters by `Customer Grouping == view`, and re-filtering an
            # already-filtered slice by the same predicate is a no-op — so the
            # aggregate it fits is byte-identical while the full-frame filter and
            # aggregate stop being repeated here (they were already done by the
            # loop, and again inside _forecast_one_group). It also narrows the
            # cache key from "the whole snapshot" to "this group's rows", which is
            # what the result actually depends on.
            fitted = run_autofit(sub, group, today_ts, path, min_weeks, data_sig)
            if fitted:
                alpha, beta, phi = fitted.get("alpha"), fitted.get("beta"), fitted.get("phi")
        summary, weekly, agg = _forecast_one_group(
            sub, today_ts, path, group, prices, alpha, beta, phi, min_weeks,
            data_sig,
        )
        if summary is not None and not summary.empty:
            # Copy before stamping: _forecast_one_group's result is cached and
            # shared, so mutating it in place would leak MODEL_USED_COL into every
            # other caller's copy.
            summary = summary.copy()
            summary[MODEL_USED_COL] = model_display(label)
        return summary, weekly, agg

    combined, weekly_by_group, agg_by_group = _by_customer_frames(
        df, list(resolved), _forecast_best, progress_cb
    )
    if combined is None:
        return None, None, None, None, None, excluded

    # Give every group BOTH descriptive averages regardless of its winning model.
    combined = attach_descriptive_averages(combined, agg_by_group, today_ts)

    # Surface the model used right after the customer group for readability.
    if "Customer Grouping" in combined.columns:
        cols = [c for c in combined.columns if c != MODEL_USED_COL]
        pos = cols.index("Customer Grouping") + 1
        cols.insert(pos, MODEL_USED_COL)
        combined = combined[cols]

    # Stitch the per-group series into one total per (SKU, WeekDate). min_count=1
    # keeps a genuinely-absent cell NaN rather than coercing it to 0. Naming the
    # value columns explicitly is what drops the "Customer Grouping" the per-group
    # frames carry, so these two come out at (SKU, WeekDate) grain.
    weekly_all = (
        weekly_by_group
        .groupby(["SKU", "WeekDate"], as_index=False)["projected_pos"].sum()
    )
    agg_all = (
        agg_by_group
        .groupby(["SKU", "WeekDate"], as_index=False)[["POS", "Orders", "Projection"]]
        .sum(min_count=1)
    )
    return combined, weekly_all, agg_all, weekly_by_group, agg_by_group, excluded


def _live_best_model(df, group, today_ts, prices=None):
    """Run the 5-model backtest live for ONE group and return its
    ``(label, model_path)``, or None if no model was scoreable (history too short).

    Used when the group has no published ``agent_summary_<group>.json`` yet. Calls
    the same three agent nodes the graph uses (skipping ingest / LLM / publish) on a
    hand-built state; ``run_all_models`` slices the group via ``view_frame``. Imported
    lazily so the dashboard doesn't pull the agent/LangGraph stack unless this runs.
    """
    from agent.nodes.forecast import run_all_models
    from agent.nodes.evaluate import evaluate_models
    from agent.nodes.select import select_best_model

    state = {"cleaned_df": df, "view": group, "today_ts": today_ts,
             "prices": prices, "errors": []}
    state.update(run_all_models(state))
    state.update(evaluate_models(state))
    state.update(select_best_model(state))
    label = state.get("best_model")
    path = MODEL_OPTIONS.get(label) if label else None
    if not label or path is None:
        return None
    return label, path


def optimal_projection_for(df, group, sku, today_ts, prices=None, min_weeks=None,
                           data_sig=None):
    """Optimized (best-model) 15-week forecast for ONE (SKU, Customer Grouping).

    Reuses the Optimized Projections view's machinery so a group already forecast
    this session is an instant cache hit: resolve the group's backtest winner from
    ``agent_summary_<group>.json`` (else run the 5-model backtest live for the group),
    then forecast that winner through the SAME cached ``run_autofit`` /
    ``_forecast_one_group`` calls ``compute_by_customer_best`` uses, and slice the SKU.

    Returns a dict:
      * ``{"status": "ok", "label", "weekly", "optimized_avg"}`` — ``weekly`` is the
        SKU's ``SKU/WeekDate/projected_pos`` frame; ``optimized_avg`` its 15-week mean.
      * ``{"status": "no_model"}`` — no model could be backtested for the group.
      * ``{"status": "no_data", "label"}`` — winner produced no forecast for this SKU.
    """
    best = _best_model_for_group(group)
    if best is None:
        best = _live_best_model(df, group, today_ts, prices)
        if best is None:
            return {"status": "no_model"}
    label, path = best

    sub = df[df["Customer Grouping"] == group]
    alpha = beta = phi = None
    P = load_pipeline(path)
    if _supports_autofit(P):
        # Group slice, not the whole frame — see the note in
        # compute_by_customer_best: same aggregate, narrower cache key, and it
        # shares the entry that view already populated for this group.
        fitted = run_autofit(sub, group, today_ts, path, min_weeks, data_sig)
        if fitted:
            alpha, beta, phi = fitted.get("alpha"), fitted.get("beta"), fitted.get("phi")
    _, weekly, _ = _forecast_one_group(
        sub, today_ts, path, group, prices, alpha, beta, phi, min_weeks, data_sig
    )
    if weekly is None or weekly.empty:
        return {"status": "no_data", "label": label}
    wk = weekly[weekly["SKU"].astype(str) == str(sku)][
        ["SKU", "WeekDate", "projected_pos"]
    ].copy()
    if wk.empty:
        return {"status": "no_data", "label": label}
    wk["WeekDate"] = pd.to_datetime(wk["WeekDate"])
    return {"status": "ok", "label": label, "weekly": wk,
            "optimized_avg": float(wk["projected_pos"].mean())}


def _agent_summary_path(view):
    """Path publish.py writes for a given view (same view->filename mangling)."""
    safe_view = view.replace(" ", "_").replace("/", "-")
    return os.path.join(REPO_ROOT, "outputs", f"agent_summary_{safe_view}.json")


def _load_agent_summary(view):
    """Last agent run for this view, or None if it hasn't run / is unreadable."""
    path = _agent_summary_path(view)
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return None
